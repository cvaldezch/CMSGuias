#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
import datetime
import json

from django.shortcuts import (
                              render_to_response,
                              get_object_or_404,
                              render,
                              get_list_or_404)
from django.template import RequestContext, TemplateDoesNotExist
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse, HttpResponseRedirect, Http404
from django.contrib import messages
# from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.db.models import Count, Max, Sum, Q
from django.utils import simplejson
from django.core.paginator import EmptyPage, Paginator, PageNotAnInteger
from django.views.generic import TemplateView, ListView, View
from django.core import serializers

from openpyxl import load_workbook

from CMSGuias.apps.almacen.models import *
from CMSGuias.apps.home.models import *
from CMSGuias.apps.ventas.models import Proyecto, Sectore, Subproyecto
from CMSGuias.apps.almacen import forms
from CMSGuias.apps.tools import genkeys, globalVariable, uploadFiles
from CMSGuias.apps.logistica.models import Compra, DetCompra
from CMSGuias.apps.operations.models import *


##
#  Declare variables
##

class JSONResponseMixin(object):
    def render_to_json_response(self, context, **response_kwargs):
        return HttpResponse(
            self.convert_context_to_json(context),
            content_type='application/json',
            mimetype='application/json',
            **response_kwargs
        )

    def convert_context_to_json(self, context):
        return simplejson.dumps(context, encoding='utf-8')

FORMAT_DATE_STR = '%Y-%m-%d'

class StorageHome(View):
    template_name = 'almacen/storage.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        return render_to_response(
            self.template_name, context_instance=RequestContext(request))

@login_required(login_url='/SignUp/')
def view_pedido(request):
    try:
        if request.method == 'GET':
            return render_to_response(
                                    'almacen/pedido.html',
                                    context_instance=RequestContext(request))
        if request.method == 'POST':
            data = {}
            form = forms.addOrdersForm(request.POST, request.FILES)
            if form.is_valid():
                # bedside Orders
                add = form.save(commit=False)
                id = genkeys.GenerateIdOrders()
                add.pedido_id = id
                add.status = 'PE'
                add.flag = True
                add.save()
                # detail Orders Details
                tmpd = tmppedido.objects.filter(
                        empdni__exact=request.user.get_profile().empdni_id)
                for x in tmpd:
                    det = Detpedido()
                    det.pedido_id = id
                    det.materiales_id = x.materiales_id
                    det.cantidad = x.cantidad
                    det.cantshop = x.cantidad
                    det.brand_id = x.brand_id
                    det.model_id = x.model_id
                    det.save()
                # saved niples of tmpniple
                tmpn = tmpniple.objects.filter(
                        empdni__exact=request.user.get_profile().empdni_id)
                for x in tmpn:
                    nip = Niple()
                    nip.pedido_id = id
                    nip.proyecto_id = request.POST.get('proyecto')
                    nip.subproyecto_id = request.POST.get('subproyecto')
                    nip.sector_id = request.POST.get('sector')
                    nip.empdni = request.user.get_profile().empdni_id
                    nip.materiales_id = x.materiales_id
                    nip.cantidad = x.cantidad
                    nip.cantshop = x.cantidad
                    nip.metrado = x.metrado
                    nip.tipo = x.tipo.strip()
                    nip.save()
                # deleting tmps
                tmp = tmppedido.objects.filter(
                        empdni__exact=request.user.get_profile().empdni_id)
                tmp.delete()
                tmp = tmpniple.objects.filter(
                        empdni__exact=request.user.get_profile().empdni_id)
                tmp.delete()
                data['status'] = True
            else:
                data['status'] = False
            return HttpResponse(
                    simplejson.dumps(data),
                    mimetype='application/json')
    except TemplateDoesNotExist, e:
        raise Http404(e)

@login_required(login_url='/SignUp/')
def view_keep_customers(request):
    try:
        if request.method == 'GET':
            lista = Cliente.objects.values(
                        'ruccliente_id',
                        'razonsocial',
                        'direccion',
                        'telefono').filter(flag=True).order_by('razonsocial')
            ctx = {'lista': lista}
            return render_to_response(
                'almacen/customers.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            if 'ruc' in request.POST:
                data = {}
                try:
                    obj = Cliente.objects.get(pk=request.POST.get('ruc'))
                    obj.flag = False
                    obj.save()
                    data['status'] = True
                except ObjectDoesNotExist, e:
                    data['status'] = False
                return HttpResponse(simplejson.dumps(data),
                                    mimetype='application/json')
    except TemplateDoesNotExist, e:
        messages.error(request, str(e))
        raise Http404('Method no proccess')

@login_required(login_url='/SignUp/')
def view_keep_add_customers(request):
    try:
        info = 'Iniciando'
        if request.method == 'POST':
            countc = Cliente.objects.filter(
                        pk=request.POST.get('ruccliente_id')).count()
            if countc > 0:
                add = Cliente.objects.get(pk=request.POST.get('ruccliente_id'))
                add.razonsocial = request.POST.get('razonsocial')
                add.pais_id = request.POST.get('pais')
                add.departamento_id = request.POST.get('departamento')
                add.provincia_id = request.POST.get('provincia')
                add.distrito_id = request.POST.get('distrito')
                add.direccion = request.POST.get('direccion')
                add.telefono = request.POST.get('telefono')
                add.flag = True
                add.save()
            else:
                form = forms.addCustomersForm(request.POST)
                if form.is_valid():
                    add = form.save(commit=False)
                    add.flag = True
                    add.save()
            # form.save_m2m() # esto es para guardar las relaciones ManyToMany
            return HttpResponseRedirect('/almacen/keep/customers/')
        if request.method == 'GET':
            form = forms.addCustomersForm()
        ctx = {'form': form, 'info': info}
        return render_to_response(
            'almacen/addcustomers.html',
            ctx,
            context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages.error(request, str(e))
        raise Http404('Method no proccess')

@login_required(login_url='/SignUp/')
def view_keep_edit_customers(request, ruc):
    try:
        c = Cliente.objects.get(pk__exact=ruc)
        if request.method == 'GET':
            form = forms.addCustomersForm(instance=c)
        elif request.method == 'POST':
            form = forms.addCustomersForm(request.POST, instance=c)
            if form.is_valid():
                edit = form.save(commit=False)
                edit.flag = True
                edit.save()
                return HttpResponseRedirect('/almacen/keep/customers/')
        ctx = {'form': form}
        return render_to_response(
            'almacen/editcustomers.html',
            ctx,
            context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages.error(request, str(e))
        raise Http404('Method no proccess')

# Project keep views
@login_required(login_url='/SignUp/')
def view_keep_project(request):
    try:
        if request.method == 'GET':
            lista = Proyecto.objects.filter(flag=True).order_by('nompro')
            return render_to_response(
                'almacen/project.html',
                {'lista': lista, 'tsid': u''},
                context_instance=RequestContext(request))

        if request.method == 'POST':
            if request.is_ajax():
                if 'proid' in request.POST:
                    data = {}
                    try:
                        # sectores
                        obj = Sectore.objects.filter(
                            proyecto_id=request.POST.get('proid'))
                        if obj:
                            for x in obj:
                                x.status = 'DL'
                                x.flag = False
                                x.save()
                        obj = Subproyecto.objects.filter(
                            proyecto_id=request.POST.get('proid'))
                        if obj:
                            for x in obj:
                                obj.status = 'DL'
                                obj.flag = False
                                obj.save()
                        obj = Proyecto.objects.get(
                            pk=request.POST.get('proid'))
                        obj.status = 'DL'
                        obj.flag = False
                        obj.save()
                        data['status'] = True
                    except ObjectDoesNotExist, e:
                        data['status'] = False
                        data['raise'] = e.__str__()
                    return HttpResponse(
                        simplejson.dumps(data),
                        mimetype='application/json')
    except TemplateDoesNotExist, e:
        messages.error(request, str(e))
        raise Http404('Method no proccess')

@login_required(login_url='/SignUp/')
def view_keep_add_project(request):
    try:
        if request.method == 'POST':
            project = Proyecto.objects.filter(
                        pk=request.POST.get('proyecto_id'))
            if project.count() > 0:
                add = project
                add.ruccliente_id = request.POST.get('ruccliente')
                add.nompro = request.POST.get('nompro')
                add.comienzo = datetime.datetime.strptime(
                    request.POST.get('comienzo'),
                    FORMAT_DATE_STR).date() if request.POST.get(
                        'comienzo') is not None else datetime.datetime.today(
                            ).date()
                add.fin = datetime.datetime.strptime(
                    request.POST.get('fin'),
                    FORMAT_DATE_STR).date() if request.POST.get(
                        'fin') is not None else None
                add.pais_id = request.POST.get('pais')
                add.departamento_id = request.POST.get('departamento')
                add.provincia_id = request.POST.get('provincia')
                add.distrito_id = request.POST.get('distrito')
                add.direccion = request.POST.get('direccion')
                add.obser = request.POST.get('obser')
                add.status = request.POST.get('status')
                add.contact = request.POST.get('contact')
                add.flag = True
                add.save()
            else:
                form = forms.addProjectForm(request.POST)
                if form.is_valid():
                    add = form.save(commit=False)
                    # for table project have generate id
                    cod = Proyecto.objects.all().aggregate(Max('proyecto_id'))
                    if cod['proyecto_id__max'] is not None:
                        aa = cod['proyecto_id__max'][2:4]
                        an = datetime.datetime.today().strftime('%y')
                        cou = cod['proyecto_id__max'][4:7]
                        if int(an) > int(aa):
                            aa = an
                            cou = 1
                        else:
                            cou = (int(cou) + 1)
                    else:
                        aa = datetime.datetime.today().strftime('%y')
                        cou = 1
                    cod = '%s%s%s' % ('PR', str(aa), '{:0>3d}'.format(cou))
                    add.proyecto_id = cod.strip()
                    add.flag = True
                    add.save()
            return HttpResponseRedirect('/almacen/keep/project/')
        if request.method == 'GET':
            form = forms.addProjectForm()
        ctx = {'form': form}
        return render_to_response(
            'almacen/addproject.html',
            ctx,
            context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages.error(request, str(e))
        raise Http404('Method no proccess')

@login_required(login_url='/SignUp/')
def view_keep_edit_project(request, proid):
    try:
        c = Proyecto.objects.get(pk__exact=proid)
        if request.method == 'GET':
            form = forms.addProjectForm(instance=c)
        elif request.method == 'POST':
            # print request.POST.get('proyecto_id')
            form = forms.addProjectForm(request.POST, instance=c)
            if form.is_valid():
                edit = form.save(commit=False)
                edit.flag = True
                # edit.proyecto_id = request.POST.get('proyecto_id')
                edit.save()
                return HttpResponseRedirect('/almacen/keep/project/')
        ctx = {'form': form, 'idpro': proid}
        return render_to_response(
            'almacen/editproject.html',
            ctx,
            context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages.error(request, str(e))
        raise Http404('Method no proccess')

# Sectors keep views
@login_required(login_url='/SignUp/')
def view_keep_sec_project(request, pid, sid):
    try:
        if request.method == 'GET':
            lista = Sectore.objects.filter(
                flag=True,
                proyecto__flag=True,
                proyecto_id__exact=pid,
                subproyecto_id=None if sid.strip() == '' else sid).order_by(
                    'sector_id')
            return render_to_response(
                'almacen/sectores.html',
                {'lista': lista, 'pid': pid, 'sid': sid},
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            if 'sec' in request.POST:
                data = {}
                try:
                    obj = Sectore.objects.get(
                        proyecto_id=pid,
                        subproyecto_id=None if sid.strip() == '' else sid,
                        sector_id=request.POST.get('sec'))
                    obj.delete()
                    data['status'] = True
                except ObjectDoesNotExist, e:
                    data['status'] = False
                return HttpResponse(
                    simplejson.dumps(data),
                    mimetype='application/json')
    except TemplateDoesNotExist, e:
        messages.error(request, str(e))
        raise Http404('Method no proccess')

@login_required(login_url='/SignUp/')
def view_keep_add_sector(request, proid, sid):
    try:
        if request.method == 'POST':
            form = forms.addSectoresForm(request.POST)
            if form.is_valid():
                add = form.save(commit=False)
                add.proyecto_id = request.POST.get('proyecto_id')
                add.subproyecto_id = request.POST.get('subproyecto_id')
                add.flag = True
                add.save()
                url = '/almacen/keep/sectores/%s/%s/' % (proid, sid)
                return HttpResponseRedirect(url)
            else:
                form = forms.addSectoresForm(request.POST)
                msg = 'No se a podido realizar la transacción.'
                ctx = {'form': form, 'pid': proid, 'sid': sid, 'msg': msg}
                return render_to_response(
                    'almacen/addsector.html',
                    ctx,
                    context_instance=RequestContext(request))
        if request.method == 'GET':
            form = forms.addSectoresForm()
            ctx = {'form': form, 'pid': proid, 'sid': sid}
            return render_to_response(
                'almacen/addsector.html',
                ctx,
                context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages.error(request, str(e))
        raise Http404('Method no proccess')

@login_required(login_url='/SignUp')
def view_keep_edit_sector(request, pid, sid, cid):
    try:
        sec = Sectore.objects.get(
            proyecto_id=pid,
            subproyecto_id=None if sid == '' else sid,
            sector_id=cid)
        if request.method == 'GET':
            form = forms.addSectoresForm(instance=sec)
            ctx = {'form': form, 'pid': pid, 'sid': sid, 'cid': cid}
            return render_to_response(
                'almacen/editsector.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            form = forms.addSectoresForm(request.POST, instance=sec)
            if form.is_valid():
                edit = form.save(commit=False)
                edit.proyecto_id = request.POST.get('proyecto_id')
                edit.subproyecto_id = request.POST.get('subproyecto_id')
                edit.flag = True
                edit.save()
                url = '/almacen/keep/sectores/%s/%s/' % (pid, sid)
                return HttpResponseRedirect(url)
            else:
                form = forms.addSectoresForm(request.POST)
                msg = 'No se a podido realizar la transacción.'
                ctx = {'form': form, 'pid': proid, 'sid': sid, 'msg': msg}
                return render_to_response(
                    'almacen/addsector.html',
                    ctx,
                    context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages.error(request, 'No se puede mostrar la pagina.' + str(e))
        raise Http404('Method not proccess')

# Subproyectos keep views
@login_required(login_url='/SignUp/')
def view_keep_sub_project(request, pid):
    try:
        if request.method == 'GET':
            lista = Subproyecto.objects.filter(
                flag=True,
                proyecto__flag=True,
                proyecto_id__exact=pid).order_by('subproyecto_id')
            return render_to_response(
                'almacen/subproject.html',
                {'lista': lista, 'pid': pid, 'sid': ''},
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            if 'sid' in request.POST:
                data = {}
                try:
                    obj = Sectore.objects.filter(
                            proyecto_id=pid,
                            subproyecto_id=request.POST.get('sid'))
                    obj.delete()
                    obj = Subproyecto.objects.get(
                            subproyecto_id=request.POST.get('sid'),
                            proyecto_id=pid)
                    obj.delete()
                    data['status'] = True
                except ObjectDoesNotExist, e:
                    data['status'] = False
                return HttpResponse(
                        simplejson.dumps(data),
                        mimetype='application/json')
    except TemplateDoesNotExist, e:
        messages.error(request, str(e))
        raise Http404('Method no proccess')

@login_required(login_url='/SignUp/')
def view_keep_add_subproyeto(request, pid):
    try:
        if request.method == 'GET':
            form = forms.addSubprojectForm()
            ctx = {'form': form, 'pid': pid}
            return render_to_response(
                'almacen/addsubproyecto.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            form = forms.addSubprojectForm(request.POST)
            if form.is_valid():
                add = form.save(commit=False)
                add.proyecto_id = request.POST.get('proyecto_id')
                add.flag = True
                add.save()
                url = '/almacen/keep/subproyectos/%s/' % (pid)
                return HttpResponseRedirect(url)
            else:
                print 'Form no valid'
    except TemplateDoesNotExist:
        # messages.error(request, str(e))
        raise Http404('Method no proccess')

@login_required(login_url='/SignUp')
def view_keep_edit_subproyecto(request, pid, sid):
    try:
        sub = Subproyecto.objects.get(
                flag=True,
                proyecto__flag=True,
                proyecto_id__exact=pid,
                subproyecto_id=sid)
        if request.method == 'GET':
            form = forms.addSubprojectForm(instance=sub)
            ctx = {'form': form, 'pid': pid}
            return render_to_response(
                'almacen/editsubproyecto.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            form = forms.addSubprojectForm(request.POST, instance=sub)
            if form.is_valid():
                edit = form.save(commit=False)
                edit.proyecto_id = request.POST.get('proyecto_id')
                edit.flag = True
                edit.save()
                url = '/almacen/keep/subproyectos/%s/' % (pid)
                return HttpResponseRedirect(url)
            else:
                form = forms.addSubprojectForm(request.POST)
                msg = 'No se a podido realizar la transacción.'
                ctx = {'form': form, 'pid': pid, 'msg': msg}
                return render_to_response(
                    'almacen/addsector.html',
                    ctx,
                    context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages.error(request, 'No se puede mostrar la pagina.' + str(e))
        raise Http404('Method not proccess')

# Almacenes
@login_required(login_url='/SignUp/')
def view_stores(request):
    try:
        if request.method == 'GET':
            lista = Almacene.objects.filter(flag=True).order_by('nombre')
            ctx = {'lista': lista}
            return render_to_response(
                'upkeep/almacenes.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            data = {}
            try:
                obj = Almacene.objects.get(pk=request.POST.get('aid'))
                obj.delete()
                data['status'] = True
            except ObjectDoesNotExist, e:
                data['status'] = False
            return HttpResponse(
                simplejson.dumps(data),
                mimetype='application/json')
    except TemplateDoesNotExist, e:
        messages('Template not found' + str(e))
        raise Http404

@login_required(login_url='/SignUp/')
def view_stores_add(request):
    try:
        if request.method == 'GET':
            form = forms.addAlmacenesForm()
            ctx = {'form': form}
            return render_to_response(
                'upkeep/addalmacen.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            form = forms.addAlmacenesForm(request.POST)
            if form.is_valid():
                add = form.save(commit=False)
                # generate id for Stores
                c_old = Almacene.objects.all().aggregate(Max('almacen_id'))
                if c_old['almacen_id__max'] is not None:
                    cont = c_old['almacen_id__max'][2:4]
                    cont = int(cont) + 1
                else:
                    cont = 1
                add.almacen_id = 'AL%s' % ('{:0>2d}'.format(cont))
                add.flag = True
                add.save()
                return HttpResponseRedirect('/almacen/upkeep/stores/')
            else:
                form = forms.addAlmacenesForm(request.POST)
                ctx = {'form': form, 'msg': 'Transaction unrealized.'}
                return render_to_response(
                    'upkeep/addalmacen.html',
                    ctx,
                    context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages('Template not found' + str(e))
        raise Http404

@login_required(login_url='/SignUp/')
def view_stores_edit(request, aid):
    try:
        al = Almacene.objects.get(pk=aid)
        if request.method == 'GET':
            form = forms.addAlmacenesForm(instance=al)
            ctx = {'form': form, 'aid': aid}
            return render_to_response(
                'upkeep/editalmacen.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            form = forms.addAlmacenesForm(request.POST, instance=al)
            if form.is_valid():
                edit = form.save(commit=False)
                edit.flag = True
                edit.save()
                url = '/almacen/upkeep/stores/'
                return HttpResponseRedirect(url)
            else:
                form = forms.addSubprojectForm(request.POST)
                msg = 'No se a podido realizar la transacción.'
                ctx = {'form': form, 'almacen_id': almacen_id, 'msg': msg}
                return render_to_response(
                    'almacen/addsector.html',
                    ctx,
                    context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages('Template not found' + str(e))
        raise Http404

# Transportistas
@login_required(login_url='/SignUp/')
def view_carrier(request):
    try:
        if request.method == 'GET':
            lista = Transportista.objects.filter(flag=True).order_by('tranom')
            ctx = {'lista': lista}
            return render_to_response(
                'upkeep/transportista.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            data = {}
            if 'ruc' in request.POST:
                try:
                    # delete conductors
                    obj = Conductore.objects.filter(
                            traruc_id__exact=request.POST.get('ruc'))
                    obj.delete()
                    # delete Transport
                    obj = Transporte.objects.filter(
                            traruc_id__exact=request.POST.get('ruc'))
                    obj.delete()
                    # delete carrier
                    obj = Transportista.objects.get(pk=request.POST.get('ruc'))
                    obj.delete()
                    data['status'] = True
                except ObjectDoesNotExist, e:
                    data['status'] = False
                return HttpResponse(
                    simplejson.dumps(data),
                    mimetype='application/json')
    except TemplateDoesNotExist, e:
        messages('Template not found' + str(e))
        raise Http404

@login_required(login_url='/SignUp/')
def view_carrier_add(request):
    try:
        if request.method == 'GET':
            form = forms.addCarrierForm()
            ctx = {'form': form}
            return render_to_response(
                'upkeep/addcarrier.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            form = forms.addCarrierForm(request.POST)
            if form.is_valid():
                add = form.save(commit=False)
                add.flag = True
                add.save()
                return HttpResponseRedirect('/almacen/upkeep/carrier/')
            else:
                form = forms.addCarrierForm(request.POST)
                ctx = {'form': form}
                return render_to_response(
                    'upkeep/addcarrier.html',
                    ctx,
                    context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages('Template not found' + str(e))
        raise Http404

@login_required(login_url='/SignUp/')
def view_carrier_edit(request, ruc):
    try:
        t = Transportista.objects.get(pk=ruc)
        if request.method == 'GET':
            form = forms.addCarrierForm(instance=t)
            return render_to_response(
                'upkeep/editcarrier.html',
                {'form': form, 'ruc': ruc},
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            form = forms.addCarrierForm(request.POST, instance=t)
            if form.is_valid():
                edit = form.save(commit=False)
                edit.flag = True
                edit.save()
                return HttpResponseRedirect('/almacen/upkeep/carrier/')
            else:
                form = forms.addSubprojectForm(request.POST)
                msg = 'No se a podido realizar la transacción.'
                ctx = {'form': form, 'almacen_id': almacen_id, 'msg': msg}
                return render_to_response(
                    'almacen/addsector.html',
                    ctx,
                    context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages('Template not found' + str(e))
        raise Http404

# Transport
@login_required(login_url='/SignUp/')
def view_transport(request, ruc):
    try:
        if request.method == 'GET':
            lista = Transporte.objects.filter(flag=True, traruc_id=ruc)
            # print lista
            ctx = {
                'lista': lista,
                'ruc': ruc,
                'nom': lista[0].traruc.tranom if len(lista) > 0 else ''}
            return render_to_response(
                'upkeep/transport.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            if 'nropla' in request.POST:
                data = {}
                try:
                    obj = Transporte.objects.get(
                        traruc_id=ruc, condni_id=request.POST.get('nropla'))
                    obj.delete()
                    data['status'] = True
                except ObjectDoesNotExist, e:
                    data['status'] = False
                return HttpResponse(
                    simplejson.dumps(data),
                    mimetype='application/json')
    except TemplateDoesNotExist, e:
        messages('Template not found' + str(e))
        raise Http404

@login_required(login_url='/SignUp/')
def view_transport_add(request, tid):
    try:
        if request.method == 'GET':
            form = forms.addTransportForm()
            ctx = {'form': form, 'tid': tid}
            return render_to_response(
                'upkeep/addtransport.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            form = forms.addTransportForm(request.POST)
            if form.is_valid():
                add = form.save(commit=False)
                add.traruc_id = request.POST.get('traruc_id')
                add.flag = True
                add.save()
                return HttpResponseRedirect(
                    '/almacen/upkeep/transport/%s' % tid)
            else:
                form = forms.addTransportForm(request.POST)
                ctx = {'form': form, 'tid': tid}
                return render_to_response(
                    'upkeep/addtransport.html',
                    ctx,
                    context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages('Template not found')
        raise Http404(e)

@login_required(login_url='/SignUp/')
def view_transport_edit(request, cid, tid):
    try:
        t = Transporte.objects.get(traruc_id=cid, nropla_id=tid)
        if request.method == 'GET':
            form = forms.addTransportForm(instance=t)
            return render_to_response(
                'upkeep/edittransport.html',
                {'form': form, 'cid': cid, 'tid': tid},
                context_instance=RequestContext(request))
        if request.method == 'POST':
            form = forms.addTransportForm(request.POST, instance=t)
            if form.is_valid():
                form.save(commit=True)
                return HttpResponseRedirect(
                    '/almacen/upkeep/transport/%s/' % cid)
            else:
                form = forms.addTransportForm(request.POST)
                return render_to_response(
                    'upkeep/edittransport.html',
                    {'form': form, 'cid': cid, 'tid': tid},
                    context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages('Template not found')
        raise Http404(e)

# Conductor
@login_required(login_url='/SignUp/')
def view_conductor(request, ruc):
    try:
        if request.method == 'GET':
            lista = Conductore.objects.filter(flag=True, traruc_id=ruc)
            # print lista
            ctx = {
                'lista': lista,
                'ruc': ruc,
                'nom': lista[0].traruc.tranom if len(lista) > 0 else ''}
            return render_to_response(
                'upkeep/conductor.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            if 'condni' in request.POST:
                data = {}
                try:
                    obj = Conductore.objects.get(
                        traruc_id=ruc,
                        condni_id=request.POST.get('condni'))
                    obj.delete()
                    data['status'] = True
                except ObjectDoesNotExist, e:
                    data['status'] = False
                return HttpResponse(
                    simplejson.dumps(data),
                    mimetype='application/json')
    except TemplateDoesNotExist, e:
        messages('Template not found')
        raise Http404(e)

@login_required(login_url='/SignUp/')
def view_conductor_add(request, tid):
    try:
        if request.method == 'GET':
            form = forms.addConductorForm()
            ctx = {'form': form, 'tid': tid}
            return render_to_response(
                'upkeep/addconductor.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            form = forms.addConductorForm(request.POST)
            if form.is_valid():
                add = form.save(commit=False)
                add.traruc_id = request.POST.get('traruc_id')
                add.flag = True
                add.save()
                return HttpResponseRedirect(
                    '/almacen/upkeep/conductor/%s' % tid)
            else:
                form = forms.addConductorForm(request.POST)
                ctx = {'form': form, 'tid': tid}
                return render_to_response(
                    'upkeep/addconductor.html',
                    ctx,
                    context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages('Template not found')
        raise Http404(e)

@login_required(login_url='/SignUp/')
def view_conductor_edit(request, cid, tid):
    try:
        t = Conductore.objects.get(traruc_id=cid, condni_id=tid)
        if request.method == 'GET':
            form = forms.addConductorForm(instance=t)
            return render_to_response(
                'upkeep/editconductor.html',
                {'form': form, 'cid': cid, 'tid': tid},
                context_instance=RequestContext(request))
        if request.method == 'POST':
            form = forms.addConductorForm(request.POST, instance=t)
            if form.is_valid():
                form.save(commit=True)
                return HttpResponseRedirect(
                    '/almacen/upkeep/conductor/%s/' % cid)
            else:
                form = forms.addConductorForm(request.POST)
                return render_to_response(
                    'upkeep/editconductor.html',
                    {'form': form, 'cid': cid, 'tid': tid},
                    context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages('Template not found')
        raise Http404(e)

"""
  request Orders
"""

# pending request Orders
@login_required(login_url='/SignUp/')
def view_orders_pending(request):
    try:
        if request.method == 'GET':
            if request.is_ajax():
                context = dict()
                try:
                    lst = Pedido.objects.filter(
                            flag=True,
                            status='PE').order_by('-registrado')
                    context['list'] = json.loads(serializers.serialize(
                            'json', lst, relations=('proyecto', 'almacen')))
                    context['status'] = True
                except ObjectDoesNotExist as e:
                    context['raise'] = str(e)
                    context['status'] = False
                return HttpResponse(
                    simplejson.dumps(context), mimetype='application/json')
            return render_to_response(
                            'almacen/slopeorders.html',
                            context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        messages('Error template not found')
        raise Http404(e)

# list ortders attend request Orders
@login_required(login_url='/SignUp/')
def view_orders_list_approved(request):
    try:
        if request.method == 'GET':
            context = dict()
            if request.is_ajax():
                lst = Pedido.objects.filter(
                    flag=True).exclude(
                    Q(status='PE') | Q(status='AN') | Q(status='CO') |
                    Q(status='DI')).order_by('-registrado')
                context['list'] = json.loads(
                    serializers.serialize(
                        'json', lst, relations=('proyecto', 'almacen')))
                context['status'] = True
                return HttpResponse(
                        json.dumps(context),
                        mimetype='application/json')
            return render_to_response(
                'almacen/listorderattend.html',
                context_instance=RequestContext(request))
    except TemplateDoesNotExist:
        messages('Error template not found')
        raise Http404('Process Error')

# meet Orders
@login_required(login_url='/SignUp/')
def view_attend_order(request, oid):
    try:
        if request.method == 'GET':
            obj = get_object_or_404(Pedido, pk=oid, flag=True)
            # det= get_list_or_404(Detpedido, pedido_id__exact=oid,flag=True)
            # use sintaxis sql for PostgreSQL
            det = Detpedido.objects.filter(
                pedido_id__exact=oid,
                flag=True).extra(select={'stock': "SELECT stock FROM almacen_inventario WHERE almacen_detpedido.materiales_id LIKE almacen_inventario.materiales_id AND periodo LIKE to_char(now(), 'YYYY')"})
            radio = ''
            for x in det:
                if x.cantshop <= 0:
                    radio = 'disabled'
                    break
            # nipples= get_list_or_404(
            # Niple.objects.order_by('metrado'),pedido_id__exact=oid,flag=True)
            nipples = Niple.objects.filter(
                        pedido_id__exact=oid, flag=True).order_by('metrado')
            usr = userProfile.objects.get(empdni__exact=obj.empdni)
            tipo = {'A': 'Roscado', 'B': 'Ranurado', 'C': 'Rosca-Ranura'}
            ctx = {
                'orders': obj,
                'det': det,
                'nipples': nipples,
                'usr': usr,
                'tipo': tipo,
                'radio': radio}
            return render_to_response(
                'almacen/attendorder.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            # Here change method Attend order stores #mark
            try:
                data = dict()
                # recover list of materials
                mat = json.loads(request.POST.get('materials'))
                # recover list of nipples
                nip = json.loads(request.POST.get('nipples'))
                # variables
                cnm = 0
                cnn = 0
                ctn = 0
                # we walk the list materials and update items
                # materials of details orders
                for c in range(len(mat)):
                    # cs = 0
                    # for x in range(len(nip)):
                    #     if mat[c]['matid'] == nip[x]['matid']:
                    #         cs += (
                    #             float(nip[x]['quantityshop']) * float(
                    #                 nip[x]['meter']))
                    # ctn += cs
                    obj = Detpedido.objects.get(
                            pedido_id__exact=request.POST.get('oid'),
                            materiales_id__exact=mat[c]['matid'])
                    # aqui hacer otro if
                    obj.cantshop -= float(mat[c]['quantityshop'])
                    # (float(mat[c]['quantity']) - float(
                    #    mat[c]['quantityshop'])) if (cs / 100) == float(
                    #    mat[c]['quantity']) else (cs/100) if mat[c][
                    #        'matid'][0:3] == '115' else (
                    #            float(mat[c]['quantity']) - float(
                    #                mat[c]['quantityshop']))
                    # print (cs / 100 ) if mat[c]['matid'][0:3] == '115' else (
                    # float(mat[c]['quantity'])- float(mat[c]['quantityshop']))
                    obj.cantguide = float(mat[c]['quantityshop'])
                    if obj.cantshop < 0:
                        obj.tag = '2'
                    else:
                        obj.tag = '1'
                    obj.save()
                    cnm += 1
                # we walk the list nipples and update tag of tables nipples
                for n in range(len(nip)):
                    obj = Niple.objects.get(pk=nip[n]['nid'])
                    # obj.cantshop = int(
                    #     float(nip[n]['quantity']) - float(
                    #         nip[n]['quantityshop']))
                    obj.cantguide = float(nip[n]['quantityshop'])
                    obj.cantshop -= float(nip[n]['quantityshop'])
                    obj.tag = '1'
                    obj.save()
                    cnn += 1
                # evaluation status orders
                # recover number of materials
                status = ''
                onm = Detpedido.objects.filter(
                    pedido_id__exact=request.POST.get('oid'),
                    cantshop__gt=0).exclude(tag='2').count()
                if onm > 0:
                    status = 'IN'
                else:
                    status = 'CO'
                # update status Bedside Orders
                obj = Pedido.objects.get(pk=request.POST.get('oid'))
                obj.status = status
                obj.save()
                data['sts'] = status
                data['pass'] = status
                data['status'] = True
            except ObjectDoesNotExist, e:
                data['status'] = False
            return HttpResponse(
                simplejson.dumps(data),
                mimetype='application/json')
    except TemplateDoesNotExist, e:
        message('Error template not found')
        raise Http404

"""
    guide remision
"""

# generate guide remision of a orders
@login_required(login_url='/SignUp/')
def view_generate_guide_orders(request):
    try:
        if request.method == 'GET':
            # lst= get_list_or_404(Pedido.objects.exclude(Q(status='PE')|
            # Q(status='AN')).order_by('-pedido_id'), flag=True )
            lst = Detpedido.objects.filter(tag='1').order_by(
                    '-pedido').distinct('pedido')
            ctx = {'orders': lst}
            return render_to_response(
                'almacen/generateGuide.html',
                ctx,
                context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        message('Error Template not found')
        raise Http404

# request generate guide remision
@login_required(login_url='/SignUp/')
def view_generate_document_out(request, oid):
    try:
        if request.method == 'GET':
            orders = get_object_or_404(Pedido, flag=True, pedido_id__exact=oid)
            trans = get_list_or_404(Transportista.objects.values(
                'traruc_id', 'tranom'), flag=True)
            ctx = {'oid': oid, 'trans': trans, 'orders': orders}
            return render_to_response(
                'almacen/documentout.html',
                ctx,
                context_instance=RequestContext(request))
        elif request.method == 'POST':
            if request.is_ajax():
                form = forms.addGuideReferral(request.POST)
                if form.is_valid():
                    data = dict()
                    try:
                        add = form.save(commit=False)
                        guidekeys = request.POST.get('guia_id')
                        # genkeys.GenerateSerieGuideRemision()
                        code = guidekeys.split('-')
                        code = (
                            '%s-%s' % (
                                '{:0>3d}'.format(int(code[0])),
                                '{:0>8d}'.format(int(code[1]))))
                        add.guia_id = code
                        add.status = 'GE'
                        add.flag = True
                        add.registrado = datetime.datetime.today()
                        # commit true save bedside guide
                        add.save()
                        # save details guide referral
                        # recover details orders
                        det = Detpedido.objects.filter(
                            pedido_id__exact=request.POST.get(
                                'pedido'), tag='1', flag=True)
                        for x in det:
                            obj = DetGuiaRemision()
                            obj.guia_id = code
                            obj.materiales_id = x.materiales_id
                            obj.cantguide = x.cantguide
                            obj.flag = True
                            obj.save()
                            ob = Detpedido.objects.get(pk__exact=x.id)
                            ob.tag = '2' if x.cantshop <= 0 else '0'
                            ob.save()
                            # here discount inventory
                            # get nro orders
                            store = ''
                            try:
                                store = Pedido.objects.get(
                                        pk=request.POST.get('pedido'))
                                store = store.almacen_id
                            except ObjectDoesNotExist, e:
                                store = 'AL01'
                            try:
                                inv = Inventario.objects.get(
                                    materiales_id=x.materiales_id,
                                    periodo=globalVariable.get_year,
                                    almacen_id=store)
                                stock = inv.stock
                                inv.stock = (stock - float(x.cantguide))
                                inv.save()
                            except ObjectDoesNotExist, e:
                                print e
                        # brands = InventoryBrand.objects.filter(
                        # materiales_id=x.materiales_id,
                        # periodo=globalVariable.get_year)

                        # recover details nipples
                        nip = Niple.objects.filter(
                            pedido_id__exact=request.POST.get('pedido'),
                            tag='1', flag=True)
                        print 'DETAILS NIP ', nip, nip.count()
                        for x in nip:
                            obj = NipleGuiaRemision()
                            obj.guia_id = code
                            obj.materiales_id = x.materiales_id
                            obj.metrado = x.metrado
                            obj.cantguide = x.cantguide
                            obj.tipo = x.tipo
                            obj.flag = True
                            obj.brand_id = x.brand_id
                            obj.model_id = x.model_id
                            # save details niples for guide referral
                            obj.save()
                            ob = Niple.objects.get(pk__exact=x.id)
                            ob.tag = '2' if x.cantshop <= 0 else '0'
                            ob.save()
                        data['status'] = True
                        data['guide'] = code
                    except ObjectDoesNotExist, e:
                        print 'Error to ingress NIPPLE'
                        print e
                        data['status'] = False
                else:
                    data['status'] = False
                    data['raise'] = 'Form invalid'
                return HttpResponse(
                    simplejson.dumps(data),
                    mimetype='application/json')
    except TemplateDoesNotExist, e:
        message('Error: Template not found')
        raise Http404(e)

# recover list guide referral for view and annular
@login_required(login_url='/SignUp/')
def view_list_guide_referral_success(request):
    try:
        if request.method == 'GET':
            if request.is_ajax():
                data = dict()
                ls = list()
                lst = list()
                try:
                    if request.GET.get('tra') == 'series':
                        lst = GuiaRemision.objects.filter(
                                pk=request.GET.get('series'),
                            status='GE', flag=True)
                        data['status'] = True
                    elif request.GET['tra'] == 'project':
                        lst = GuiaRemision.objects.filter(
                            pedido__proyecto__nompro__icontains=request.GET[
                            'project'], status='GE', flag=True).order_by('-traslado')
                        data['status'] = True
                    elif request.GET.get('tra') == 'dates':
                        fecf = request.GET.get('fecf')
                        feci = request.GET.get('feci')
                        if fecf == '' and feci != '':
                            star = datetime.datetime.strptime(
                                    request.GET.get('feci'),
                                    FORMAT_DATE_STR).date()
                            lst = GuiaRemision.objects.filter(
                                    traslado=star, status='GE', flag=True)
                        elif fecf != '' and feci != '':
                            star = datetime.datetime.strptime(
                                    request.GET.get('feci'),
                                    FORMAT_DATE_STR).date()
                            end = datetime.datetime.strptime(
                                    request.GET.get('fecf'),
                                    FORMAT_DATE_STR).date()
                            lst = GuiaRemision.objects.filter(
                                    traslado__range=[star, end],
                                    status='GE',
                                    flag=True)
                        data['status'] = True
                    i = 1
                    for x in lst:
                        ls.append({
                            'item': i,
                            'guia_id': x.guia_id,
                            'nompro': x.pedido.proyecto.nompro,
                            'traslado': x.traslado.strftime(
                                            FORMAT_DATE_STR),
                            'connom': x.condni.connom})
                        i += 1
                except ObjectDoesNotExist, e:
                    data['status'] = False
                data['list'] = ls
                return HttpResponse(
                    simplejson.dumps(data),
                    mimetype='application/json')
            lst = GuiaRemision.objects.filter(
                    status='GE',
                    flag=True).order_by('-guia_id')[:10]
            ctx = {'guide': lst}
            return render_to_response(
                'almacen/listguide.html',
                ctx,
                context_instance=RequestContext(request))

        if request.method == 'POST':
            data = dict()
            try:
                obj = GuiaRemision.objects.get(
                        pk=request.POST.get('series'), status='GE', flag=True)
                obj.status = 'AN'
                obj.flag = False
                obj.save()
                det = DetGuiaRemision.objects.filter(
                        guia_id__exact=request.POST.get('series'))
                det.update(flag=False)
                nip = NipleGuiaRemision.objects.filter(
                    guia_id__exact=request.POST.get('series'))
                nip.update(flag=False)
                # generate beside restoration
                add = Restoration()
                res = genkeys.keyRestoration().strip()
                add.restoration_id = res
                if obj.pedido_id:
                    add.almacen_id = obj.pedido.almacen_id
                else:
                    add.almacen_id = 'AL01'
                add.ndocument_id = obj.guia_id
                add.observation = request.POST['observation']
                add.performed_id = request.user.get_profile().empdni_id
                add.save()
                for x in det:
                    # return quantity at inventory
                    inv = Inventario.objects.filter(
                            periodo=globalVariable.get_year,
                            materiales_id=x.materiales_id).order_by(
                            '-ingreso')[0]
                    if x.brand_id:
                        brand = x.brand_id
                    else:
                        brand = 'BR000'
                    if x.model_id:
                        model = x.model_id
                    else:
                        model = 'MO000'
                    # return quantity at inventory brand
                    ibm = InventoryBrand.objects.filter(
                        period=globalVariable.get_year,
                        materials_id=x.materiales_id,
                        brand_id=brand,
                        model_id=model).order_by('-ingress')
                    if ibm:
                        ibm = ibm[0]
                        inv.stock = (float(inv.stock) + float(x.cantguide))
                        inv.save()
                        ibm.stock = (float(ibm.stock) + float(x.cantguide))
                        ibm.save()
                    # Save Details Restoration
                    dt = DetRestoration()
                    dt.restoration_id = res
                    dt.materials_id = x.materiales_id
                    dt.brand_id = brand
                    dt.model_id = model
                    dt.quantity = x.cantguide
                    dt.save()
                    # Return quantity at order if exists
                    if obj.pedido_id:
                        ords = Detpedido.objects.get(
                            pedido_id=obj.pedido_id,
                            materiales_id=x.materiales_id,
                            brand_id=brand,
                            model_id=model)
                        ords.cantshop = (ords.cantshop + x.cantguide)
                        ords.cantguide = (ords.cantguide - x.cantguide)
                        ords.tag = '1'
                        ords.save()
                # if guide nipple exist for material
                if obj.pedido_id:
                    gn = NipleGuiaRemision.objects.filter(guia_id=obj.guia_id)
                    if gn:
                        for n in gn:
                            try:
                                np = Niple.objects.get(
                                    pedido_id=obj.pedido_id,
                                    materiales_id=n.materiales_id,
                                    metrado=n.metrado,
                                    tipo=n.tipo)
                                np.cantguide = (np.cantguide - n.cantguide)
                                np.cantshop = (np.cantshop + n.cantguide)
                                np.save()
                            except Niple.DoesNotExist:
                                raise e
                            n.flag = False
                            n.save()
                    pe = Pedido.objects.get(
                        pedido_id=obj.pedido_id)
                    pe.status = 'IN'
                    pe.save()
                data['status'] = True
            except ObjectDoesNotExist, e:
                print e
                data['raise'] = str(e)
                data['status'] = False
            return HttpResponse(
                simplejson.dumps(data), mimetype='application/json')
    except TemplateDoesNotExist, e:
        raise Http404(e)

# recover list guide referral for view and annular
@login_required(login_url='/SignUp/')
def view_list_guide_referral_canceled(request):
    try:
        if request.method == 'GET':
            if request.is_ajax():
                data = dict()
                ls = list()
                lst = list()
                try:
                    if request.GET.get('tra') == 'series':
                        lst = GuiaRemision.objects.filter(
                            pk=request.GET.get('series'),
                            status='AN', flag=False)
                        data['status'] = True
                    elif request.GET['tra'] == 'project':
                        lst = GuiaRemision.objects.filter(
                            pedido__proyecto__nompro__icontains=request.GET[
                            'project'], status='AN').order_by('-traslado')
                        data['status'] = True
                    elif request.GET.get('tra') == 'dates':
                        if request.GET.get('fecf') == '' and request.GET.get('feci') != '':
                            star = datetime.datetime.strptime(
                                request.GET.get('feci'),
                                FORMAT_DATE_STR).date()
                            lst = GuiaRemision.objects.filter(
                                traslado=star, status='AN', flag=False)
                        elif request.GET.get('fecf') != '' and request.GET.get('feci') != '':
                            star = datetime.datetime.strptime(
                                request.GET.get('feci'),
                                FORMAT_DATE_STR).date()
                            end = datetime.datetime.strptime(
                                request.GET.get('fecf'),
                                FORMAT_DATE_STR).date()
                            lst = GuiaRemision.objects.filter(
                                traslado__range=[star, end],
                                status='AN',
                                flag=False)
                        data['status'] = True
                    i = 1
                    for x in lst:
                        ls.append({
                            'item': i,
                            'guia_id': x.guia_id,
                            'nompro': x.pedido.proyecto.nompro,
                            'traslado': x.traslado.strftime(
                                            FORMAT_DATE_STR),
                            'connom': x.condni.connom})
                        i += 1
                except ObjectDoesNotExist, e:
                    data['status'] = False
                    data['raise'] = str(e)
                data['list'] = ls
                return HttpResponse(
                    simplejson.dumps(data),
                    mimetype='application/json')
            lst = GuiaRemision.objects.filter(
                status='AN', flag=False).order_by('-guia_id')[:10]
            ctx = {'guide': lst}
            return render_to_response(
                'almacen/listguidecanceled.html',
                ctx,
                context_instance=RequestContext(request))
    except TemplateDoesNotExist, e:
        raise Http404(e)

###########################
# Views natives of stores #
###########################
class InventoryView(ListView, JSONResponseMixin):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            if request.GET.get('tipo') == 'desc':
                if bool(int(request.GET.get('stkzero'))) and request.GET.get('stkmin') == 'None':
                    model = Inventario.objects.filter(
                        materiales__matnom__icontains=request.GET.get('omat'),
                        periodo=request.GET.get('periodo'),
                        almacen=request.GET.get('almacen'), stock__lte=0)
                elif bool(int(request.GET.get('stkzero'))) and request.GET.get('stkmin') != 'None':
                    model = Inventario.objects.filter(
                        materiales__matnom__icontains=request.GET.get('omat'),
                        periodo=request.GET.get('periodo'),
                        almacen=request.GET.get('almacen'),
                        stock__lte=0,
                        stkmin=request.GET.get('stkmin'))
                elif not bool(int(request.GET.get('stkzero'))) and request.GET.get('stkmin') != 'None':
                    model = Inventario.objects.filter(
                        materiales__matnom__icontains=request.GET.get('omat'),
                        periodo=request.GET.get('periodo'),
                        almacen=request.GET.get('almacen'))
                else:
                    model = Inventario.objects.filter(
                        materiales__matnom__icontains=request.GET.get('omat'),
                        periodo=request.GET.get('periodo'),
                        almacen=request.GET.get('almacen'))
            elif request.GET.get('tipo') == 'cod':
                if bool(int(request.GET.get('stkzero'))) and request.GET.get('stkmin') == 'None':
                    model = Inventario.objects.filter(
                        materiales__materiales_id__startswith=request.GET.get(
                            'omat'), periodo=request.GET.get('periodo'),
                        almacen=request.GET.get('almacen'), stock__lte=0)
                elif bool(int(request.GET.get('stkzero'))) and request.GET.get('stkmin') != 'None':
                    model = Inventario.objects.filter(
                        materiales__materiales_id__startswith=request.GET.get(
                            'omat'), periodo=request.GET.get('periodo'),
                        almacen=request.GET.get('almacen'),
                        stock__lte=0, stkmin=request.GET.get('stkmin'))
                elif not bool(int(request.GET.get('stkzero'))) and request.GET.get('stkmin') != 'None':
                    model = Inventario.objects.filter(
                        materiales__materiales_id__startswith=request.GET.get(
                            'omat'),
                        periodo=request.GET.get('periodo'),
                        almacen=request.GET.get('almacen'))
                else:
                    model = Inventario.objects.filter(
                        materiales__materiales_id__startswith=request.GET.get(
                            'omat'),
                        periodo=request.GET.get('periodo'),
                        almacen=request.GET.get('almacen'))
            counter = 0
            paginator = Paginator(model, 20)
            page = request.GET.get('page')
            try:
                materials = paginator.page(page)
            except PageNotAnInteger:
                materials = paginator.page(1)
            except EmptyPage:
                materials = paginator.page(paginator.num_pages)
            data = {'list': []}
            for x in materials:
                data['list'].append({
                    'materiales_id': x.materiales_id,
                    'matnom': x.materiales.matnom,
                    'matmed': x.materiales.matmed,
                    'unid': x.materiales.unidad_id,
                    'stkmin': x.stkmin,
                    'stock': x.stock,
                    'ingreso': x.ingreso.strftime(FORMAT_DATE_STR),
                    'compra_id': x.compra_id,
                    'spptag': x.spptag})
            data['has_previous'] = materials.has_previous()
            if materials.has_previous():
                data['previous_page_number'] = materials.previous_page_number()
            data['number'] = materials.number
            data['num_pages'] = paginator.num_pages
            data['has_next'] = materials.has_next()
            if materials.has_next():
                data['next_page_number'] = materials.next_page_number()
            return HttpResponse(
                simplejson.dumps(data),
                mimetype='application/json')
        else:
            model = Inventario.objects.filter(
                periodo=datetime.datetime.today().date().year.__str__(),
                flag=True).order_by('materiales')
        context = dict()
        context['periodo'] = [
            x['periodo'] for x in Inventario.objects.values(
                'periodo').order_by('-periodo').distinct('periodo')]
        context['almacen'] = [{
            'alid': x.almacen_id,
            'nom': x.nombre} for x in Almacene.objects.filter(flag=True)]
        paginator = Paginator(model, 20)
        # Show 25 materials per page
        page = request.GET.get('page')
        try:
            materials = paginator.page(page)
        except PageNotAnInteger:
            # If page not is an integer, delivery first page
            materials = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), delibery last page of result
            materials = paginator.page(paginator.num_pages)
        context['inventory'] = materials
        return render_to_response(
            'almacen/inventory.html',
            context,
            context_instance=RequestContext(request))

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        data = dict()
        try:
            tipo = request.POST.get('tipo')
            if tipo == 'save-tmp':
                obj = tmpsuministro()
                obj.empdni = request.user.get_profile().empdni_id
                obj.materiales_id = request.POST.get('add-id')
                obj.cantidad = request.POST.get('add-cant')
                obj.origin_id = request.POST.get('add-oid')
                obj.origin = request.POST.get('add-ori')
                obj.save()
                obj = Inventario.objects.get(
                    materiales_id=request.POST.get('add-id'),
                    almacen_id=request.POST.get('add-oid'),
                    periodo=datetime.datetime.today().date().strftime('%Y'))
                obj.spptag = True
                obj.save()
                data['status'] = True
            if tipo == 'all':
                sts = Inventario.register_all_list_materilas(
                    request.POST.get('alid'), request.POST.get('quantity'))
                data['status'] = sts
            if tipo == 'per':
                print 'ingress step'
                sts = Inventario.register_period_past(
                    request.POST.get('alcp'),
                    request.POST.get('pewh'),
                    request.POST.get('alwh')
                )
                print sts
                data['status'] = sts
            if 'uploadInventory' in request.POST:
                path = '/storage/Temp/'
                options = {'name': 'tmpstorage'}
                filename = uploadFiles.upload(
                            path,
                            request.FILES['inventory'],
                            options)
                if uploadFiles.fileExists(filename):
                    wb = load_workbook(filename)
                    ws = wb.worksheets[0]
                    for x in range(1, ws.max_row+1):
                        cell = ws.cell(row=x, column=1).value
                        if len(str(cell)) == 15 and cell != None:
                            if float(str(ws.cell(row=x, column=4).value)) <= 0 or ws.cell(row=x, column=4).value == None:
                                continue
                            else:
                                purchase = 0
                                try:
                                    p = MetProject.objects.filter(
                                            materiales_id=str(cell)).order_by(
                                            'proyecto__registrado')
                                    if len(p):
                                        p = p[0]
                                        purchase = p.precio
                                except MetProject.DoesNotExist:
                                    raise
                                # inventory general
                                try:
                                    inv = Inventario.objects.get(
                                        materiales_id=str(cell),
                                        periodo=globalVariable.get_year)
                                    inv.stock += float(str(ws.cell(row=x, column=4).value))
                                    inv.save()
                                except Inventario.DoesNotExist:
                                    inv = Inventario()
                                    inv.materiales_id = str(cell)
                                    inv.almacen_id = 'AL01'
                                    inv.precompra = purchase
                                    inv.stkmin = 0
                                    inv.stock = float(str(ws.cell(row=x, column=4).value))
                                    inv.stkpendiente = 0
                                    inv.stkdevuelto = 0
                                    inv.periodo = globalVariable.get_year
                                    inv.ingreso = globalVariable.date_now()
                                    inv.save()
                                # inventory brand
                                bc = None
                                brand = str(ws.cell(row=x, column=3).value)
                                ml = 'MO000'
                                if brand == None:
                                    bc = 'BR000'
                                else:
                                    if len(brand.split('/')) == 2:
                                        try:
                                            bc = Brand.objects.get(brand__exact=brand.split('/')[-1].strip())
                                            bc = bc.brand_id
                                        except Brand.DoesNotExist:
                                            bc = Brand()
                                            bc.brand_id = genkeys.GenerateIdBrand()
                                            bc.brand = brand.strip().upper()
                                            bc.save()
                                            bc = bc.brand_id
                                        try:
                                            ml = Model.objects.get(
                                                model__exact=brand.split('/')[0].strip().upper())
                                            print ml.model
                                            print brand.split('/')[1].upper()
                                            if ml.model == brand.split('/')[0].strip().upper():
                                                ml = ml.model_id
                                            else:
                                                ml = Model()
                                                ml.model_id = genkeys.GenerateIdModel()
                                                ml.brand_id = bc
                                                ml.model = brand.split('/')[0].strip().upper()
                                                ml.save()
                                                ml = ml.model_id
                                        except ObjectDoesNotExist:
                                            ml = Model()
                                            ml.model_id = genkeys.GenerateIdModel()
                                            ml.brand_id = bc
                                            ml.model = brand.split('/')[0].strip().upper()
                                            ml.save()
                                            ml = ml.model_id
                                    else:
                                        try:
                                            bc = Brand.objects.get(brand__exact=brand)
                                            bc = bc.brand_id
                                        except Brand.DoesNotExist:
                                            bc = Brand()
                                            bc.brand_id = genkeys.GenerateIdBrand()
                                            bc.brand = brand.strip().upper()
                                            bc.save()
                                            bc = bc.brand_id
                                try:
                                    ib = InventoryBrand.objects.get(
                                        materials_id=str(cell),
                                        brand_id=bc,
                                        model_id=ml,
                                        period=globalVariable.get_year)
                                    ib.stock += float(str(ws.cell(row=x, column=4).value))
                                    ib.save()
                                except InventoryBrand.DoesNotExist:
                                    ib = InventoryBrand()
                                    ib.storage_id = 'AL01'
                                    ib.period = globalVariable.get_year
                                    ib.materials_id = str(cell)
                                    ib.brand_id = bc
                                    ib.model_id = ml
                                    ib.ingress = globalVariable.date_now()
                                    ib.stock = float(str(ws.cell(row=x, column=4).value))
                                    ib.purchase = purchase
                                    ib.sale = ((purchase * 0.1)+purchase)
                                    ib.save()
                                    bc = None
                                    ml = None
                                    print '------------------'
                        else:
                            continue
                    data['status'] = True
                else:
                    data['status'] = False
            if 'delInventory' in request.POST:
                Inventario.objects.filter(
                    periodo=globalVariable.date_now().strftime('%Y')).delete()
                InventoryBrand.objects.filter(
                    period=globalVariable.date_now().strftime('%Y')).delete()
                data['status'] = True
        except ObjectDoesNotExist, e:
            data['raise'] = str(e)
            data['status'] = False
        # data = simplejson.dumps(data)
        return self.render_to_json_response(data)

class SupplyView(ListView):
    template_name = 'almacen/supply.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            data = dict()
            try:
                arr = json.loads(request.GET.get('mats'))
                queryset = tmpsuministro.objects.extra(select={
                    'stock': "SELECT stock FROM almacen_inventario WHERE almacen_tmpsuministro.materiales_id LIKE almacen_inventario.materiales_id AND periodo LIKE to_char(now(), 'YYYY')"}).filter(empdni__exact=request.user.get_profile().empdni_id, materiales_id__in=arr)
                queryset = queryset.values(
                    'materiales_id',
                    'materiales__matnom',
                    'materiales__matmed',
                    'materiales__unidad_id',
                    'brand__brand',
                    'model__model',
                    'stock')
                queryset = queryset.annotate(
                    cantidad=Sum('cantidad')).order_by('materiales__matnom')
                data['list'] = [
                    {
                        'materiales_id': x['materiales_id'],
                        'matnom': x['materiales__matnom'],
                        'matmed': x['materiales__matmed'],
                        'unidad': x['materiales__unidad_id'],
                        'brand': x['brand__brand'],
                        'model': x['model__model'],
                        'cantidad': x['cantidad'],
                        'stock': x['stock']
                    }
                    for x in queryset
                ]
                data['status'] = True
            except ObjectDoesNotExist:
                data['status'] = False
            data = simplejson.dumps(data)
            return HttpResponse(
                data,
                mimetype='application/json',
                content_type='application/json')
        context = {}
        context['tmp'] = tmpsuministro.objects.filter(
            empdni__exact=request.user.get_profile().empdni_id
            ).order_by('materiales__matnom')
        context['almacen'] = Almacene.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            data = {}
            try:
                if request.POST.get('tipo') == 'deltmp':
                    obj = tmpsuministro.objects.filter(
                        empdni__exact=request.user.get_profile().empdni_id)
                    if obj:
                        for x in obj:
                            det = None
                            if x.origin == 'PE':
                                det = Detpedido.objects.get(
                                    pedido_id__exact=x.origin_id,
                                    materiales_id=x.materiales_id)
                            elif x.origin == 'AL':
                                det = Inventario.objects.get(
                                    almacen_id=x.origin_id,
                                    materiales_id=x.materiales_id,
                                    periodo=globalVariable.date_now(
                                        type='str', format='%Y'))
                            else:
                                continue
                            det.spptag = False
                            det.save()
                        obj.delete()
                        data['status'] = True
                    else:
                        data['status'] = False
                    data = simplejson.dumps(data)
                    return HttpResponse(data, mimetype='application/json')
                if 'generateSupply' in request.POST:
                    # save bedside of supply
                    idsp = genkeys.GenerateKeySupply()
                    bed = Suministro()
                    bed.suministro_id = idsp
                    bed.almacen_id = request.POST.get('almacen')
                    bed.empdni = request.user.get_profile().empdni_id
                    bed.ingreso = request.POST.get('ingreso')
                    bed.obser = request.POST.get('obser')
                    bed.flag = True
                    bed.asunto = request.POST.get('asunto')
                    bed.status = 'PE'
                    # details supply
                    obj = tmpsuministro.objects.filter(
                        empdni=request.user.get_profile().empdni_id)
                    proj = obj.values('orders__proyecto__proyecto_id')
                    proj = proj.distinct(
                        'orders__proyecto__proyecto_id'
                        ).order_by('orders__proyecto__proyecto_id')
                    # print ','.join([x['orders__proyecto__proyecto_id
                    # '] for x in proj])
                    print proj
                    bed.orders = ','.join(
                        [x['orders__proyecto__proyecto_id'] for x in proj]
                        ) if proj[0][
                        'orders__proyecto__proyecto_id'] is not None else ''
                    bed.save()
                    for x in obj:
                        det = DetSuministro()
                        det.suministro_id = idsp
                        det.materiales_id = x.materiales_id
                        det.cantidad = float(x.cantidad)
                        det.cantshop = float(x.cantidad)
                        det.tag = '1'
                        det.origin = x.origin
                        det.save()

                    obj.delete()
                    data['status'] = True
                    data['nro'] = idsp
            except ObjectDoesNotExist, e:
                context['raise'] = e.__str__()
                data['status'] = False
            data = simplejson.dumps(data)
            return HttpResponse(
                data,
                mimetype='application/json',
                content_type='application/json')

class ListOrdersSummary(TemplateView):
    template_name = 'almacen/listorderssupply.html'
    context_object_name = 'Orders'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        context[self.context_object_name] = Pedido.objects.filter(
                                Q(flag=True),
                                Q(status='AP') |
                                Q(status='IN')).order_by('-registrado')
        return render_to_response(
                self.template_name,
                context,
                context_instance=RequestContext(request))

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            data = {}
            try:
                obj = tmpsuministro()
                # arr = json.loads(request.POST.get('add-oid'))
                # for x in arr.__len__():
                obj.empdni = request.user.get_profile().empdni_id
                obj.materiales_id = request.POST.get('id-add')
                obj.cantidad = (request.POST.get('cant-add'))
                # obj.origin_id = arr[x]
                obj.origin_id = request.POST.get('add-ori')
                obj.save()
                arr = json.loads(request.POST.get('orders'))
                Detpedido.objects.filter(
                    Q(flag=True) & Q(pedido_id__in=arr) & Q(
                        materiales_id=request.POST.get('id-add'))
                    ).update(spptag=True)
                data['status'] = True
            except ObjectDoesNotExist:
                data['status'] = False
            return HttpResponse(
                simplejson.dumps(data),
                mimetype='application/json')

class ListDetOrders(JSONResponseMixin, TemplateView):
    template_name = 'almacen/listdetailsOrders.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        context['orders'] = Pedido.objects.filter(
            Q(flag=True) & Q(status='AP') | Q(status='IN')
            ).order_by('-pedido_id')
        orders = [x.pedido_id for x in context['orders']]
        context['details'] = Detpedido.objects.filter(
            pedido_id__in=orders).extra(
            select={
                'stock': "SELECT stock FROM almacen_inventario WHERE almacen_detpedido.materiales_id LIKE almacen_inventario.materiales_id AND periodo LIKE to_char(now(), 'YYYY')"}).order_by('materiales__matnom')
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            data = dict()
            try:
                mats = json.loads(request.POST.get('mats'))
                for x in range(mats.__len__()):
                    obj = tmpsuministro()
                    obj.empdni = request.user.get_profile().empdni_id
                    obj.materiales_id = mats[x]['mid']
                    obj.cantidad = float(mats[x]['cant'].__str__())
                    obj.origin = request.POST.get('addori')
                    obj.orders_id = mats[x]['oid']
                    obj.save()
                    dor = Detpedido.objects.get(
                        pedido_id=mats[x]['oid'], materiales_id=mats[x]['mid'])
                    dor.spptag = True
                    dor.save()
                data['status'] = True
            except ObjectDoesNotExist, e:
                data['raise'] = e.__str__()
                data['status'] = False
            # data = simplejson.dumps(data)
            return self.render_to_json_response(data)

# Input order purchase
class InputOrderPurchase(JSONResponseMixin, TemplateView):
    template_name = 'almacen/inputpurchase.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'type' in request.GET:
                        result = list()
                        if request.GET.get('type') == 'code':
                            result = Compra.objects.filter(
                                    compra_id=request.GET.get('code'))
                            context['status'] = True
                        elif request.GET.get('type') == 'dates':
                            if 'end' not in request.GET and 'start' in request.GET:
                                result = Compra.objects.filter(
                                        registrado__startswith=globalVariable.format_str_date(
                                            request.GET.get('start')))
                                context['status'] = True
                            elif 'end' in request.GET and 'start' in request.GET:
                                result = Compra.objects.filter(
                                    registrado__range=(
                                        globalVariable.format_str_date(
                                            request.GET.get('start')),
                                        globalVariable.format_str_date(
                                            request.GET.get('end'))))
                        elif request.GET['type'] == 'supplier':
                            result = Compra.objects.filter(proveedor__razonsocial__icontains=request.GET['supplier'])
                            context['status'] = True
                        context['list'] = [{
                            'purchase': x.compra_id,
                            'reason': x.proveedor.razonsocial,
                            'supplier': x.proveedor_id,
                            'document': x.documento.documento,
                            'transfer': globalVariable.format_date_str(x.traslado)
                        } for x in result]
                    if 'purchase' in request.GET:
                        com = Compra.objects.get(
                                pk=request.GET.get('purchase'))
                        context['head'] = {
                            'supplier': com.proveedor_id,
                            'quote': com.cotizacion_id if com.cotizacion_id else 'None',
                            'place': com.lugent,
                            'document': com.documento.documento,
                            'payment': com.pagos.pagos,
                            'currency': com.moneda.moneda,
                            'register': globalVariable.format_date_str(
                                            com.registrado),
                            'transfer': globalVariable.format_date_str(
                                            com.traslado),
                            'contact': com.contacto,
                            'deposit': '%s' % com.deposito,
                            'performed': '%s, %s' % (
                                com.empdni.firstname, com.empdni.lastname)}
                        context['details'] = [{
                            'materials': x.materiales_id,
                            'name': x.materiales.matnom,
                            'measure': x.materiales.matmed,
                            'unit': x.materiales.unidad.uninom,
                            'dunit': x.unit.uninom if x.unit is not None else '',
                            'quantity': x.cantidad,
                            'static': x.cantstatic,
                            'price': x.precio,
                            'discount': float(x.discount),
                            'brand': x.brand.brand,
                            'brand_id': x.brand_id,
                            'model': x.model.model,
                            'model_id': x.model_id}
                            for x in DetCompra.objects.filter(
                                compra_id=request.GET.get('purchase')
                                ).order_by(
                                    'materiales__matnom') if x.flag != '2']
                        context['status'] = True
                except ObjectDoesNotExist, e:
                    context['raise'] = e.__str__()
                    context['status'] = False
                return self.render_to_json_response(context)

            context['purchase'] = Compra.objects.filter(
                    Q(flag=True),
                    Q(status='PE') | Q(status='IN')
                ).order_by('-compra_id')
            context['storage'] = Almacene.objects.filter(flag=True)
            context['employee'] = Employee.objects.filter(
                                    flag=True, 
                                    charge__area__icontains='almacen')
            return render_to_response(
                    self.template_name,
                    context,
                    context_instance=RequestContext(request))
        except TemplateDoesNotExist, e:
            raise Http404('Template not found')

    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            context = dict()
            try:
                # print request.POST
                if 'ingress' in request.POST:
                    form = forms.addNoteIngress(request.POST)
                    print form
                    if form.is_valid():
                        add = form.save(commit=False)
                        ingress = genkeys.GenerateIdNoteIngress()
                        add.ingress_id = ingress
                        add.status = 'CO'
                        add.save()
                        # save to bedside Note Ingress
                        # save details Note Ingress
                        details = json.loads(request.POST.get('details'))
                        # counter = 0
                        for x in details:
                            # details Note Ingress
                            det = DetIngress()
                            det.ingress_id = ingress
                            det.materials_id = x['materials']
                            det.quantity = x['quantity']
                            det.brand_id = x[
                                'brand'] if 'brand' in x else 'BR000'
                            det.model_id = x[
                                'model'] if 'model' in x else 'MO000'
                            det.report = '0'
                            det.purchase = float(x['price'])
                            det.sales = round(float(x['price']) * 1.15, 2)
                            det.save()
                        #     inv = Inventario.objects.filter(
                        #             materiales_id=x['materials'],
                        #             periodo=globalVariable.get_year,
                        #             almacen_id=request.POST.get('storage'),
                        #             flag=True)
                        #     if inv:
                        #         inv[0].stock = (
                        #             inv[0].stock + float(x['quantity']))
                        #         inv[0].save()
                        #     else:
                        #         inv = Inventario()
                        #         inv.almacen_id = request.POST.get('storage')
                        #         inv.materiales_id = x['materials']
                        #         inv.stock = x['quantity']
                        #         inv.precompra = float(x['price'])
                        #         inv.preventa = (
                        #             float(
                        #                 x['price']) + (
                        #                     float(x['price']) * 0.15))
                        #         inv.stkmin = 0
                        #         inv.stkpendiente = 0
                        #         inv.stkdevuelto = 0
                        #         inv.periodo = globalVariable.get_year
                        #         inv.compra_id = request.POST.get('purchase')
                        #         inv.save()
                        #     wbm = InventoryBrand.objects.filter(
                        #         storage_id=request.POST.get('storage'),
                        #         materials_id=x['materials'],
                        #         period=globalVariable.get_year,
                        #         brand_id=x[
                        #             'brand'] if 'brand' in x else 'BR000',
                        #         model_id=x[
                        #             'model'] if 'model' in x else 'MO000')
                        #     if wbm:
                        #         wbm[0].stock = wbm[0].stock + float(
                        #             x['quantity'])
                        #         wbm[0].save()
                        #     else:
                        #         bm = InventoryBrand()
                        #         bm.storage_id = request.POST.get('storage')
                        #         bm.materials_id = x['materials']
                        #         bm.period = globalVariable.get_year
                        #         bm.brand_id = x[
                        #             'brand'] if 'brand' in x else 'BR000'
                        #         bm.model_id = x[
                        #             'model'] if 'model' in x else 'MO000'
                        #         bm.stock = x['quantity']
                        #         bm.purchase = float(x['price'])
                        #         bm.sale = (
                        #             float(x['price']) + (
                        #                 float(x['price']) * 0.15))
                        #         bm.save()
                        #     dbuy = DetCompra.objects.get(
                        #             compra_id=request.POST.get('purchase'),
                        #             materiales_id=x['materials'])
                        #     dbuy.flag = x['tag']
                        #     dbuy.cantidad = (
                        #         dbuy.cantidad - float(x['quantity']))
                        #     dbuy.save()
                        # compra = Compra.objects.get(
                        #         pk=request.POST.get('purchase'))
                        # det = DetCompra.objects.filter(
                        #     Q(compra_id=request.POST.get('purchase')),
                        #     Q(flag='1') | Q(flag='0')).aggregate(
                        #         counter=Count('flag'))
                        # if det['counter'] > 0:
                        #     compra.status = 'IN'
                        #     compra.save()
                        # else:
                        #     compra.status = 'CO'
                        #     compra.save()
                        context['status'] = True
                        context['ingress'] = ingress
                    else:
                        context['raise'] = 'Formato'
                        context['status'] = False
            except ObjectDoesNotExist, e:
                context['raise'] = e.__str__()
                context['status'] = False
            return self.render_to_json_response(context)

class NoteIngressView(JSONResponseMixin, TemplateView):
    template_name = 'almacen/listnoteingress.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'search' in request.GET:
                        if 'nro' in request.GET:
                            obj = NoteIngress.objects.filter(
                                    pk=request.GET.get('nro'), status='CO')
                        if 'status' in request.GET:
                            obj = NoteIngress.objects.filter(
                                    status=request.GET.get('status'))
                        if 'sdate' in request.GET:
                            if 'edate' in request.GET:
                                obj = NoteIngress.objects.filter(
                                        register__range=(
                                            globalVariable.format_str_date(
                                                request.GET.get(
                                                    'sdate'), '%d-%m-%Y'),
                                            globalVariable.format_str_date(
                                                request.GET.get('edate'),
                                                '%d-%m-%Y')))
                            else:
                                # print globalVariable.format_str_date(
                                #        request.GET.get('sdate'),'%d-%m-%Y')
                                date = globalVariable.format_str_date(
                                        request.GET.get('sdate'), '%d-%m-%Y')
                                obj = NoteIngress.objects.filter(
                                    register__startswith=date)
                        if obj:
                            context['list'] = [{
                                'ingress': x.ingress_id,
                                'purchase': x.purchase_id,
                                'invoice': x.invoice,
                                'register': x.register.strftime('%d-%m-%Y'),
                                'status': x.status}
                                for x in obj]
                            context['status'] = True
                        else:
                            context['status'] = False
                    if 'details' in request.GET:
                        bedside = NoteIngress.objects.get(
                                    pk=request.GET['ingress'])
                        context['ingress'] = bedside.ingress_id
                        context['storage'] = bedside.storage.nombre
                        context['purchase'] = bedside.purchase_id
                        context['guide'] = bedside.guide
                        context['invoice'] = bedside.invoice
                        context['motive'] = bedside.motive
                        context['observation'] = bedside.observation
                        context['details'] = [{
                            'materials': x.materials_id,
                            'name': x.materials.matnom,
                            'meter': x.materials.matmed,
                            'unit': x.materials.unidad.uninom,
                            'brand': x.brand.brand,
                            'model': x.model.model,
                            'quantity': x.quantity}
                            for x in DetIngress.objects.filter(
                                ingress_id=request.GET.get('ingress'))]
                        context['status'] = True
                except ObjectDoesNotExist, e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
            context['note'] = NoteIngress.objects.filter(
                                status='CO').order_by('-register')[:10]
            return render_to_response(
                self.template_name,
                context,
                context_instance=RequestContext(request))
        except TemplateDoesNotExist, e:
            raise Http404(e)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'editingress' in request.POST:
                    obj = NoteIngress.objects.get(
                            pk=request.POST.get('ingress'))
                    obj.guide = request.POST.get('guide')
                    obj.invoice = request.POST.get('invoice')
                    obj.motive = request.POST.get('motive')
                    obj.observation = request.POST.get('observation')
                    obj.save()
                    context['status'] = True
                if 'annularNI' in request.POST:
                    ingress = NoteIngress.objects.get(
                                pk=request.POST.get('ingress'))
                    details = DetIngress.objects.filter(
                                ingress_id=request.POST.get('ingress'))
                    for x in details:
                        inv = Inventario.objects.get(
                                materiales_id=x.materials_id,
                                periodo=ingress.register.strftime('%Y'))
                        stock = (inv.stock - x.quantity)
                        if stock >= 0:
                            inv.stock = stock
                        else:
                            inv.stock = 0
                        try:
                            brand = InventoryBrand.objects.get(
                                        materials_id=x.materials_id,
                                        period=ingress.register.strftime('%Y'),
                                        brand_id=x.brand_id,
                                        model_id=x.model_id)
                            stock = (brand.stock - x.quantity)
                            if stock >= 0:
                                brand.stock = stock
                            else:
                                brand.stock = 0
                            brand.save()
                        except ObjectDoesNotExist:
                            pass
                        inv.save()
                    ingress.status = 'AN'
                    ingress.save()
                    purchase = Compra.objects.get(
                                compra_id=ingress.purchase_id)
                    purchase.status = 'AN'
                    purchase.save()
                    context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = True
            return self.render_to_json_response(context)

class GuideSingle(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        try:
            if request.is_ajax():
                try:
                    if 'customers' in request.GET:
                        context['customers'] = json.loads(
                            serializers.serialize(
                                'json',
                                Cliente.objects.filter(flag=True)))
                        context['status'] = True
                    if 'carrier' in request.GET:
                        context['carrier'] = json.loads(
                            serializers.serialize(
                                'json',
                                Transportista.objects.filter(flag=True)))
                        context['status'] = True
                    if 'detCarrier' in request.GET:
                        context['transport'] = json.loads(
                            serializers.serialize(
                                'json',
                                Transporte.objects.filter(
                                    traruc_id=request.GET['tra'],
                                    flag=True)))
                        context['driver'] = json.loads(
                            serializers.serialize(
                                'json',
                                Conductore.objects.filter(
                                    traruc_id=request.GET['tra'],
                                    flag=True)))
                        context['status'] = True
                    if 'listTemp' in request.GET:
                        context['list'] = json.loads(
                            serializers.serialize(
                                'json',
                                TmpDetGuia.objects.filter(flag=True),
                                relations=('materials', 'brand', 'model')))
                        context['status'] = True
                    if 'brandandmodel' in request.GET:
                        context['brand'] = json.loads(
                            serializers.serialize(
                                'json',
                                Brand.objects.filter(flag=True)))
                        context['model'] = json.loads(
                            serializers.serialize(
                                'json',
                                Model.objects.filter(flag=True)))
                        context['status'] = True
                    if 'gstock' in request.GET:
                        context['stocka'] = json.loads(
                            serializers.serialize(
                                'json',
                                Inventario.objects.filter(
                                    materiales_id=request.GET['code'],
                                    periodo=globalVariable.get_year)))
                        det = InventoryBrand.objects.filter(
                            period=globalVariable.get_year,
                            materials_id=request.GET['code'])
                        context['list'] = json.loads(
                            serializers.serialize(
                                'json', det, relations=('brand', 'model')))
                        context['exact'] = json.loads(
                            serializers.serialize(
                                'json',
                                det.filter(
                                    brand_id=request.GET['brand'],
                                    model_id=request.GET['model'])))
                        context['status'] = True
                except ObjectDoesNotExist as e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
            context['address'] = request.session['company']['address']
            return render(request, 'almacen/guideSingle.html', context)
        except TemplateDoesNotExist, e:
            raise Http404(e)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'saveMaterial' in request.POST:
                    try:
                        if 'obrand' in request.POST:
                            obj = TmpDetGuia.objects.get(
                                materials_id=request.POST['materials'],
                                brand_id=request.POST['obrand'],
                                model_id=request.POST['omodel'])
                        else:
                            obj = TmpDetGuia.objects.get(
                                materials_id=request.POST['materials'],
                                brand_id=request.POST['brand'],
                                model_id=request.POST['model'])
                        form = forms.addTempGuide(request.POST, instance=obj)
                    except TmpDetGuia.DoesNotExist:
                        form = forms.addTempGuide(request.POST)
                    if form.is_valid():
                        form.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                if 'delItem' in request.POST:
                    TmpDetGuia.objects.get(
                        materials_id=request.POST['materials'],
                        brand_id=request.POST['brand'],
                        model_id=request.POST['model']).delete()
                    context['status'] = True

                if 'valid' in request.POST:
                    code = request.POST['code']
                    try:
                        if len(code) == 12:
                            if GuiaRemision.objects.get(guia_id=code):
                                context['status'] = False
                        elif len(code) > 0:
                            try:
                                code = code.split('-')
                                GuiaRemision.objects.get(
                                    guia_id='%s-%s' % (
                                        code[0], '{:0>8d}'.format(
                                            int(code[1]))))
                                context['status'] = False
                            except GuiaRemision.DoesNotExist:
                                GuiaRemision.objects.get(
                                    guia_id=request.POST['code'])
                                context['status'] = False
                        else:
                            context['status'] = False
                    except GuiaRemision.DoesNotExist, e:
                        context['status'] = True
                        context['raise'] = str(e)
                if 'genGuide' in request.POST:
                    # Generate Guide Single
                    det = TmpDetGuia.objects.filter(flag=True)
                    if det.count():
                        code = request.POST['guide']
                        if len(code) < 12:
                            code = code.split('-')
                            code = (
                                '%s-%s' % (
                                    '{:0>3d}'.format(int(code[0])),
                                    '{:0>8d}'.format(int(code[1]))))
                        form = forms.addGuideReferral(request.POST)
                        if form.is_valid():
                            add = form.save(commit=False)
                            add.guia_id = code
                            add.flag = True
                            add.status = 'GE'
                            add.observation = request.POST['observation']
                            add.nota = request.POST['nota']
                            add.save()
                            # save details
                            for x in det:
                                # get Stock Inventory of brand and model
                                inv = Inventario.objects.get(
                                        periodo=globalVariable.get_year,
                                        materiales_id=x.materials_id).order_by(
                                        '-ingreso')
                                ibm = InventoryBrand.objects.get(
                                    period=globalVariable.get_year,
                                    materials_id=x.materials_id,
                                    brand_id=x.brand_id,
                                    model_id=x.model_id).order_by(
                                        '-ingress')
                                dg = DetGuiaRemision()
                                inv.stock = (
                                    float(inv.stock) - float(x.quantity))
                                inv.save()
                                ibm.stock = (
                                    float(ibm.stock) - float(x.quantity))
                                ibm.save()
                                dg.guia_id = code
                                dg.materiales_id = x.materials_id
                                dg.cantguide = x.quantity
                                dg.brand_id = x.brand_id
                                dg.model_id = x.model_id
                                if x.observation:
                                    dg.observation = x.observation
                                dg.save()
                            det.delete()
                            context['status'] = True
                        else:
                            context['raise'] = 'fields invalid.'
                            context['status'] = False
                    else:
                        context['status'] = False
                        context['raise'] = 'No se a encontrado materiales' \
                            ' para generar la Guia Remision'
                if 'saveObs' in request.POST:
                    obs = TmpDetGuia.objects.get(
                        materials_id=request.POST['materials'],
                        brand_id=request.POST['brand'],
                        model_id=request.POST['model'])
                    obs.observation = request.POST['observation']
                    obs.save()
                    context['status'] = True
            except ObjectDoesNotExist as e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

class MaterialBrand(JSONResponseMixin, TemplateView):

    def get(self, request, *args, **kwargs):
        context = dict()
        try:
            # if request.is_ajax():
                # try:
                #     pass
                # except ObjectDoesNotExist as e:
                #     context['raise'] = str(e)
                #     context['status'] = False
                # return self.render_to_json_response(context)
            context['brand'] = InventoryBrand.objects.filter(
                                materials_id=kwargs['mid'],
                                period=globalVariable.get_year)
            return render(request, 'almacen/materialbrand.html', context)
        except TemplateDoesNotExist, e:
            raise Http404(e)

class EditBedsideGuide(JSONResponseMixin, View):

    def get(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                if 'details' in request.GET:
                    context['list'] = json.loads(serializers.serialize(
                        'json',
                        DetGuiaRemision.objects.filter(guia_id=kwargs['guide']),
                        relations=('materiales',)))
                    context['status'] = True
                return self.render_to_json_response(context)
            context['guide'] = GuiaRemision.objects.get(guia_id=kwargs['guide'])
            return render(request, 'almacen/guideedit.html', context)
        except TemplateDoesNotExist, e:
            raise Http404(e)

    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            context = dict()
            try:
                if 'saveGuide' in request.POST:
                    o = GuiaRemision.objects.get(guia_id=kwargs['guide'])
                    o.traslado = request.POST['traslado']
                    o.puntollegada = request.POST['puntollegada']
                    o.traruc_id = request.POST['traruc_id']
                    o.condni_id = request.POST['condni_id']
                    o.nropla_id = request.POST['nropla_id']
                    o.comment = request.POST['comment']
                    o.nota = request.POST['nota']
                    o.save()
                    context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

class ReturnItemOrders(JSONResponseMixin, TemplateView):
    """docstring for ReturnItemOrders"""
    def get(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'getorder' in request.GET:
                        det = Detpedido.objects.filter(pedido_id=kwargs['order'])
                        if det[0].pedido.status != 'CO' and det[0].pedido.status != 'AN':
                            det = det.exclude(cantshop=0)
                            context['details'] = json.loads(
                                serializers.serialize(
                                    'json',
                                    det,
                                    relations=('materiales','brand','model')))
                        else:
                            context['details'] = '[]'
                        context['status'] = True
                    if 'getNipples' in request.GET:
                        gnp = json.loads(request.GET['check'])
                        context['gnp'] = list()
                        for x in gnp:
                            nip = Niple.objects.filter(
                                pedido_id=kwargs['order'],
                                materiales_id=x['materials'],
                                brand_id=x['brand'],
                                model_id=x['model']).exclude(tag='2', cantshop__lte=0)
                            if nip.count():
                                context['gnp'].append(json.loads(
                                    serializers.serialize(
                                        'json',
                                        nip, relations=('materiales'))))
                        context['valid'] = True if len(context['gnp']) else False
                        context['status'] = True
                except (ObjectDoesNotExist, Exception) as e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
            context['order'] = Pedido.objects.get(pedido_id=kwargs['order'], flag=True)
            context['usr'] = userProfile.objects.get(empdni__exact=context['order'].empdni)
            return render(request, 'almacen/returnorder.html', context)
        except TemplateDoesNotExist, e:
            raise Http404(e)

    def post(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'saveReturn' in request.POST:
                    """ save changes return list to project"""
                    # validating quantity in ddbb
                    # requeriment before return materials to project
                    data = json.loads(request.POST['details'])
                    nip = json.loads(request.POST['nip'])
                    od = Pedido.objects.filter(pedido_id=kwargs['order'])
                    if len(od) > 0:
                        od = od[0]
                    for d in data:
                        valid = False
                        nq = 0
                        for n in nip:
                            if d['materials'] in n:
                                for x in n[d['materials']]:
                                    nq += (float(x['import']) / 100)
                                    # part return quantity at table nipple
                                    try:
                                        np = Nipple.objects.filter(
                                                proyecto_id=od.proyecto_id,
                                                sector_id=od.sector_id,
                                                materiales_id=x['materials'],
                                                metrado=x['meter'],
                                                tipo=x['type'])
                                        if np.count() > 0:
                                            np = np[0]
                                            np.cantshop += float(x['quantity'])
                                            np.tag = '1'
                                            np.save()
                                        # remove or decrease
                                        # items from orders nipple
                                        rno = Niple.objects.get(id=x['id'])
                                        rno.cantshop += float(x['quantity'])
                                        rno.cantidad -= float(x['quantity'])
                                        if rno.cantidad <= 0:
                                            rno.delete()
                                        else:
                                            rno.save()
                                    except Exception as e:
                                        print 'NIPPLE DONT EXISTS'
                                        print str(e)
                                valid = True
                        if valid is True:
                            d['quantity'] = round(nq, 2)
                    notProc = list()
                    for x in data:
                        valid = True
                        queryset = Detpedido.objects.get(id=x['id'], pedido_id=kwargs['order'])
                        print '---------------------------'
                        print 'HERE QUERY ', queryset
                        print '---------------------------'
                        print queryset.cantshop, 'CANTIDAD PROCESS ', x['quantity']
                        if queryset.cantshop < float(x['quantity']):
                            valid = False
                        if valid:
                            if queryset.cantidad == x['quantity']:
                                print 'ITEM DELETE'
                                queryset.delete()
                            elif queryset.cantshop == x['quantity']:
                                print 'ITEM COMPLETE'
                                queryset.cantshop = (queryset.cantshop - x['quantity'])
                                queryset.cantidad = (queryset.cantidad - x['quantity'])
                                queryset.tag = '2'
                                queryset.save()
                            if x['quantity'] < queryset.cantshop:
                                print 'ITEM MEDIO'
                                queryset.cantshop = (queryset.cantshop - x['quantity'])
                                queryset.cantidad = (queryset.cantidad - x['quantity'])
                                queryset.save()
                        else:
                            print 'ITEM NOT FOUND'
                            notProc.append(x)
                    # consult and update status order
                    order = Pedido.objects.get(pedido_id=kwargs['order'])
                    print order
                    qdp = Detpedido.objects.filter(pedido_id=kwargs['order'])
                    if qdp.filter(cantshop__gt=0).count():
                        if qdp.count() == qdp.filter(cantshop__gt=0).count():
                            order.status = 'AP'
                        else:
                            order.status = 'IN'
                    if qdp.filter(cantshop=0).count() == qdp.count():
                        order.status = 'CO'
                    order.save()
                    # register in table ReturnItemsProject
                    obj = ReturnItemsProject()
                    obj.pedido_id = kwargs['order']
                    obj.observation = request.POST['observation']
                    obj.listsend = request.POST['details']
                    obj.notpro = json.dumps(notProc)
                    obj.save()
                    # return materials to project
                    opr = None
                    # print request.POST
                    if isinstance(order.dsector_id, str) and len(str(order.dsector_id).strip()) == 10:
                        # this part return item to area table operations "dsmetrado"
                        print 'table dsmetrado'
                        for x in data:
                            try:
                                opr =  dsmetrado.objects.get(
                                    dsecto_id=order.dsecto_id,
                                    materials_id=x['materials'],
                                    brand_id=x['brand_id'],
                                    model_id=x['model_id'])
                                opr.qorder += float(x['quantity'])
                                if opr.quantity > opr.qorder:
                                    opr.tag = '1'
                                else:
                                    opr.tag = '0'
                                opr.save()
                            except Exception, e:
                                print e
                                continue
                    else:
                        print 'table metProject'
                        # this part return item to project for table operations "metproject"
                        for x in data:
                            try:
                                opr = MetProject.objects.get(
                                    sector_id=order.sector_id,
                                    materiales_id=x['materials'],
                                    brand_id=x['brand_id'],
                                    model_id=x['model_id'])
                                opr.quantityorder += x['quantity']
                                if opr.cantidad > opr.quantityorder:
                                    opr.tag = '1'
                                else:
                                    opr.tag = '0'
                                opr.save()
                            except Exception, e:
                                print e
                                continue
                    context['status'] = True
            except (ObjectDoesNotExist, Exception), e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

class AttendOrder(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        try:
            if request.is_ajax():
                try:
                    if 'details' in request.GET:
                        det = Detpedido.objects.filter(
                                    pedido_id=kwargs['order']).extra(
                                    select={'stock': "SELECT stock FROM almacen_inventario WHERE almacen_detpedido.materiales_id LIKE almacen_inventario.materiales_id"})
                        context['details'] = [{
                            'pk': x.id,
                            'fields': {
                                'materiales': {
                                    'pk': x.materiales_id,
                                    'fields': {
                                        'matnom': x.materiales.matnom,
                                        'matmed': x.materiales.matmed,
                                        'unidad': x.materiales.unidad.uninom
                                    }
                                },
                                'brand': {
                                    'pk': x.brand_id,
                                    'fields': {
                                        'brand': x.brand.brand,
                                    }
                                },
                                'model': {
                                    'pk': x.model_id,
                                    'fields': {
                                        'model': x.model.model
                                    }
                                },
                                'cantidad': x.cantidad,
                                'cantshop': x.cantshop,
                                'tag': x.tag,
                                'stock': x.stock
                                }
                            } for x in det]
                        context['status'] = True
                    if 'detailsnip' in request.GET:
                        context['nip'] = json.loads(
                            serializers.serialize(
                                'json',
                                Niple.objects.filter(
                                    pedido_id=kwargs['order']), relations=('materiales', 'brand', 'model')))
                        context['types'] = globalVariable.tipo_nipples
                        context['status'] = True
                    if 'gstock' in request.GET:
                        # here obtend stock
                        context['stock'] = json.loads(
                            serializers.serialize(
                                'json',
                                InventoryBrand.objects.filter(
                                    storage_id='AL01',
                                    #period=globalVariable.get_year,
                                    materials_id=request.GET['materials']),
                                relations=('materials','brand','model')))
                        context['status'] = True
                    if 'validNumber' in request.GET:
                        print request.GET
                        import re
                        nguia = request.GET['guia']
                        mt = re.match('[0-9]{3}[-]{1}[0-9]{8}$', nguia)
                        # print
                        if mt:
                            guia = GuiaRemision.objects.filter(guia_id=nguia)
                            if len(guia) > 0:
                                guia = guia[0]
                                # verify pedido is self project
                                ord = Pedido.objects.get(pedido_id=kwargs['order'])
                                if ord.proyecto_id == guia.pedido.proyecto_id:
                                    context['status'] = True
                                else:
                                    context['status'] = False
                                    context['raise'] = 'La guia pertenece a otro proyecto'
                            else:
                                context['status'] = True
                        else:
                            context['raise'] = 'Formato Invalido.'
                            context['status'] = False
                except ObjectDoesNotExist, e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
            else:
                context['order'] = Pedido.objects.get(pedido_id=kwargs['order'])
                return render(request, 'almacen/attends/vieworderattend.html', context)
        except TemplateDoesNotExist as e:
            raise Http404(e)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        context = dict()
        try:
            if 'generateGuide' in request.POST:
                # Here get data
                # save bedside Guide Remision
                code = str(request.POST['nro']).strip()
                code = code.split('-')
                code = '%s-%s' % ('{:0>3d}'.format(int(code[0])), '{:0>8d}'.format(int(code[1])))
                guide = GuiaRemision(
                    guia_id=code,
                    pedido_id=kwargs['order'],
                    ruccliente_id=request.POST['customer'],
                    registrado=datetime.datetime.today(),
                    puntollegada=request.POST['dotdelivery'],
                    traslado=request.POST['transfer'],
                    traruc_id=request.POST['carrier'],
                    condni_id=request.POST['conductor'],
                    nropla_id=request.POST['transport'],
                    status='GE',
                    motive='VENTA',
                    orders=kwargs['order'],
                    perreg_id=request.user.get_profile().empdni_id)
                guide.save()
                # here save details guide
                det = json.loads(request.POST['details'])
                print 'DETALLE DE GUIA', det
                for d in det:
                    for x in d['details']:
                        dg = DetGuiaRemision(
                            guia_id=code,
                            materiales_id=x['materials'],
                            cantguide=x['quantity'],
                            flag=True,
                            brand_id=x['brand'],
                            model_id=x['model'],
                            observation=x['observation'] if 'observation' in x else '',
                            order_id=kwargs['order'],
                            obrand_id=d['brand'],
                            omodel_id=d['model'])
                        dg.save()
                nip = json.loads(request.POST['nipp'])
                print 'NIPPLES', nip
                if len(nip) > 0:
                    for n in nip:
                        for x in n['details']:
                            if float(x['guide']) > 0:
                                ng = NipleGuiaRemision(
                                    guia_id=code,
                                    materiales_id=x['materials'],
                                    metrado=x['meter'],
                                    cantguide=x['guide'],
                                    tipo=x['tipo'],
                                    flag=True,
                                    brand_id=x['brand'],
                                    model_id=x['model'],
                                    related=x['id'],
                                    order_id=kwargs['order'])
                                ng.save()
                context['code'] = code
                context['status'] = True
        except (ObjectDoesNotExist, Exception) as e:
            context['raise'] = str(e)
            context['status'] = False
        return self.render_to_json_response(context)

class LoadInventoryBrand(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        try:
            if request.is_ajax():
                try:
                    if 'getmat' in request.GET:
                        context['materials'] = json.loads(
                            serializers.serialize(
                            'json',
                            Inventario.objects.filter(
                                Q(materiales_id=request.GET['desc']) |
                                Q(
                                materiales__matnom__icontains=request.GET[
                                    'desc']) | Q(
                                    materiales__matmed__icontains=request.GET[
                                    'desc'])).order_by('materiales__matmed'),
                                relations=('materiales')))
                        context['status'] = True
                    if 'details' in request.GET:
                        ib = InventoryBrand.objects.filter(materials_id=request.GET['materials'])
                        context['materials'] = json.loads(
                            serializers.serialize(
                                'json',
                                ib.order_by('brand__brand'),
                                    relations=('materials', 'brand', 'model')))
                        context['amount'] = ib.aggregate(Sum('stock'))['stock__sum']
                        context['status'] = True
                except (ObjectDoesNotExist, Exception) as e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
            else:
                return render(request, 'almacen/uploadStock.html', context)
        except TemplateDoesNotExist as e:
            raise Http404(e)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'delInventory' in request.POST:
                    erase = InventoryBrand.eraseAllInventory()
                    # print erase
                    context['raise'] = erase
                    context['status'] = erase
                if 'loadInventory' in request.POST:
                    # print request.FILES
                    path = '/storage/Temp/'
                    options = {'name': 'storage'}
                    filename = uploadFiles.upload(
                                path,
                                request.FILES['fload'],
                                options)
                    if uploadFiles.fileExists(filename):
                        wb = load_workbook(filename=filename, read_only=True)
                        # print filename
                        ws = wb.worksheets[0]
                        for x in range(1, (ws.max_row + 1)):
                            print x
                            cell = str('%s'%ws.cell(row=x, column=1).value).encode('utf-8')
                            if len(cell) >= 14 and cell != None:
                                stk = float(str('%s'%ws.cell(row=x, column=7).value).encode('utf-8'))
                                if float(str(stk)) <= 0 or stk == None:
                                    continue
                                else:
                                    # get brand and model
                                    brd = str('%s'%ws.cell(row=x, column=5).value).encode('utf-8').upper().strip()
                                    if brd == '-' or brd == "":
                                        brd = 'BR000'
                                    else:
                                        sbrd = Brand.objects.filter(brand=brd)
                                        if len(sbrd):
                                            brd = sbrd[0].brand_id
                                        else:
                                            b = Brand(
                                                brand_id=genkeys.GenerateIdBrand(),
                                                brand=brd)
                                            b.save()
                                            brd = b.brand_id
                                    mdl = str('%s'%ws.cell(row=x, column=4).value).encode('utf-8').upper().strip()
                                    if mdl == '-' or mdl == "":
                                        mdl = 'MO000'
                                    else:
                                        smdl = Model.objects.filter(model=mdl)
                                        if len(smdl):
                                            mdl = smdl[0].model_id
                                        else:
                                            b = Model(
                                                model_id=genkeys.GenerateIdModel(),
                                                brand_id=brd,
                                                model=mdl)
                                            b.save()
                                            mdl = b.model_id
                                    # evaluate code exists
                                    ib = InventoryBrand.objects.filter(
                                        materials_id=cell,
                                        brand_id=brd,
                                        model_id=mdl)
                                    if len(ib):
                                        ib = ib[0]
                                        ib.stock = (ib.stock + stk)
                                        ib.save()
                                    else:
                                        ppurchase = float(ws.cell(row=x, column=8).value)
                                        psales = float(ws.cell(row=x, column=9).value)
                                        ib = InventoryBrand(
                                            storage_id='AL01',
                                            period=globalVariable.get_year,
                                            materials_id=cell,
                                            brand_id=brd,
                                            model_id=mdl,
                                            stock=stk,
                                            purchase=ppurchase,
                                            sale=psales)
                                        ib.save()
                        uploadFiles.removeTmp(filename)
                    context['status'] = True
                if 'putorder' in request.POST:
                    dt = DetPedido.objects.filter(pedido_id__in=request.POST['order'].split(','))
                    for x in dt:
                        ib = InventoryBrand.objects.filter(
                            materials_id=x.materiales_id,
                            brand_id='BR000',
                            model_id='MO000')
                        if len(ib):
                            ib = ib[0]
                            ib.stock = (ib.stock + x.cantidad)
                            ib.save()
                        else:
                            ib = InventoryBrand(
                                storage_id='AL01',
                                period=globalVariable.get_year,
                                materials_id=x.materiales_id,
                                brand_id='BR000',
                                model_id='MO000',
                                stock=x.cantidad,
                                purchase=1,
                                sale=1)
                            ib.save()
                    context['status'] = True
            except (ObjectDoesNotExist, Exception) as e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)
