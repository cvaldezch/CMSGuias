#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
import json

from django.core.exceptions import ObjectDoesNotExist
from django.core import serializers
# from django.contrib import messages
# from django.contrib.auth.mod import User
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse
from django.db.models import Q, Sum, Avg
from django.shortcuts import render_to_response, render
from django.utils import simplejson
from django.utils.decorators import method_decorator
from django.template import RequestContext, TemplateDoesNotExist
from django.views.generic import TemplateView, View
from django.core.serializers.json import DjangoJSONEncoder
# from xlrd import open_workbook, XL_CELL_EMPTY
from openpyxl import load_workbook, Workbook, cell, styles
import time

from CMSGuias.apps.home.models import *
from .models import *
from .forms import *
from CMSGuias.apps.almacen.models import *
from CMSGuias.apps.ventas.models import Metradoventa, SectorFiles
from CMSGuias.apps.tools import genkeys, uploadFiles


# Class Bases Views Generic

class JSONResponseMixin(object):

    def render_to_json_response(self, context, **response_kwargs):
        return HttpResponse(
            self.convert_context_to_json(context),
            content_type='application/json',
            mimetype='application/json',
            **response_kwargs
        )

    def convert_context_to_json(self, context):
        return simplejson.dumps(context,
                                encoding='utf-8',
                                cls=DjangoJSONEncoder)


# View home Operations
class OperationsHome(TemplateView):
    template_name = 'operations/home.html'

    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(OperationsHome, self).dispatch(request, *args, **kwargs)


# View list pre orders
class ListPreOrders(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            context = dict()
            context['list'] = PreOrders.objects.filter(
                project_id=kwargs['pro'],
                sector_id=kwargs['sec'],
                status='PE'
            ).order_by('-register')
            context['status'] = globalVariable.status
            return render_to_response(
                'operations/listpreorders.html',
                context,
                context_instance=RequestContext(request))
        except TemplateDoesNotExist, e:
            raise Http404(e)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'anullarPreOrders' in request.POST:
                    obj = PreOrders.objects.get(
                        preorder_id=request.POST['pre'])
                    obj.annular = request.POST.get('annular')
                    obj.status = 'AN'
                    obj.save()
                    context['status'] = True
                if 'changeComplete' in request.POST:
                    obj = PreOrders.objects.get(
                        preorder_id=request.POST['pre'])
                    obj.status = 'CO'
                    obj.save()
                    context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)
        else:
            context['list'] = PreOrders.objects.filter(
                project_id=kwargs['pro'],
                sector_id=kwargs['sec'],
                status=request.POST.get('status')
            ).order_by('-register')
            context['search'] = request.POST.get('status')
            context['status'] = globalVariable.status
            return render_to_response(
                'operations/listpreorders.html',
                context,
                context_instance=RequestContext(request))


class ProgramingProject(JSONResponseMixin, View):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'listg' in request.GET:
                    sg = SGroup.objects.filter(
                            project_id=kwargs['pro'],
                            subproject_id=kwargs[
                                'sub'] if unicode(kwargs[
                                    'sub']) != 'None' and unicode(kwargs[
                                        'sub']) != '' else None,
                            sector_id=kwargs['sec']).order_by('name')
                    context['sg'] = json.loads(serializers.serialize(
                                                                'json', sg))
                    context['status'] = True
                if 'listds' in request.GET:
                    ds = DSector.objects.filter(
                            project_id=kwargs['pro'],
                            sgroup__subproject_id=kwargs[
                                'sub'] if unicode(kwargs[
                                    'sub']) != 'None' and unicode(kwargs[
                                        'sub']) != '' else None,
                            sgroup__sector_id=kwargs['sec']).order_by(
                            'sgroup')
                    context['ds'] = json.loads(serializers.serialize(
                                        'json',
                                        ds,
                                        relations=('sgroup',)))
                    context['status'] = True
                if 'getAreasByGroup' in request.GET:
                    ds = DSector.objects.filter(
                            project_id=kwargs['pro'],
                            sgroup_id=request.GET['sgroup'],
                            sgroup__sector_id=kwargs['sec']).order_by(
                            'sgroup')
                    context['areas'] = json.loads(serializers.serialize(
                                        'json',
                                        ds,
                                        relations=('sgroup',)))
                    context['status'] = True
                if 'valPrices' in request.GET:
                    # get sgroup
                    sg = [x[0] for x in SGroup.objects.filter(
                            project_id=kwargs['pro'],
                            subproject_id=kwargs[
                                'sub'] if unicode(kwargs[
                                    'sub']) != 'None' and unicode(kwargs[
                                        'sub']) != '' else None,
                            sector_id=kwargs['sec']).values_list('sgroup_id',)]
                    sec = [x[0] for x in DSector.objects.filter(
                            sgroup_id__in=sg).values_list('dsector_id')]
                    without = DSMetrado.objects.filter(
                        Q(dsector_id__in=sec),
                        Q(ppurchase=0) | Q(psales=0) | Q(quantity=0)
                        ).order_by('materials__matnom').distinct(
                        'materials__matnom')
                    if without:
                        context['list'] = json.loads(
                            serializers.serialize(
                                'json',
                                without,
                                relations=('materials',),
                                indent=4))
                        context['status'] = True
                    else:
                        context['status'] = False
            except ObjectDoesNotExist as e:
                context['raise'] = str(e)
                context['status'] = Falses
            return self.render_to_json_response(context)
        else:
            try:
                context['project'] = Proyecto.objects.get(
                    proyecto_id=kwargs['pro'])
                context['sector'] = Sectore.objects.get(
                    proyecto_id=kwargs['pro'],
                    subproyecto_id=kwargs['sub'] if kwargs[
                        'sub'] is None else None,
                    sector_id=kwargs['sec'])
                ds = DSector.objects.filter(
                        project_id=kwargs['pro'],
                        sector_id=kwargs['sec'])
                t = ds.count()
                if t == 0:
                    context['status'] = 'PE'
                else:
                    ts = ds.filter(status='AC').count()
                    context['status'] = 'AC' if t == ts else 'PE'
                context['sales'] = MetProject.objects.filter(
                    proyecto_id=kwargs['pro'],
                    sector_id=kwargs['sec']).aggregate(
                        amount=Sum(
                            'cantidad', field='cantidad*precio'))['amount']
                if ds:
                    context['operations'] = DSMetrado.objects.filter(
                        dsector_id__in=[x.dsector_id for x in ds]).aggregate(
                        amount=Sum(
                            'quantity',
                            field='quantity*ppurchase'))['amount']
                    if context['operations'] <= 0 or context['operations'] == None:
                        context['operations'] = 0
                    context['diff'] = (context[
                                        'sector'].amount-context['operations'])
                return render(
                    request,
                    'operations/programinggroup.html',
                    context)
            except TemplateDoesNotExist, e:
                raise Http404(e)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'saveg' in request.POST:
                    try:
                        if 'sgroup_id' in request.POST:
                            sg = SGroup.objects.get(
                                    sgroup_id=request.POST['sgroup_id'])
                            form = SGroupForm(request.POST, instance=sg)
                        else:
                            form = SGroupForm(request.POST)
                    except ObjectDoesNotExist:
                        form = SGroupForm(request.POST)
                    print form, form.is_valid()
                    if form.is_valid():
                        if 'sgroup_id' not in request.POST:
                            add = form.save(commit=False)
                            key = genkeys.genSGroup(
                                    kwargs['pro'], kwargs['sec'])
                            add.sgroup_id = key.strip()
                            add.project_id = kwargs['pro']
                            if unicode(kwargs['sub']) != 'None':
                                add.subproject_id = kwargs['sub']
                            add.sector_id = kwargs['sec']
                            add.colour = request.POST['rgba']
                            add.save()
                        else:
                            edit = form.save(commit=False)
                            edit.colour = request.POST['rgba']
                            edit.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                if 'saveds' in request.POST:
                    try:
                        if 'dsector_id' in request.POST:
                            ds = DSector.objects.get(
                                    dsector_id=request.POST['dsector_id'])
                            form = DSectorForm(
                                    request.POST, request.FILES, instance=ds)
                        else:
                            form = DSectorForm(request.POST, request.FILES)
                    except ObjectDoesNotExist as e:
                        form = DSectorForm(request.POST, request.FILES)
                    if form.is_valid():
                        if 'dsector_id' not in request.POST:
                            add = form.save(commit=False)
                            key = genkeys.genDSector(
                                    kwargs['pro'],
                                    request.POST['sgroup'])
                            add.dsector_id = key.strip()
                            add.project_id = kwargs['pro']
                            add.sector_id = kwargs['sec']
                            add.save()
                        else:
                            form.save()
                        try:
                            sg = SGroup.objects.get(
                                    sgroup_id=request.POST['sgroup'])
                            if sg.datestart is None:
                                sg.datestart = request.POST['datestart']
                            else:
                                st = globalVariable.format_str_date(
                                        request.POST['datestart'])
                                if st < sg.datestart:
                                    sg.datestart = st
                            if sg.dateend is None:
                                sg.dateend = request.POST['dateend']
                            else:
                                ed = globalVariable.format_str_date(
                                        request.POST['dateend'])
                                if ed > sg.dateend:
                                    sg.dateend = ed
                            sg.save()
                        except SGroup.DoesNotExist, e:
                            context['raise'] = str(e)
                        context['status'] = True
                    else:
                        context['status'] = False
                        context['raise'] = 'Fields empty'
                if 'savePricewithout' in request.POST:
                    sg = [x[0] for x in SGroup.objects.filter(
                            project_id=kwargs['pro'],
                            subproject_id=kwargs[
                                'sub'] if unicode(kwargs[
                                    'sub']) != 'None' and unicode(kwargs[
                                        'sub']) != '' else None,
                            sector_id=kwargs['sec']).values_list('sgroup_id',)]
                    sec = [x[0] for x in DSector.objects.filter(
                            sgroup_id__in=sg).values_list('dsector_id')]
                    dsm = DSMetrado.objects.filter(
                        dsector_id__in=sec,
                        materials_id=request.POST['materials'])
                    if request.POST['field'] == 'purchase':
                        dsm.update(ppurchase=request.POST['value'])
                    if request.POST['field'] == 'sales':
                        dsm.update(psales=request.POST['value'])
                    context['status'] = True
                if 'approvedAreas' in request.POST:
                    # print kwargs['pro'], kwargs['sec']
                    sg = SGroup.objects.filter(
                            project_id=kwargs['pro'],
                            sector_id=kwargs['sec'])
                    for x in sg:
                        x.status = 'AC'
                        x.save()
                    ds = DSector.objects.filter(
                            project_id=kwargs['pro'],
                            sector_id=kwargs['sec'])
                    for x in ds:
                        x.status = 'AC'
                        x.save()
                    context['status'] = True
                if 'DiscapprovedAreas' in request.POST:
                    sg = SGroup.objects.filter(
                            project_id=kwargs['pro'],
                            sector_id=kwargs['sec'])
                    for x in sg:
                        x.status = 'PE'
                        x.save()
                    ds = DSector.objects.filter(
                            project_id=kwargs['pro'],
                            sector_id=kwargs['sec'])
                    for x in ds:
                        x.status = 'PE'
                        x.save()
                    context['status'] = True
                if 'uploadFile' in request.POST:
                    path = '/storage/Temp/'
                    opt = {'name': 'tmpa%s' % kwargs['pro']}
                    filename = uploadFiles.upload(
                        path,
                        request.FILES['upload'],
                        opt)
                    context['name'] = filename
                    context['status'] = uploadFiles.fileExists(filename)
                if 'processData' in request.POST:
                    # import time
                    print 'kwargs'
                    print kwargs
                    print 'print kwargs', kwargs
                    print 'this is sector id'
                    print kwargs['sec']
                    book = load_workbook(request.POST['filename'])
                    sheet = book['AREAS']
                    nrow = sheet.max_row
                    ncol = sheet.max_column
                    print 'row and col', nrow, ncol
                    sgroup = dict()
                    sec = kwargs['sec']
                    for x in range(1, nrow + 1):
                        print x, '-----------------'
                        if x == 2:
                            for c in range(4, ncol):
                                name = sheet.cell(row=x, column=c).value
                                if c >= 4 and name != None:
                                    # crea los grupos
                                    # print sheet.cell(row=x, column=c).value
                                    name = name.upper().strip()
                                    print name
                                    try:
                                        sg = SGroup.objects.filter(
                                                project_id=kwargs['pro'],
                                                sector_id=kwargs['sec'],
                                                name__exact=name)
                                        print 'COUNTER SGROUP', sg.count()
                                        if sg.count() > 0:
                                            sgroup[name] = {'id': sg[0].sgroup_id}
                                        else:
                                            print "CREA NUEVO GROUP DEFECT"
                                            nw = SGroup()
                                            sgroup[name] = {'id': genkeys.genSGroup(kwargs['pro'], sec)}
                                            # time.sleep(1)
                                            nw.sgroup_id = sgroup[name]['id']
                                            nw.project_id = kwargs['pro']
                                            nw.subproject_id = kwargs['sub'] if unicode(kwargs['sub']) != 'None' else ''
                                            nw.sector_id = kwargs['sec']
                                            nw.name = name
                                            nw.colour = 'rgba(254,255,180,0.8)'
                                            nw.status = 'PE'
                                            nw.save()
                                    except (ObjectDoesNotExist, Exception) as e:
                                        print "NEW GROUP By EXCEPT"
                                        nw = SGroup()
                                        sgroup[name] = {'id': genkeys.genSGroup(kwargs['pro'], sec)}
                                        # time.sleep(1)
                                        nw.sgroup_id = sgroup[name]['id']
                                        nw.project_id = kwargs['pro']
                                        nw.subproject_id = kwargs['sub'] if unicode(kwargs['sub']) != 'None' else ''
                                        nw.sector_id = kwargs['sec']
                                        nw.name = name
                                        nw.colour = 'rgba(254,255,180,0.8)'
                                        nw.status = 'PE'
                                        nw.save()
                                    time.sleep(4)
                        elif x == 3:
                            tng = None
                            group = None
                            for c in range(4, ncol):
                                if c > 3:
                                    print c, ncol
                                    if c == ncol:
                                        continue
                                    # crea las areas
                                    # print sheet.cell(row=x, column=c).value
                                    name = sheet.cell(row=2, column=c).value
                                    if name != None:
                                        name = name.upper().strip()
                                        group = sgroup[name]['id']
                                        tng = str(name).upper().strip()

                                    ds = genkeys.genDSector(
                                        kwargs['pro'], group)
                                    dsn = sheet.cell(row=x, column=c).value
                                    dsn = dsn.upper().strip()
                                    if str(dsn).strip().upper() == 'TOTAL':
                                        break
                                    sgroup[tng].update({dsn: {'id': ds}})
                                    nds = DSector()
                                    nds.dsector_id = sgroup[tng][dsn]['id']
                                    nds.sgroup_id = group
                                    nds.project_id = kwargs['pro']
                                    nds.sector_id = kwargs['sec']
                                    nds.name = dsn
                                    nds.datestart = globalVariable.date_now()
                                    nds.dateend = globalVariable.date_now()
                                    nds.description = ''
                                    nds.status = 'PE'
                                    nds.save()
                        elif x > 3:
                            tgn = None
                            for c in range(4, ncol):
                                if c == ncol:
                                    continue
                                name = sheet.cell(row=2, column=c).value
                                if name != None:
                                    tgn = name
                                    tgn = name.upper().strip()
                                cell = sheet.cell(row=x, column=c).value
                                if cell == None or float(str(cell)) == 0:
                                    continue
                                else:
                                    cell = float(str(cell))
                                    cm = str(sheet.cell(row=x, column=1).value)
                                    if len(cm) == 15:
                                        # ingreso de materiales
                                        ns = sheet.cell(row=3, column=c).value
                                        ns = ns.upper().strip()
                                        purchase = 0
                                        sales = 0
                                        try:
                                            dsm = DSMetrado.objects.filter(
                                                materials_id=str(cm)).order_by(
                                                    '-dsector__register')
                                            if len(dsm):
                                                dsm = dsm[0]
                                                purchase = float(dsm.ppurchase)
                                                sales = float(dsm.psales)
                                        except DSMetrado.DoesNotExist:
                                            raise e
                                        if purchase == 0 or sales == 0:
                                            try:
                                                dsm = MetProject.objects.filter(
                                                    materiales_id=str(cm)).order_by(
                                                    '-proyecto__registrado')
                                                if len(dsm):
                                                    dsm = dsm[0]
                                                    purchase = float(dsm.precio)
                                                    sales = float(dsm.sales)
                                            except MetProject.DoesNotExist:
                                                raise e

                                        try:
                                            odsm = DSMetrado.objects.filter(dsector_id=sgroup[tgn][ns]['id'],
                                                sector_id=kwargs['sec'],
                                                materials_id=cm, brand_id='BR000',
                                                model_id='MO000')
                                            if len(odsm) > 0:
                                                odsm = odsm[0]
                                                odsm.quantity += cell
                                                odsm.qorder += cell
                                                odsm.ppurchase = purchase
                                                odsm.psales = sales
                                                if odsm.tag == '2':
                                                    odsm.tag = '1'
                                                odsm.save() 
                                            else:
                                                dm = DSMetrado()
                                                dm.dsector_id = sgroup[tgn][ns]['id']
                                                dm.sector_id = kwargs['sec']
                                                dm.materials_id = cm
                                                dm.brand_id = 'BR000'
                                                dm.model_id = 'MO000'
                                                dm.quantity = cell
                                                dm.qorder = cell
                                                dm.qguide = 0
                                                dm.ppurchase = purchase
                                                dm.psales = sales
                                                dm.tag = '0'
                                                dm.save()
                                        except (ObjectDoesNotExist, Exception) as e:
                                            dm = DSMetrado()
                                            dm.dsector_id = sgroup[tgn][ns]['id']
                                            dm.sector_id = kwargs['sec']
                                            dm.materials_id = cm
                                            dm.brand_id = 'BR000'
                                            dm.model_id = 'MO000'
                                            dm.quantity = cell
                                            dm.qorder = cell
                                            dm.qguide = 0
                                            dm.ppurchase = purchase
                                            dm.psales = sales
                                            dm.tag = '0'
                                            dm.save()
                                    else:
                                        continue
                    context['status'] = True
                if 'delarea' in request.POST:
                    DSMetrado.objects.filter(
                        dsector_id=request.POST['ds']).delete()
                    DSector.objects.get(dsector_id=request.POST['ds']).delete()
                    context['status'] = True
                if 'delsgroup' in request.POST:
                    try:
                        ds = DSector.objects.filter(
                                project_id=kwargs['pro'],
                                sgroup_id=request.POST['sgroup'])
                        ds = [x.dsector_id for x in ds]
                        DSMetrado.objects.filter(
                            dsector_id__in=ds).delete()
                        DSector.objects.filter(dsector_id__in=ds).delete()
                    except DSector.DoesNotExist as e:
                        context['raise'] = str(e)
                    SGroup.objects.get(
                            project_id=kwargs['pro'],
                            subproject_id=kwargs[
                                'sub'] if unicode(kwargs[
                                    'sub']) != 'None' and unicode(kwargs[
                                        'sub']) != '' else None,
                            sector_id=kwargs['sec'],
                            sgroup_id=request.POST['sgroup']).delete()
                    context['status'] = True
            except ObjectDoesNotExist as e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)


class AreaProjectView(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'lplanes' in request.GET:
                    context['lplane'] = json.loads(serializers.serialize(
                        'json',
                        SectorFiles.objects.filter(dsector_id=kwargs['area'])))
                    context['status'] = True
                if 'dslist' in request.GET:
                    context['list'] = json.loads(serializers.serialize(
                        'json',
                        DSMetrado.objects.filter(
                            dsector_id=kwargs['area']).order_by(
                            'materials__matnom'), relations=(
                            'materials', 'brand', 'model', 'dsector')))
                    context['status'] = True
                if 'lstnipp' in request.GET:
                    context['nip'] = json.loads(
                        serializers.serialize(
                            'json',
                            Nipple.objects.filter(
                                area_id=kwargs['area'],
                                materiales_id=request.GET['materials']),
                            relations=('materiales')))
                    context['dnip'] = globalVariable.tipo_nipples
                    context['status'] = True
                if 'typeNipple' in request.GET:
                    context['type'] = globalVariable.tipo_nipples
                    context['status'] = True
                if 'modifyList' in request.GET:
                    context['modify'] = json.loads(
                        serializers.serialize(
                            'json',
                            MMetrado.objects.filter(
                                dsector_id=kwargs['area']).order_by(
                                'materials__matnom'),
                            relations=('materials', 'brand', 'model')))
                    context['status'] = True
                if 'samountp' in request.GET:
                    context['sec'] = json.loads(serializers.serialize(
                        'json',
                        [Sectore.objects.get(sector_id=kwargs['sec'])]))
                    sg = [x[0] for x in SGroup.objects.filter(
                            project_id=kwargs['pro'],
                            sector_id=kwargs['sec']).values_list('sgroup_id',)]
                    sec = [x[0] for x in DSector.objects.filter(
                            sgroup_id__in=sg).values_list('dsector_id')]
                    dsal = DSMetrado.objects.filter(
                            dsector_id__in=sec).aggregate(
                            tpurchase=Sum(
                                'quantity', field='quantity * ppurchase'),
                            tsales=Sum('quantity', field='quantity * psales'))
                    rds = DSMetrado.objects.filter(dsector_id=kwargs[
                        'area']).aggregate(tpurchase=Sum(
                            'quantity',
                            field='quantity * ppurchase'),
                            tsales=Sum('quantity', field='quantity * psales'))
                    print rds
                    context['msector'] = dsal
                    context['maarea'] = rds
                    mm = MMetrado.objects.filter(
                        dsector_id=kwargs['area']).order_by(
                        'materials__matnom').aggregate(
                        apurchase=Sum(
                            'quantity', field='quantity * ppurchase'),
                        asale=Sum(
                            'quantity', field='quantity * psales'))
                    context['mmodify'] = mm
                    context['status'] = True
                if 'nippleOrders' in request.GET:
                    context['nipple'] = json.loads(
                        serializers.serialize(
                            'json',
                            Nipple.objects.filter(
                                area_id=kwargs['area'],
                                materiales_id=request.GET['materials'],
                                cantshop__gt=0),
                            relations=('materiales')))
                    context['status'] = True
            except ObjectDoesNotExist as e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)
        else:
            try:
                context['dsector'] = DSector.objects.get(
                                        dsector_id=kwargs['area'])
                context['modify'] = MMetrado.objects.filter(
                                        dsector_id=kwargs['area']).order_by(
                                            '-register')[:5]
                if request.user.get_profile().empdni.charge.area.lower() == 'ventas' or request.user.get_profile().empdni.charge.area .lower() == 'administrator':
                    context['amount'] = DSMetrado.objects.filter(
                        dsector_id=kwargs['area']).aggregate(apurchase=Sum(
                            'quantity', field='quantity*ppurchase'),
                            asales=Sum('quantity', field='quantity*psales'))
                context['storage'] = Almacene.objects.filter(flag=True)
                return render(request, 'operations/dsector.html', context)
            except TemplateDoesNotExist as e:
                raise Http404(e)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'savepmat' in request.POST:
                    try:
                        if 'editmat' in request.POST:
                            dsm = DSMetrado.objects.get(
                                dsector_id=kwargs['area'],
                                materials_id=request.POST['code'],
                                brand_id=request.POST['obrand'],
                                model_id=request.POST['omodel'])
                            dsm.brand_id = request.POST['brand']
                            dsm.model_id = request.POST['model']
                            dsm.quantity = float(request.POST['quantity'])
                            dsm.qorder = float(request.POST['quantity'])
                        else:
                            dsm = DSMetrado.objects.get(
                                dsector_id=kwargs['area'],
                                materials_id=request.POST['code'],
                                brand_id=request.POST['brand'],
                                model_id=request.POST['model'])
                            dsm.quantity = (
                                dsm.quantity + float(request.POST['quantity']))
                            dsm.qorder = (
                                dsm.qorder + float(request.POST['quantity']))
                        pp = (dsm.ppurchase != float(request.POST['ppurchase']))
                        ps = (dsm.ppurchase != float(request.POST['psales']))
                        dsm.ppurchase = request.POST['ppurchase']
                        dsm.psales = request.POST['psales']
                        # function for update price to all the project
                        print 'price PURCHASE', request.POST['ppurchase']
                        print 'price SALES', request.POST['psales']
                        dsm.save()
                        if pp or ps:
                            sg = [x[0] for x in SGroup.objects.filter(
                                project_id=kwargs['pro'],
                                subproject_id=kwargs[
                                    'sub'] if unicode(kwargs[
                                        'sub']) != 'None' and unicode(kwargs[
                                            'sub']) != '' else None,
                                sector_id=kwargs['sec']).values_list('sgroup_id',)]
                            sec = [x[0] for x in DSector.objects.filter(
                                    sgroup_id__in=sg).values_list('dsector_id')]
                            dsmup = DSMetrado.objects.filter(
                                dsector_id__in=sec,
                                materials_id=request.POST['code'])
                            if 'ppurchase' in request.POST:
                                dsmup.update(ppurchase=request.POST['ppurchase'])
                            if 'psales' in request.POST:
                                dsmup.update(psales=request.POST['psales'])
                    except DSMetrado.DoesNotExist as e:
                        context['raise'] = str(e)
                        dsm = DSMetrado()
                        dsm.dsector_id = kwargs['area']
                        dsm.sector_id = kwargs['sec']
                        dsm.materials_id = request.POST['code']
                        dsm.brand_id = request.POST['brand']
                        dsm.model_id = request.POST['model']
                        dsm.quantity = request.POST['quantity']
                        dsm.qorder = request.POST['quantity']
                        dsm.qguide = 0
                        dsm.ppurchase = request.POST['ppurchase']
                        dsm.psales = request.POST['psales']
                        dsm.save()
                    context['status'] = True
                if 'delmat' in request.POST:
                    try:
                        DSMetrado.objects.filter(
                            dsector_id=kwargs['area'],
                            materials_id=request.POST['materials'],
                            brand_id=request.POST['brand'],
                            model_id=request.POST['model']).delete()
                        context['status'] = True
                    except DSMetrado.DoesNotExist, e:
                        context['raise'] = str(e)
                if 'copysector' in request.POST:
                    sec = MetProject.objects.filter(
                        proyecto_id=request.POST['project'],
                        sector_id=request.POST['sector'])
                    if not sec:
                        sec = Metradoventa.objects.filter(
                            proyecto_id=request.POST['project'],
                            sector_id=request.POST['sector'])
                    if sec:
                        for x in sec:
                            try:
                                ds = DSMetrado.objects.get(
                                        dsector_id=kwargs['area'],
                                        materials_id=x.materiales_id)
                            except DSMetrado.DoesNotExist:
                                ds = DSMetrado()
                            ds.dsector_id = kwargs['area']
                            ds.sector_id = kwargs['sec']
                            ds.materials_id = x.materiales_id
                            ds.brand_id = x.brand_id
                            ds.model_id = x.model_id
                            ds.quantity = x.cantidad
                            ds.qorder = x.cantidad
                            ds.qguide = 0
                            ds.ppurchase = x.precio
                            ds.psales = x.sales
                            ds.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                if 'delAreaMA' in request.POST:
                    DSMetrado.objects.filter(
                        dsector_id=kwargs['area']).delete()
                    context['status'] = True
                if 'availableNipple' in request.POST:
                    dsm = DSMetrado.objects.get(
                        dsector_id=kwargs['area'],
                        materials_id=request.POST['materials'],
                        brand_id=request.POST['brand'],
                        model_id=request.POST['model'])
                    if dsm.nipple:
                        dsm.nipple = False
                        dsm.save()
                    else:
                        dsm.nipple = True
                        dsm.save()
                    context['status'] = True
                if 'nipplesav' in request.POST:
                    try:
                        area = DSector.objects.get(
                                dsector_id=kwargs['area'])
                    except DSector.DoesNotExist:
                        area = {'sector_id': ''}
                    # kw = request.POST
                    request.POST._mutable = True
                    request.POST['proyecto'] = kwargs['area'][:7]
                    request.POST['area'] = kwargs['area']
                    request.POST['sector'] = area.sgroup.sector_id
                    request.POST._mutable = False
                    if 'edit' in request.POST:
                        Np = Nipple.objects.get(
                                id=request.POST['id'],
                                area_id=kwargs['area'],
                                materiales_id=request.POST['materiales'])
                        nip = NippleForm(request.POST, instance=Np)
                    else:
                        nip = NippleForm(request.POST)
                    if nip.is_valid():
                        nip.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                if 'delnipp' in request.POST:
                    Nipple.objects.get(
                        id=request.POST['id'],
                        area_id=kwargs['area'],
                        materiales_id=request.POST['materials']).delete()
                    context['status'] = True
                if 'modifyArea' in request.POST:
                    valid = MMetrado.objects.filter(dsector_id=kwargs['area'])
                    if not valid:
                        clipboard = DSMetrado.objects.filter(
                                        dsector_id=kwargs['area'])
                        if clipboard:
                            for x in clipboard:
                                add = MMetrado()
                                add.dsector_id = kwargs['area']
                                add.materials_id = x.materials_id
                                add.brand_id = x.brand_id
                                add.model_id = x.model_id
                                add.quantity = x.quantity
                                add.qorder = x.qorder
                                add.qguide = x.qguide
                                add.ppurchase = x.ppurchase
                                add.psales = x.psales
                                add.comment = x.comment
                                add.tag = x.tag
                                add.nipple = x.nipple
                                add.flag = x.flag
                                add.sector_id = kwargs['sec']
                                add.save()
                            context['status'] = True
                        else:
                            context['status'] = False
                    else:
                        context['status'] = False
                if 'editMM' in request.POST:
                    update = MMetrado.objects.filter(
                        dsector_id=kwargs['area'],
                        materials_id=request.POST['materials'],
                        brand_id=request.POST['brand'],
                        model_id=request.POST['model'])
                    if update:
                        if request.POST['name'] == 'brand':
                            update[0].brand_id = request.POST['value']
                        if request.POST['name'] == 'model':
                            update[0].model_id = request.POST['value']
                        if request.POST['name'] == 'quantity':
                            update[0].quantity = request.POST['value']
                        if request.POST['name'] == 'ppurchase':
                            update[0].ppurchase = request.POST['value']
                        if request.POST['name'] == 'psales':
                            update[0].psales = request.POST['value']
                        update[0].save()
                        context['status'] = True
                    else:
                        context['status'] = False
                        context['raise'] = 'Data not found'
                if 'delMM' in request.POST:
                    MMetrado.objects.filter(
                        dsector_id=kwargs['area'],
                        materials_id=request.POST['materials'],
                        brand_id=request.POST['brand'],
                        model_id=request.POST['model']).delete()
                    context['status'] = True
                if 'annModify' in request.POST:
                    MMetrado.objects.filter(dsector_id=kwargs['area']).delete()
                    context['status'] = True
                if 'savemmat' in request.POST:
                    try:
                        mm = MMetrado.objects.get(
                            dsector_id=kwargs['area'],
                            materials_id=request.POST['code'],
                            brand_id=request.POST['brand'],
                            model_id=request.POST['model'])
                        mm.quantity = (
                            mm.quantity + float(request.POST['quantity']))
                        mm.qorder = (
                            mm.qorder + float(request.POST['quantity']))
                        mm.ppurchase = request.POST['ppurchase']
                        mm.psales = request.POST['psales']
                    except MMetrado.DoesNotExist, e:
                        context['raise'] = str(e)
                        mm = MMetrado()
                        mm.dsector_id = kwargs['area']
                        mm.sector_id = kwargs['sec']
                        mm.materials_id = request.POST['code']
                        mm.brand_id = request.POST['brand']
                        mm.model_id = request.POST['model']
                        mm.quantity = request.POST['quantity']
                        mm.qorder = request.POST['quantity']
                        mm.qguide = 0
                        mm.ppurchase = request.POST['ppurchase']
                        mm.psales = request.POST['psales']
                    mm.save()
                    context['status'] = True
                if 'approvedModify' in request.POST:
                    lm = MMetrado.objects.filter(dsector_id=kwargs['area'])
                    mn = DSMetrado.objects.filter(
                            dsector_id=kwargs['area'])
                    kc = globalVariable.get_Token()
                    for o in mn:
                        # save history and delete area
                        his = HistoryDSMetrado()
                        his.qcode = kc
                        his.dsector_id = kwargs['area']
                        his.materials_id = o.materials_id
                        his.brand_id = o.brand_id
                        his.model_id = o.model_id
                        his.quantity = o.quantity
                        his.qorder = o.qorder
                        his.qguide = o.qguide
                        his.ppurchase = o.ppurchase
                        his.psales = o.psales
                        his.comment = o.comment
                        his.tag = o.tag
                        his.nipple = o.nipple
                        his.flag = o.flag
                        his.save()
                        try:
                            m = lm.get(
                                dsector_id=kwargs['area'],
                                materials_id=o.materials_id,
                                brand_id=o.brand_id,
                                model_id=o.model_id)
                            ds = DSMetrado()
                            ds.dsector_id = kwargs['area']
                            ds.sector_id = kwargs['sec']
                            ds.materials_id = m.materials_id
                            ds.brand_id = m.brand_id
                            ds.model_id = m.model_id
                            ds.quantity = m.quantity
                            if m.quantity < o.quantity:
                                ds.qorder = (m.qorder-(o.quantity-m.quantity))
                            elif m.quantity > o.quantity:
                                ds.qorder = ((m.quantity-o.quantity)+m.qorder)
                            else:
                                ds.qorder = m.qorder
                            ds.qguide = m.qguide
                            ds.ppurchase = m.ppurchase
                            ds.psales = m.psales
                            ds.comment = m.comment
                            if ds.qorder == ds.quantity:
                                ds.tag = '0'
                            if (ds.qorder < ds.quantity) and (ds.qorder > 0):
                                ds.tag = '1'
                            if ds.qorder <= 0:
                                ds.qorder = 0
                                ds.tag = '2'
                            ds.flag = m.flag
                            ds.nipple = m.nipple
                            o.delete()
                            ds.save()
                        except ObjectDoesNotExist:
                            o.delete()
                    lds = DSMetrado.objects.filter(dsector_id=kwargs['area'])
                    if lds.count() != lm.count():
                        for x in lm:
                            try:
                                DSMetrado.objects.get(
                                    dsector_id=kwargs['area'],
                                    materials_id=x.materials_id,
                                    brand_id=x.brand_id,
                                    model_id=x.model_id)
                            except Exception:
                                ds = DSMetrado()
                                ds.dsector_id = kwargs['area']
                                ds.sector_id = kwargs['sec']
                                ds.materials_id = x.materials_id
                                ds.brand_id = x.brand_id
                                ds.model_id = x.model_id
                                ds.quantity = x.quantity
                                ds.qorder = x.quantity
                                ds.qguide = 0
                                ds.ppurchase = x.ppurchase
                                ds.psales = x.psales
                                ds.comment = x.comment
                                ds.tag = '0'
                                ds.flag = x.flag
                                ds.nipple = x.nipple
                                ds.save()
                    lm.delete()
                    context['status'] = True
                if 'saveComment' in request.POST:
                    dm = DSMetrado.objects.get(
                        dsector_id=kwargs['area'],
                        materials_id=request.POST['materials'],
                        brand_id=request.POST['brand'],
                        model_id=request.POST['model'])
                    dm.comment = request.POST['comment']
                    dm.save()
                    context['status'] = True
                if 'saveOrders' in request.POST:
                    # get id new orders
                    gkey = genkeys.GenerateIdOrders()
                    orders = Pedido()
                    orders.pedido_id = gkey
                    orders.proyecto_id = kwargs['pro']
                    orders.sector_id = kwargs['sec']
                    orders.dsector_id = kwargs['area']
                    orders.almacen_id = request.POST['storage']
                    orders.asunto = request.POST['issue']
                    orders.empdni_id = request.user.get_profile().empdni_id
                    orders.traslado = request.POST['transfer']
                    orders.obser = request.POST['observation']
                    orders.status = 'PE'
                    if 'ordersf' in request.FILES:
                        orders.orderfile = request.FILES['ordersf']
                    orders.save()
                    # save orders details
                    det = json.loads(request.POST['details'])
                    # print det
                    for x in det:
                        ds = DSMetrado.objects.get(
                                dsector_id=kwargs['area'],
                                materials_id=x['materials'])
                        d = Detpedido()
                        d.pedido_id = gkey
                        d.materiales_id = x['materials']
                        d.brand_id = ds.brand_id
                        d.model_id = ds.model_id
                        d.cantidad = x['quantity']
                        d.cantshop = x['quantity']
                        d.tag = '0'
                        d.comment = ds.comment if ds.comment else ''
                        ds.qorder -= float(d.cantshop)
                        if ds.qorder > 0:
                            ds.tag = '1'
                        else:
                            ds.tag = '2'
                        ds.save()
                        d.save()
                    # save nipples
                    if 'nipples' in request.POST:
                        nipples = json.loads(request.POST['nipples'])
                        print nipples
                        for n in nipples:
                            npo = Nipple.objects.get(id=n['id'])
                            nip = Niple()
                            nip.pedido_id = gkey
                            nip.proyecto_id = kwargs['pro']
                            nip.sector_id = kwargs['sec']
                            nip.dsector_id = kwargs['area']
                            nip.empdni = request.user.get_profile().empdni_id
                            nip.materiales_id = n['m']
                            nip.cantidad = n['quantity']
                            nip.metrado = n['measure']
                            nip.cantshop = n['quantity']
                            nip.cantguide = 0
                            nip.tipo = npo.tipo
                            nip.comment = npo.comment
                            nip.tag = '0'
                            npo.cantshop -= float(nip.cantshop)
                            if npo.cantshop > 0:
                                npo.tag = '1'
                            else:
                                npo.tag = '2'
                            npo.save()
                            nip.save()
                    context['orders'] = gkey
                    context['status'] = True
                if 'uploadPlane' in request.POST:
                    sf = SectorFiles()
                    sf.dsector_id = kwargs['area']
                    sf.sector_id = kwargs['sec']
                    sf.proyecto_id = kwargs['pro']
                    sf.files = request.FILES['plane']
                    sf.note = request.POST['note']
                    sf.flag = True
                    sf.save()
                    context['status'] = True
                if 'delplane' in request.POST:
                    sf = SectorFiles.objects.filter(pk=request.POST['plane'])
                    for x in sf:
                        uploadFiles.deleteFile(str(x.files), partial=True)
                        x.delete()
                    context['status'] = True
            except ObjectDoesNotExist as e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)


class CompareMaterials(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        try:
            if request.is_ajax():
                if 'lbrand' in request.GET:
                    context['brand'] = json.loads(
                                    serializers.serialize(
                                        'json',Brand.objects.filter(flag=True).distinct('brand').order_by('brand')))
                    context['status'] = True
                if 'saveChange' in request.GET:
                    d = DSMetrado.objects.filter(
                            dsector__dsector_id__startswith=kwargs['pro'],
                            materials_id=request.GET['materials'],
                            brand_id=request.GET['obrand'],
                            model_id=request.GET['omodel'])
                    print request.GET
                    print d.count()
                    print 'COUNT DSECTOR'
                    if d:
                        for x in d:
                            x.brand_id = request.GET['brand']
                            x.model_id = request.GET['model']
                            x.ppurchase = request.GET['ppurchase']
                            x.psales = request.GET['psales']
                            x.save()
                            # time.sleep(3)
                    context['status'] = True
                if 'glist' in request.GET:
                    sales = MetProject.objects.filter(
                        proyecto_id=kwargs['pro'], sector_id=kwargs['sec'])
                    ds = DSector.objects.filter(
                        project_id=kwargs['pro'], sector_id=kwargs['sec'])
                    print ds
                    operations = DSMetrado.objects.filter(
                        dsector_id__in=[x.dsector_id for x in ds])
                    # print 'count opertaions ', operations.count()
                    # .order_by(
                    #    'materials__materiales_id').distinct(
                    #    'materials__materiales_id')
                    lst = list()
                    for s in sales:
                        lst.append({
                            'materials': s.materiales_id,
                            'name': '%s %s' % (
                                s.materiales.matnom, s.materiales.matmed),
                            'brand_id': s.brand_id,
                            'brand': s.brand.brand,
                            'model_id': s.model_id,
                            'model': s.model.model,
                            'unit': s.materiales.unidad.uninom,
                            'sales': s.cantidad,
                            'purchase': s.precio,
                            'operations': '-'})
                    for o in operations:
                        c = 0
                        for x in lst:
                            mat = (x['materials'] == o.materials_id)
                            brand = (x['brand_id'] == o.brand_id)
                            model = (x['model_id'] == o.model_id)
                            if mat and brand and model:
                                if x['operations'] == '-':
                                    x['operations'] = 0
                                x['purchase'] = o.ppurchase
                                x['psales'] = o.psales
                                x['operations'] = (x['operations'] + o.quantity)
                                x['amount'] = (x['operations'] * float(x['purchase']))
                                break
                            else:
                                c += 1
                        if len(lst) == c:
                            lst.append({
                                'materials': o.materials_id,
                                'name': '%s %s' % (
                                    o.materials.matnom, o.materials.matmed),
                                'brand_id': o.brand_id,
                                'brand': o.brand.brand,
                                'model_id': o.model_id,
                                'model': o.model.model,
                                'unit': o.materials.unidad.uninom,
                                'sales': '-',
                                'purchase': float(o.ppurchase),
                                'operations': o.quantity,
                                'amount': (o.quantity*float(o.ppurchase))})

                    context['lst'] = lst
                    context['sales'] = sales.aggregate(amount=Sum(
                                        'cantidad', field='cantidad*precio'))['amount']
                    # context['operations'] = operations.aggregate(amount=Sum(
                    #                'quantity', field='quantity*ppurchase'))['amount']
                    am = 0
                    for x in lst:
                        if not type(x['operations']) is str:
                            am += x['amount']
                        # am += (x.quantity*float(x.ppurchase))
                    context['operations'] = am
                    context['currency'] = sales[0].proyecto.currency.moneda
                    context['symbol'] = sales[0].proyecto.currency.simbolo
                    context['salesap'] = sales[0].sector.amount
                    context['diff'] = (sales[0].sector.amount - context['operations'])
                return self.render_to_json_response(context)
            if 'export' in request.GET:
                response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=DATOS-%s-%s.xlsx' %(kwargs['pro'], kwargs['sec'])
                wb = Workbook()
                ws = wb.active
                ws.title = 'Materiales'
                ws.sheet_properties.tabColor = '1072BA'
                # ws = wb.create_sheet()
                # sales = MetProject.objects.filter(
                #         proyecto_id=kwargs['pro'], sector_id=kwargs['sec'])
                # ds = DSector.objects.filter(
                #        project_id=kwargs['pro'], sector_id=kwargs['sec'])
                operations = DSMetrado.objects.filter(sector_id=kwargs['sec']).order_by('materials__matnom')
                # operations = operations
                lst = list()
                for o in operations:
                    c = 0
                    for x in lst:
                        mat = (x['materials'] == o.materials_id)
                        brand = (x['brand_id'] == o.brand_id)
                        model = (x['model_id'] == o.model_id)
                        if mat and brand and model:
                            x['operations'] = (x['operations'] + o.quantity)
                            # x['amount'] = (x['operations'] * float(x['purchase']))
                            break
                        else:
                            c += 1
                    if len(lst) == c:
                        lst.append({
                            'materials': o.materials_id,
                            'name': '%s %s' % (
                                o.materials.matnom, o.materials.matmed),
                            'brand_id': o.brand_id,
                            'brand': o.brand.brand,
                            'model_id': o.model_id,
                            'model': o.model.model,
                            'unit': o.materials.unidad.uninom,
                            #'sales': '-',
                            #'purchase': float(o.ppurchase),
                            'operations': o.quantity,})
                            # 'amount': (o.quantity*float(o.ppurchase))})
                columns = [
                    ('Código', 16),
                    ('Descripción', 60),
                    ('Unidad', 9),
                    ('Marca', 9),
                    ('Modelo', 9),
                    ('Cantidad Vendida', 20),]
                border = styles.borders.Border(left=styles.borders.Side(style='thin'),
                                                right=styles.borders.Side(style='thin'),
                                                top=styles.borders.Side(style='thin'),
                                                bottom=styles.borders.Side(style='thin'))
                mstyle = styles.Style(border=border)
                # col_num = 0
                for col_num in xrange(len(columns)):
                    c = ws.cell(row=1, column=col_num+1)
                    c.value = columns[col_num][0]
                    c.style = mstyle
                    c.font = styles.Font(name='Tahoma', bold=True, size=9)
                    ws.column_dimensions[cell.get_column_letter(col_num+1)].width = columns[col_num][1]
                    # col_num+=1
                rw = 2
                for x in lst:
                    ws.cell(column=1, row=rw).value = x['materials']
                    ws.cell(column=1, row=rw).style = mstyle
                    ws.cell(column=2, row=rw).value = x['name']
                    ws.cell(column=2, row=rw).style = mstyle
                    ws.cell(column=3, row=rw).value = x['unit']
                    ws.cell(column=3, row=rw).style = mstyle
                    ws.cell(column=4, row=rw).value = x['brand']
                    ws.cell(column=4, row=rw).style = mstyle
                    ws.cell(column=5, row=rw).value = x['model']
                    ws.cell(column=5, row=rw).style = mstyle
                    ws.cell(column=6, row=rw).value = x['operations']
                    ws.cell(column=6, row=rw).style = mstyle
                    rw+=1
                # print len(ws.cell(column=1, row=2).width)
                wb.save(response)
                return response
            return render(request, 'operations/comparematerials.html', context)
        except TemplateDoesNotExist, e:
            raise Http404(e)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        context = dict()
        try:
            if 'saveBrand' in request.POST:
                brand = request.POST['brand'].strip().upper()
                o = Brand.objects.filter(brand=brand)
                if not o.count():
                    key = genkeys.GenerateIdBrand()
                    obj = Brand()
                    obj.brand = brand
                    obj.brand_id = key
                    obj.save()
                    context['id'] = key
                    context['name'] = brand
                else:
                    context['id'] = o[0].brand_id
                    context['name'] = o[0].brand
                context['status'] = True
            if 'saveModel' in request.POST: 
                model = request.POST['model'].strip().upper()
                o = Model.objects.filter(brand_id=request.POST['brand'], model=model)
                if not o.count():
                    key = genkeys.GenerateIdModel()
                    obj = Model()
                    obj.model_id = key
                    obj.brand_id = request.POST['brand']
                    obj.model = model
                    obj.save()
                    context['id'] = key
                    context['name'] = model
                else:
                    context['id'] = o[0].model_id
                    context['name'] = o[0].model
                context['status'] = True
        except ObjectDoesNotExist, e:
            context['raise'] = str(e)
            context['status'] = False
        return self.render_to_json_response(context)

class  ConsultMaterialsAreas(JSONResponseMixin, View):

    def get(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'gsgroup' in request.GET:
                        context['sg'] = json.loads(
                            serializers.serialize(
                                'json',
                                SGroup.objects.filter(project_id=kwargs['pro'],sector_id=kwargs['sec'])))
                        context['status'] = True
                    if 'gdsector' in request.GET:
                        context['ds'] = json.loads(
                            serializers.serialize(
                                'json',
                                DSector.objects.filter(project_id=kwargs['pro'], sector_id=kwargs['sec'])))
                        context['status'] = True
                    if 'getGlobal' in request.GET:
                        query = DSMetrado.objects.filter(sector_id=kwargs['sec'])
                        query = query.values_list(
                            'materials_id',
                            'materials__matnom',
                            'materials__matmed',
                            'brand__brand',
                            'model__model',
                            'materials__unidad__uninom',).annotate(sold=Sum('quantity'), pending=Sum('qorder')).order_by('materials__matnom')
                        tmp = [{
                            'materials': x[0],
                            'matnom': x[1],
                            'matmed': x[2],
                            'brand': x[3],
                            'model': x[4],
                            'unit': x[5],
                            'sold': x[6],
                            'pending': x[7],
                        } for x in query]
                        context['dataset'] = tmp
                        context['status'] = True
                    if 'getPendingData' in request.GET:
                        keys = request.GET['keys'].split(',')
                        query = None
                        count = 0
                        if request.GET['searchby'] == 'sgroup':
                            for key in keys:
                                aux = DSMetrado.objects.filter(dsector__dsector_id__startswith=key)
                                if count == 0:
                                    query = aux
                                    count+=1
                                query = query | aux
                        if request.GET['searchby'] == 'dsector':
                            for key in keys:
                                aux = DSMetrado.objects.filter(dsector__dsector_id__startswith=key)
                                if count == 0:
                                    query = aux
                                    count += 1
                                query = query | aux
                        # context['query'] = json.loads(serializers.serialize('json', query))
                        query = query.values_list(
                            'materials_id',
                            'materials__matnom',
                            'materials__matmed',
                            'brand__brand',
                            'model__model',
                            'materials__unidad__uninom',).annotate(sold=Sum('quantity'), pending=Sum('qorder')).order_by('materials__matnom')
                        # query = query.distinct('materials').extra(select={'squantity':"SELECT SUM(quantity) FROM operations_dsmetrado t WHERE t.materials_id LIKE operations_dsmetrado.materials_id AND t.dsector_id LIKE operations_dsmetrado.dsector_id"})
                        # query = query.annotate(Sum('quantity'))
                        # print query
                        tmp = [{
                            'materials': x[0],
                            'matnom': x[1],
                            'matmed': x[2],
                            'brand': x[3],
                            'model': x[4],
                            'unit': x[5],
                            'sold': x[6],
                            'pending': x[7],
                        } for x in query]
                        context['dataset'] = tmp
                        context['status'] = True
                    if 'getdsmetrado' in request.GET:
                        context['status'] = True
                    if 'getDetails' in request.GET:
                        if 'keys' in request.GET:
                            keys = request.GET['keys'].split(',')
                            query = None
                            count = 0
                            if request.GET['searchby'] == 'sgroup':
                                for key in keys:
                                    aux = DSMetrado.objects.filter(dsector__dsector_id__startswith=key, materials_id=request.GET['materials'])
                                    if count == 0:
                                        query = aux
                                        count+=1
                                    query = query | aux
                            if request.GET['searchby'] == 'dsector':
                                for key in keys:
                                    aux = DSMetrado.objects.filter(dsector__dsector_id__startswith=key, materials_id=request.GET['materials'])
                                    if count == 0:
                                        query = aux
                                        count += 1
                                    query = query | aux
                        else:
                            query = DSMetrado.objects.filter(sector_id=kwargs['sec'], materials_id=request.GET['materials'])
                        context['data'] = json.loads(
                            serializers.serialize(
                                'json',
                                query,
                                relations=('materials','brand','model','dsector',)))
                        context['status'] = True
                except (ObjectDoesNotExist, Exception) as e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
            return render(request, 'operations/listpending.html', context)
        except TemplateDoesNotExist, e:
            raise Http404(str(e))

