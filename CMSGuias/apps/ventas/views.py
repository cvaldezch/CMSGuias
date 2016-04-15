#!/usr/bin/env python
# -*- coding: utf-8 -*-

import json
import os
import shutil

from django.db.models import Q, Sum
# from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.core.exceptions import ObjectDoesNotExist
from django.core import serializers
from django.contrib import messages
# from django.contrib.auth.mod import User
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse
from django.shortcuts import render_to_response, render, redirect
from django.utils import simplejson
from django.utils.decorators import method_decorator
from django.template import RequestContext, TemplateDoesNotExist
from django.views.generic import TemplateView, View
from openpyxl import load_workbook
# from django.views.generic.edit import UpdateView, CreateView
# from decimal import Decimal
from django.core.serializers.json import DjangoJSONEncoder

from CMSGuias.apps.home.models import *
from CMSGuias.apps.operations.models import (
                                            MetProject,
                                            Nipple,
                                            Deductive,
                                            DeductiveInputs,
                                            DeductiveOutputs,
                                            Letter,
                                            LetterAnexo,
                                            DSector)
from CMSGuias.apps.almacen.models import (
                                            Inventario,
                                            Pedido,
                                            Detpedido,
                                            Niple,
                                            GuiaRemision,
                                            DetGuiaRemision,
                                            NipleGuiaRemision)
from .models import *
from .forms import *
from CMSGuias.apps.logistica.models import *
from CMSGuias.apps.almacen.forms import addOrdersForm
from CMSGuias.apps.operations.forms import (NippleForm, LetterForm,
                                            PreOrdersForm)
from CMSGuias.apps.tools import genkeys, globalVariable, uploadFiles, search


# Class Bases Views Generic

class JSONResponseMixin(object):
    def render_to_json_response(self, context, **response_kwargs):
        return HttpResponse(
            self.convert_context_to_json(context),
            content_type='application/json',
            mimetype='application/json',
            **response_kwargs
        )

    def convert_context_to_json(self, context):
        return simplejson.dumps(context,
                                encoding='utf-8',
                                cls=DjangoJSONEncoder)


class SalesHome(TemplateView):
    template_name = 'sales/home.html'

    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(SalesHome, self).dispatch(request, *args, **kwargs)


class ProjectsList(JSONResponseMixin, TemplateView):
    template_name = 'sales/projects.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        area = request.user.get_profile().empdni.charge.area.lower()
        if request.is_ajax():
            try:
                if 'lCustomers' in request.GET:
                    context['customers'] = simplejson.loads(
                                            serializers.serialize(
                                                'json',
                                                Cliente.objects.filter(
                                                    flag=True)))
                    context['status'] = True
                if 'getCustomers' in request.GET:
                    if area == 'ventas' or area == 'administrator':
                        proyectos = Proyecto.objects.filter(
                                            Q(flag=True),
                                            ~Q(status='DA')
                                            ).order_by('-proyecto_id')
                    elif area == 'operaciones':
                        cnom = request.user.get_profile(
                            ).empdni.charge.cargos.lower()
                        if cnom == 'jefe de operaciones':
                            proyectos = Proyecto.objects.filter(
                                            Q(flag=True),
                                            Q(status='AC')).order_by(
                                            '-proyecto_id')
                        else:
                            proyectos = Proyecto.objects.filter(
                                    Q(flag=True),
                                    Q(status='AC'),
                                    empdni_id=request.user.get_profile(
                                        ).empdni_id).order_by('-proyecto_id')
                    elif area == 'logistica' or area == 'almacen':
                        proyectos = Proyecto.objects.filter(
                                            Q(flag=True),
                                            Q(status='AC')).order_by(
                                            '-proyecto_id')
                    cust = proyectos
                    if area == 'operaciones':
                        if cnom == 'jefe de operaciones':
                            cust = cust.filter(status='AC')
                        else:
                            cust = cust.filter(
                                empdni_id=request.user.get_profile().empdni_id,
                                status='AC')
                    elif area == 'logistica' or area == 'almacen':
                        cust = cust.filter(status='AC')
                    cust = cust.order_by(
                            'ruccliente__razonsocial').distinct(
                            'ruccliente__razonsocial')
                    context['customers'] = simplejson.loads(
                                            serializers.serialize(
                                                'json',
                                                cust,
                                                indent=4,
                                                relations=('ruccliente',)))
                    context['status'] = True
                if 'getProjects' in request.GET:
                    if area == 'ventas' or area == 'administrator':
                        projects = Proyecto.objects.filter(
                                    Q(flag=True),
                                    ~Q(status='DA'),
                                    Q(ruccliente_id=request.GET['customer'])
                                    ).order_by('-proyecto_id')
                    elif area == 'operaciones':
                        projects = Proyecto.objects.filter(
                                    Q(flag=True),
                                    Q(status='AC'),
                                    Q(ruccliente_id=request.GET['customer']),
                                    empdni_id=request.user.get_profile(
                                        ).empdni_id).order_by('-proyecto_id')
                    elif area == 'logistica' or area == 'almacen':
                        projects = Proyecto.objects.filter(
                                    Q(flag=True),
                                    Q(status='AC'),
                                    Q(ruccliente_id=request.GET['customer'])
                                    ).order_by('-proyecto_id')
                    cnom = request.user.get_profile(
                            ).empdni.charge.cargos.lower()
                    if cnom == 'jefe de operaciones':
                        projects = Proyecto.objects.filter(
                                    Q(flag=True),
                                    Q(status='AC'),
                                    Q(ruccliente_id=request.GET['customer'])
                                    ).order_by('-proyecto_id')
                    context['projects'] = simplejson.loads(
                                            serializers.serialize(
                                                'json', projects))
                    context['status'] = True
                if 'allProjects' in request.GET:
                    if area == 'ventas' or area == 'administrator':
                        projects = Proyecto.objects.filter(
                                    Q(flag=True),
                                    ~Q(status='DA')).order_by('-proyecto_id')
                    elif area == 'operaciones':
                        projects = Proyecto.objects.filter(
                                    Q(flag=True),
                                    Q(status='AC'),
                                    empdni_id=request.user.get_profile(
                                        ).empdni_id).order_by('-proyecto_id')
                    elif area == 'logistica' or area == 'almacen':
                        projects = Proyecto.objects.filter(
                                    Q(flag=True),
                                    Q(status='AC')).order_by('-proyecto_id')
                    cnom = request.user.get_profile(
                            ).empdni.charge.cargos.lower()
                    if cnom == 'jefe de operaciones':
                        projects = Proyecto.objects.filter(
                                    Q(flag=True),
                                    Q(status='AC')).order_by('-proyecto_id')
                    context['projects'] = simplejson.loads(
                                            serializers.serialize(
                                                'json', projects))
                    context['status'] = True
                if 'ascAllProjects' in request.GET:
                    context['projects'] = json.loads(serializers.serialize(
                        'json',
                        Proyecto.objects.filter(
                            Q(flag=True),
                            ~Q(status='DA')
                            ).order_by('-proyecto_id')))
                    context['status'] = True
            except ObjectDoesNotExist as e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)
        try:
            if area == 'ventas' or area == 'administrator':
                    context['currency'] = Moneda.objects.filter(flag=True)
                    context['typep'] = globalVariable.typeProject
            return render_to_response(
                    self.template_name,
                    context,
                    context_instance=RequestContext(request))
        except TemplateDoesNotExist, e:
            messages.error(request, 'Template Does Not Exist %s' % e)
            raise Http404('Template Does Not Exist %s' % e)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            context = dict()
            try:
                if request.POST.get('type') == 'new':
                    form = ProjectForm(request.POST)
                    key = genkeys.GenerateIdPorject()
                    if form.is_valid():
                        add = form.save(commit=False)
                        add.proyecto_id = key
                        # add.empdni_id = request.user.get_profile().empdni_id
                        add.status = 'PE'
                        add.save()
                        context['status'] = True
                    else:
                        context['status'] = False
            except ObjectDoesNotExist, e:
                context['raise'] = e.__str__()
                context['status'] = False
            return self.render_to_json_response(context)


# add sectors
class SectorsView(JSONResponseMixin, View):
    template_name = 'sales/crud/sectors_form.html'

    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                obj = Sectore.objects.filter(
                        proyecto_id=request.GET.get('pro'),
                        subproyecto_id=request.GET.get(
                            'sub') if request.GET.get('sub') != '' else None,
                        flag=True).order_by('subproyecto', 'planoid')
                context['list'] = [{
                    'sector_id': x.sector_id,
                    'nomsec': x.nomsec,
                    'planoid': x.planoid} for x in obj]
                context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = e.__str__()
                context['status'] = False
            return self.render_to_json_response(context)
        context['pro'] = request.GET.get('pro')
        context['sub'] = request.GET.get('sub')
        if request.GET.get('type') == 'update':
            obj = Sectore.objects.get(
                    proyecto_id=request.GET.get('pro'),
                    subproyecto_id=request.GET.get('sub') if request.GET.get(
                        'sub') != '' else None,
                    sector_id=request.GET.get('sec'))
            context['form'] = SectoreForm(instance=obj)
            context['type'] = 'update'
        elif request.GET.get('type') == 'new':
            context['form'] = SectoreForm
            context['type'] = 'new'
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))

    def post(self, request, *args, **kwargs):
        context = dict()
        try:
            if request.POST.get('type') == 'update':
                obj = Sectore.objects.get(
                    proyecto_id=request.POST.get('proyecto'),
                    subproyecto_id=request.POST.get('sub') if request.POST.get(
                        'subproyecto') != '' else None,
                    sector_id=request.POST.get('sector_id'))
                form = SectoreForm(request.POST, instance=obj)
                # print form, form.is_valid()
                if form.is_valid():
                    form.save()
                    context['status'] = True
                    context['msg'] = 'success'
                else:
                    context['status'] = False
                    context['msg'] = 'error'
            elif request.POST.get('type') == 'new':
                form = SectoreForm(request.POST)
                # print form, form.is_valid()
                if form.is_valid():
                    add = form.save(commit=False)
                    id = '%s%s' % (add.proyecto_id, add.sector_id)
                    # print id, len(id)
                    add.sector_id = id
                    add.save()
                    context['status'] = True
                    context['msg'] = 'success'
                else:
                    context['status'] = False
                    context['msg'] = 'error'
        except ObjectDoesNotExist, e:
            context['raise'] = e.__str__()
            context['status'] = False
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


# Add Subproject
class SubprojectsView(JSONResponseMixin, View):
    template_name = 'sales/crud/subprojects_form.html'

    def get(self, request, *args, **kwargs):
        context = dict()
        context['pro'] = request.GET.get('pro')
        if request.GET.get('type') == 'update':
            obj = Subproyecto.objects.get(proyecto_id=request.GET.get('pro'), subproyecto_id=request.GET.get('sub'))
            context['form'] = SubprojectForm(instance=obj)
            context['type'] = 'update'
        elif request.GET.get('type') == 'new':
            context['form'] = SubprojectForm
            context['type'] = 'new'
        return render_to_response(self.template_name, context, context_instance = RequestContext(request))

    def post(self, request, *args, **kwargs):
        context = dict()
        try:
            if request.POST.get('type') == 'update':
                obj = Subproyecto.objects.get(proyecto_id=request.POST.get('proyecto'),subproyecto_id=request.POST.get('subproyecto_id'))
                form = SubprojectForm(request.POST, instance=obj)
                # print form, form.is_valid()
                if form.is_valid():
                    form.save()
                    context['status'] = True
                    context['msg'] = 'success'
                else:
                    context['status'] = False
                    context['msg'] = 'error'
            elif request.POST.get('type') == 'new':
                form = SubprojectForm(request.POST)
                #print form, form.is_valid()
                if form.is_valid():
                    form.save()
                    context['status'] = True
                    context['msg'] = 'success'
                else:
                    context['status'] = False
                    context['msg'] = 'error'
        except ObjectDoesNotExist, e:
            context['raise'] = e.__str__()
            context['status'] = False
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


# Manager View Project
class ProjectManager(JSONResponseMixin, View):
    template_name = 'sales/managerpro.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        try:
            if request.is_ajax():
                try:
                    if 'listPurchase' in request.GET:
                        context['list'] = [{
                            'nro': x.nropurchase,
                            'issued': globalVariable.format_date_str(x.issued),
                            'document': x.document.documento,
                            'order': str(x.order),
                            'id': x.id}
                            for x in PurchaseOrder.objects.filter(
                                        flag=True,
                                        project_id=kwargs['project']
                                        ).order_by('register')]
                        context['status'] = True
                    if 'editPurchase' in request.GET:
                        obj = PurchaseOrder.objects.get(
                                pk=request.GET.get('pk'))
                        context['nropurchase'] = obj.nropurchase
                        context['issued'] = globalVariable.format_date_str(
                                                obj.issued)
                        context['currency'] = obj.currency_id
                        context['document'] = obj.document_id
                        context['method'] = obj.method_id
                        context['observation'] = obj.observation
                        context['dsct'] = obj.dsct
                        context['igv'] = obj.igv
                        context['details'] = [{
                            'description': x.description,
                            'unit': x.unit.uninom,
                            'delivery': globalVariable.format_date_str(
                                            x.delivery),
                            'quantity': x.quantity,
                            'price': x.price,
                            'amount': (x.quantity * x.price)}
                            for x in DetailsPurchaseOrder.objects.filter(
                                        purchase_id=request.GET.get('pk'))]
                        context['status'] = True
                    if 'letterlist' in request.GET:
                        context['list'] = [{
                            'letter': x.letter_id,
                            'register': x.register.strftime('%d-%m-%Y'),
                            'froms': x.froms,
                            'fors': x.fors,
                            'file': str(x.letter),
                            'status': x.status,
                            'observation': x.observation}
                            for x in Letter.objects.filter(
                                        Q(project_id=kwargs['project']),
                                        ~Q(status='AN')).order_by('-register')]
                        context['status'] = True
                except ObjectDoesNotExist, e:
                    context['raise'] = e.__str__()
                    context['status'] = False
                return self.render_to_json_response(context)
            context['project'] = Proyecto.objects.get(
                                    pk=kwargs['project'], flag=True)
            try:
                context['subpro'] = Subproyecto.objects.filter(
                                        proyecto_id=kwargs['project'],
                                        flag=True)
            except ObjectDoesNotExist, e:
                context['subpro'] = list()
            context['sectors'] = Sectore.objects.filter(proyecto_id=kwargs['project'], flag=True).order_by('subproyecto', 'planoid')
            context['operation'] = Employee.objects.filter(
                Q(charge__area__istartswith='opera') | Q(empdni_id=72604244)).order_by('charge__area')
            context['admin'] = Employee.objects.filter(charge__area__istartswith='admin').order_by('charge__area')
            context['alerts'] = Alertasproyecto.objects.filter(Q(proyecto_id=kwargs['project']) | ~Q(subproyecto_id=None), Q(sector_id=None), Q(flag=True)).order_by('-registrado')
            context['currency'] = Moneda.objects.filter(flag=True)
            context['document'] = Documentos.objects.filter(flag=True)
            context['method'] = FormaPago.objects.filter(flag=True)
            context['unit'] = Unidade.objects.filter(flag=True)
            context['conf'] = Configuracion.objects.get(periodo=globalVariable.get_year)
            try:
                context['closure'] = CloseProject.objects.get(project_id=kwargs['project'])
            except ObjectDoesNotExist, e:
                pass
            return render_to_response(self.template_name, context, context_instance = RequestContext(request))
        except TemplateDoesNotExist, e:
            messages.error(request, 'Template not Exist %s',e)
            raise Http404('Page Not Found')

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            context = dict()
            try:
                if request.POST.get('type') == 'files':
                    year = Proyecto.objects.get(pk=kwargs['project']).registrado.strftime('%Y') #globalVariable.get_year
                    try:
                        # charge file to server
                        if request.POST.get('sub') == '':
                            admin = '/storage/projects/%s/%s/administrative/'%(year, request.POST.get('pro'))
                            opera = '/storage/projects/%s/%s/operation/'%(year, request.POST.get('pro'))
                        else:
                            admin = '/storage/projects/%s/%s/%s/administrative/'%(year, request.POST.get('pro'), request.POST.get('sub'))
                            opera = '/storage/projects/%s/%s/%s/operation/'%(year, request.POST.get('pro'), request.POST.get('sub'))

                        if 'administrative' in request.FILES:
                            fileadmin = uploadFiles.upload(admin, request.FILES['administrative'], {'name': 'admin'})
                            # descompress files in the server
                            context['descompress'] =  uploadFiles.descompressRAR(fileadmin, admin)
                            # delete files temp
                            uploadFiles.removeTmp(fileadmin)
                        if 'operation' in request.FILES:
                            fileopera = uploadFiles.upload(opera, request.FILES['operation'], {'name': 'opera'})
                            # descompress files in the server
                            context['descompress'] = uploadFiles.descompressRAR(fileopera, opera)
                            # delete files temp
                            uploadFiles.removeTmp(fileopera)
                        context['status'] = True
                    except ObjectDoesNotExist, e:
                        # print e
                        context['raise'] = e.__str__()
                        context['status'] = False
                if request.POST.get('type') == 'add':
                    form = AlertasproyectoForm(request.POST)
                    if form.is_valid():
                        add = form.save(commit=False)
                        add.empdni_id = request.user.get_profile().empdni_id
                        add.charge_id = request.user.get_profile().empdni.charge_id
                        add.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                if request.POST.get('type') == 'edit':
                    obj = Alertasproyecto.objects.get(pk=request.POST.get('edit'))
                    form = AlertasproyectoForm(request.POST, instance=obj)
                    if form.is_valid():
                        form.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                if request.POST.get('type') == 'responsible':
                    try:
                        #search.validKey(request.POST.get('passwd'),kwargs['project'],'responsible',request.user.get_profile().empdni_id)
                        # authenticate password admin
                        if search.validKey(request.POST.get('passwd'),kwargs['project'],'responsible',request.user.get_profile().empdni_id):
                            #if user.user.is_superuser:
                            pro = Proyecto.objects.get(pk=kwargs['project'])
                            pro.empdni_id = request.POST.get('responsible')
                            pro.save()
                            context['status'] = True
                        else:
                            context['raise'] = 'Código incorrecto!'
                            context['status'] = False
                    except ObjectDoesNotExist, e:
                        context['raise'] = e.__str__()
                        context['status'] = False
                if request.POST.get('type') == 'approved':
                    # this function is for approve the project
                    try:
                        #user = userProfile.objects.get(empdni_id=request.POST.get('admin'))
                        # authenticate password admin
                        if search.validKey(request.POST.get('passwd'),kwargs['project'],'approved',request.user.get_profile().empdni_id):
                            # if user.user.is_superuser:
                            pro = Proyecto.objects.get(pk=kwargs['project'])
                            pro.approved_id = request.user.get_profile().empdni_id
                            pro.status = 'AC'
                            pro.save()
                            sub = Subproyecto.objects.filter(
                                    proyecto_id=kwargs['project'])
                            if sub:
                                sub.update(status='AC')
                            sec = Sectore.objects.filter(
                                    proyecto_id=kwargs['project'])
                            if sec:
                                sec.update(status='AC')
                            # paste all list of materials to 'MetProject'
                            for x in Metradoventa.objects.filter(
                                        proyecto_id=kwargs['project']):
                                obj = MetProject()
                                obj.proyecto_id = x.proyecto_id
                                obj.subproyecto_id = x.subproyecto_id
                                obj.sector_id = x.sector_id
                                obj.materiales_id = x.materiales_id
                                obj.cantidad = x.cantidad
                                obj.precio = x.precio
                                obj.sales = x.sales
                                obj.brand_id = x.brand_id
                                obj.model_id = x.model_id
                                obj.quantityorder = x.cantidad
                                obj.save()
                            context['status'] = True
                        else:
                            context['raise'] = 'Password incorrect!'
                            context['status'] = False
                    except ObjectDoesNotExist, e:
                        context['raise'] = e.__str__()
                        context['status'] = False
                if 'delsub' in request.POST:
                    # Delete subproject
                    obj = Subproyecto.objects.get(proyecto_id=kwargs['project'], subproyecto_id=request.POST.get('sub'))
                    obj.status = 'DL'
                    obj.flag = False
                    obj.save()
                    # Delete all sectors of the Subproject
                    for x in Sectore.objects.filter(proyecto_id=kwargs['project'], subproyecto_id=request.POST.get('sub')):
                        x.status = 'DL'
                        x.flag = False
                        x.save()
                    # Delete Meter of Sales
                    for x in Metradoventa.objects.filter(proyecto_id=kwargs['project'], subproyecto_id=request.POST.get('sub')):
                        x.flag = False
                        x.save()
                    # Delete files of Sector
                    for x in SectorFiles.objects.filter(proyecto_id=kwargs['project'], subproyecto_id=request.POST.get('sub')):
                        x.flag = False
                        x.save()
                    # Delete Meter of Project
                    for x in MetProject.objects.filter(proyecto_id=kwargs['project'], subproyecto_id=request.POST.get('sub')):
                        x.flag = False
                        x.save()
                    # Delete Meter Project Nipples
                    for x in Nipple.objects.filter(proyecto_id=kwargs['project'], subproyecto_id=request.POST.get('sub')):
                        x.flag = False
                        x.save()
                    context['status'] = True
                if 'savedPurchase' in request.POST:
                    if 'editpurchse' in request.POST:
                        obj = PurchaseOrder.objects.get(pk=request.POST.get('editpurchse'))
                        form = PurchaseOrderForm(request.POST, request.FILES, instance=obj)
                    else:
                        form = PurchaseOrderForm(request.POST, request.FILES)
                    if form.is_valid():
                        add = form.save(commit=False)
                        add.project_id = kwargs['project']
                        add.save()
                        # save details purchase
                        details = json.loads(request.POST.get('details'))
                        if 'editpurchse' in request.POST:
                            # Delete details purchase
                            DetailsPurchaseOrder.objects.filter(purchase=request.POST.get('editpurchse')).delete()
                            purid = request.POST.get('editpurchse')
                        else:
                            purid = PurchaseOrder.objects.filter(project_id=kwargs['project']).order_by('-register')[:1][0].id
                        for x in details:
                            dp = DetailsPurchaseOrder()
                            dp.purchase_id = purid
                            dp.nropurchase = request.POST.get('nropurchase')
                            dp.description = x['description']
                            dp.unit_id = x['unit']
                            dp.delivery = globalVariable.format_str_date(x['date'])
                            dp.quantity = x['quantity']
                            dp.price = x['price']
                            dp.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                        context['raise'] = 'Form Incorrect.'
                if 'deletePurchase' in request.POST:
                    # Detele details
                    DetailsPurchaseOrder.objects.filter(purchase_id=request.POST.get('pk')).delete()
                    # Delete Bedside
                    PurchaseOrder.objects.get(pk=request.POST.get('pk')).delete()
                    context['status'] = True
                if 'loadPrices' in request.POST:
                    year = Proyecto.objects.get(pk=kwargs['project']).registrado.strftime('%Y')
                    prename = '/storage/projects/%s/%s/'%(year, kwargs['project'])
                    filename = uploadFiles.upload(
                        prename,
                        request.FILES['prices'],
                        {
                            'name': 'prices'
                        }
                    )
                    #print filename
                    sess = 'PRICES%s'%(kwargs['project'])
                    #del request.session[sess]
                    if sess in request.session:
                        del request.session[sess]
                    if not sess in request.session:
                        request.session[sess] = list()
                        request.session[sess] = uploadFiles.readQuotation(filename)
                        # print request.session[sess]
                    context['status'] = True
                if 'generateletter' in request.POST:
                    key = genkeys.generateLetterCode(kwargs['project'], request.session.get('company')['ruc'])
                    form = LetterForm(request.POST, request.FILES)
                    # print form
                    if form.is_valid():
                        add = form.save(commit=False)
                        add.letter_id = key
                        add.project_id = kwargs['project']
                        add.performed_id = request.user.get_profile().empdni_id
                        add.save()
                        context['code'] = key
                        context['status'] = True
                    else:
                        context['raise'] = 'Error form'
                        context['status'] = False
                if 'uploadLetter' in request.POST:
                    letter = Letter.objects.get(pk=request.POST.get('id'))
                    if request.user.get_profile().empdni_id == letter.performed_id or request.get_profile().empdni.charge.area.lower() == 'administrator':
                        letter.letter = request.FILES['letter']
                        letter.status = 'UP'
                        letter.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                if 'uploadLetterAnexo' in request.POST:
                    for x in range(0, int(request.POST.get('len'))):
                        letter = LetterAnexo()
                        letter.letter_id = request.POST.get('id')
                        letter.anexo = request.FILES['anexo%i'%x]
                        letter.save()
                    context['status'] = True
                if 'AnnulerLetter' in request.POST:
                    letter = Letter.objects.get(pk=request.POST.get('id'))
                    if request.user.get_profile().empdni_id == letter.performed_id or request.get_profile().empdni.charge.area.lower() == 'administrator':
                        letter.status = 'AN'
                        letter.flag = False
                        letter.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                        context['raise'] = 'Solo el responsable de crear la carta puede anularlo.'
                if 'editLetter' in request.POST:
                    letter = Letter.objects.get(pk=request.POST.get('letter_id'))
                    if request.user.get_profile().empdni_id == letter.performed_id:
                        form = LetterForm(request.POST, instance=letter)
                        if form.is_valid():
                            form.save()
                            context['status'] = True
                        else:
                            context['status'] = False
                if 'listAnexo' in request.POST:
                    anexo = LetterAnexo.objects.filter(letter_id=request.POST.get('id')).order_by('anexo')
                    context['list'] = [
                        {
                            'name': str(x.anexo).split('/')[-1],
                            'file': str(x.anexo)
                        }
                        for x in anexo
                    ]
                    context['status'] = True
                if 'closeStorage' in request.POST:
                    if request.user.get_profile().empdni.charge.area.lower() == 'almacen' or request.user.get_profile().empdni.charge.area.lower() == 'administrator':
                        try:
                            close = CloseProject.objects.get(project_id=kwargs['project'])
                            close.storageclose = True
                            close.performedstorage_id = request.user.get_profile().empdni_id
                            close.save()
                        except ObjectDoesNotExist, e:
                            close = CloseProject()
                            close.project_id = kwargs['project']
                            close.storageclose = True
                            close.performedstorage_id = request.user.get_profile().empdni_id
                            close.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                        context['raise'] = 'Permissions denied, close storage fail.'
                if 'letterdelivery' in request.POST:
                    if request.user.get_profile().empdni.charge.area.lower() == 'almacen' or request.user.get_profile().empdni.charge.area.lower() == 'administrator':
                        try:
                            close = CloseProject.objects.get(project_id=kwargs['project'])
                            close.letterdelivery = request.FILES['letter']
                            close.dateletter = globalVariable.date_now()
                            close.performedoperations_id = request.user.get_profile().empdni_id
                            close.save()
                        except ObjectDoesNotExist, e:
                            year = Proyecto.objects.get(proyecto_id=kwargs['project']).registrado.strftime('%Y')
                            close = CloseProject()
                            close.project_id = kwargs['project']
                            close.letterdelivery = request.FILES['letter']
                            close.dateletter = globalVariable.date_now()
                            close.performedoperations_id = request.user.get_profile().empdni_id
                            close.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                        context['raise'] = 'Permissions denied, upload letter fail.'
                if 'documentsCloser' in request.POST:
                    if request.user.get_profile().empdni.charge.area.lower() == 'operationes' or request.user.get_profile().empdni.charge.area.lower() == 'administrator':
                        if 'documents' in request.FILES or 'documents0' in request.FILES:
                            try:
                                close = CloseProject.objects.get(project_id=kwargs['project'])
                                close.documents = True
                                close.performeddocument_id = request.user.get_profile().empdni_id
                                close.save()
                                year = close.project.registrado.strftime('%Y')
                            except ObjectDoesNotExist, e:
                                year = Proyecto.objects.get(proyecto_id=kwargs['project']).registrado.strftime('%Y')
                                close = CloseProject()
                                close.project_id = kwargs['project']
                                close.documents = True
                                close.performeddocument_id = request.user.get_profile().empdni_id
                                close.save()
                            uri = '/storage/projects/%s/%s/closure/documents/'%(year, kwargs['project'])
                            if 'totalFiles' in request.POST:
                                for x in range(int(request.POST.get('totalFiles'))):
                                    filedocuments = uploadFiles.upload(uri, request.FILES['documents%i'%(x)])
                                    if filedocuments.split('.')[-1].lower() == 'rar':
                                        uploadFiles.descompressRAR(filedocuments, uri)
                                        uploadFiles.deleteFile(filedocuments)
                            else:
                                filedocuments = uploadFiles.upload(uri, request.FILES['documents'])
                                if filedocuments.split('.')[-1].lower() == 'rar':
                                    uploadFiles.descompressRAR(filedocuments, uri)
                                    uploadFiles.deleteFile(filedocuments)
                            context['status'] = True
                        else:
                            context['status'] = False
                            context['raise'] = 'File no exists'
                    else:
                        context['status'] = False
                        context['raise'] = 'Permissions denied'
                if 'saveAccounting' in request.POST:
                    if request.user.get_profile().empdni.charge.area.lower() == 'losgistica' or request.user.get_profile().empdni.charge.area.lower() == 'administrator':
                        close = None
                        try:
                            close = CloseProject.objects.get(project_id=kwargs['project'])
                        except ObjectDoesNotExist:
                            close = CloseProject()
                        close.project_id = kwargs['project']
                        close.accounting = True
                        close.tinvoice = request.POST.get('tinvoice')
                        close.tiva = request.POST.get('tiva')
                        close.otherin = request.POST.get('otherin')
                        close.otherout = request.POST.get('otherout')
                        close.retention = request.POST.get('retention')
                        close.performedaccounting_id = request.user.get_profile().empdni_id
                        if 'fileaccounting' in request.FILES:
                            close.fileaccounting = request.FILES.get('fileaccounting')
                            if request.FILES['fileaccounting'].name.split('.')[-1].lower() == 'rar':
                                pro = Proyecto.objects.get(proyecto_id=kwargs['project'])
                                path= 'storage/projects/%s/%s/closure/accounting/%s'%(pro.registrado.strftime('%Y'),pro.proyecto_id,request.FILES['fileaccounting'].name)
                                uri = '%s/%s'%(globalVariable.relative_path,path)
                                uploadFiles.descompressRAR(uri, uri.replace(uri.split('/')[-1],''))
                                uploadFiles.deleteFile(uri)
                        close.save()
                        context['status'] = True
                    else:
                        context['raise'] = 'Permissions denied'
                        context['status'] = False
                if 'genKeyConfClose' in request.POST:
                    if request.user.get_profile().empdni.charge.area.lower() == 'administrator' or request.user.get_profile().empdni.charge.area.lower() == 'ventas':
                        key = globalVariable.get_Token()
                        try:
                            close = CloseProject.objects.get(project_id=kwargs['project'])
                            close.closeconfirm = key
                            close.save()
                            context['key'] = key
                            context['status'] = True
                        except ObjectDoesNotExist, e:
                            context['status'] = False
                            context['raise'] = 'Fatal project close, missing other steps.'
                    else:
                        context['status'] = False
                        context['raise'] = 'Permissions denied'
                if 'closureproject' in request.POST:
                    if request.user.get_profile().empdni.charge.area.lower() == 'administrator' or request.user.get_profile().empdni.charge.area.lower() == 'ventas':
                        try:
                            close = CloseProject.objects.get(project_id=kwargs['project'])
                            if request.POST.get('keycon') == close.closeconfirm:
                                close.performedclose_id = request.user.get_profile().empdni_id
                                close.status = 'CO'
                                close.save()
                                context['status'] = True
                            else:
                                context['status'] = False
                                context['raise'] = 'El Código ingresado no es correcto.'
                        except ObjectDoesNotExist, e:
                            context['raise'] = str(e)
                            context['status'] = False
                    else:
                        context['status'] = False
                        context['raise'] = 'Permissions denied'
            except ObjectDoesNotExist, e:
                context['raise'] = e.__str__()
                context['status'] = False
            return self.render_to_json_response(context)


# Manager View Sectors
class SectorManage(JSONResponseMixin, View):
    template_name = 'sales/managersec.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        try:
            if request.is_ajax():
                try:
                    if 'percentsec' in request.GET:
                        items = MetProject.objects.filter(
                            proyecto_id=kwargs['pro'],
                            subproyecto_id=kwargs['sub']
                            if kwargs['sub'] != unicode(None) else None,
                            sector_id=kwargs['sec'])
                        titem = items.count()
                        attend = items.filter(tag='2').count()
                        partital = items.filter(tag='1').count()
                        attend = (attend + (partital/2))
                        if attend > 0:
                            percent = ((attend * 100) / titem)
                        else:
                            percent = 0
                        context['percent'] = '%.1f' % percent
                        context['status'] = True
                    if 'type' in request.GET:
                        if request.GET.get('type') == 'list':
                            obj = Metradoventa.objects.filter(
                                    proyecto_id=request.GET.get('pro'),
                                    subproyecto_id=request.GET.get('sub')
                                    if request.GET.get('sub') != '' else None,
                                    sector_id=request.GET.get('sec'),
                                    flag=True).order_by('materiales__matnom')
                            context['list'] = [
                                {
                                    'id': x.id,
                                    'materials_id': x.materiales_id,
                                    'name': x.materiales.matnom,
                                    'measure': x.materiales.matmed,
                                    'unit': x.materiales.unidad.uninom,
                                    'brand': x.brand.brand,
                                    'model': x.model.model,
                                    'quantity': x.cantidad,
                                    'price': x.precio,
                                    'sales': float(x.sales)
                                }
                                for x in obj
                            ]
                            context['status'] = True
                    if 'list-nip' in request.GET:
                        obj = Nipple.objects.filter(
                            proyecto_id=request.GET.get('pro'),
                            subproyecto_id=request.GET.get('sub')
                            if request.GET.get('sub') != '' else None,
                            sector_id=request.GET.get('sec'),
                            materiales_id=request.GET.get('mat'),
                            flag=True).order_by('metrado')
                        mat = MetProject.objects.get(
                                proyecto_id=request.GET.get('pro'),
                                subproyecto_id=request.GET.get('sub')
                                if request.GET.get('sub') != '' else None,
                                sector_id=request.GET.get('sec'),
                                materiales_id=request.GET.get('mat'))
                        if mat.quantityorder > 0:
                            attend = 'show'
                        else:
                            attend = 'hide'
                        context['list'] = [
                            {
                                'id': x.id,
                                'quantity': x.cantshop,
                                'diameter': x.materiales.matmed,
                                'measure': x.metrado,
                                'unit': 'cm',
                                'name': 'Niple%s %s, %s' % (
                                        's' if x.cantshop > 1 else '',
                                        globalVariable.tipo_nipples[x.tipo],
                                        x.tipo),
                                'comment': x.comment,
                                'materials': x.materiales_id,
                                'view': attend,
                                'tag': x.tag
                            } for x in obj
                        ]
                        ingress = 0
                        for x in obj:
                            ingress += (x.cantshop * x.metrado)
                        context['ingress'] = ingress
                        context['status'] = True
                    if 'withoutprices' in request.GET:
                        wprices = UpdateMetProject.objects.filter(
                            Q(proyecto_id=kwargs['pro']),
                            Q(subproyecto_id=kwargs['sub']
                                if kwargs['sub'] != unicode(None) else None),
                            Q(sector_id=kwargs['sec']), Q(price__lte=0) |
                            Q(sales__lte=0))
                        if wprices:
                            context['details'] = list(wprices.values(
                                'materials_id',
                                'materials__matnom',
                                'materials__matmed',
                                'brand_id',
                                'brand__brand',
                                'model_id',
                                'materials__unidad__uninom',
                                'price',
                                'sales'
                            ))
                            context['status'] = True
                        else:
                            context['details'] = list()
                            context['status'] = False
                            context['raise'] = ('No se han encontrado '
                                                'materiales sin precios')

                    if 'areasdsector' in request.GET:
                        context['dsectors'] = DSector.objects.filter(
                                            project_id=kwargs['pro']).count()
                        context['status'] = True
                except ObjectDoesNotExist, e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
            # block manager sector global
            context['project'] = Proyecto.objects.get(pk=kwargs['pro'])
            if kwargs['sub'] != unicode(None):
                context['subproject'] = Subproyecto.objects.get(
                                        proyecto_id=kwargs['pro'],
                                        subproyecto_id=kwargs['sub'])
            context['sector'] = Sectore.objects.get(
                                proyecto_id=kwargs['pro'],
                                subproyecto_id=kwargs['sub']
                                if kwargs['sub'] != unicode(None) else None,
                                sector_id=kwargs['sec'])
            context['planes'] = SectorFiles.objects.filter(
                                proyecto_id=kwargs['pro'],
                                subproyecto_id=kwargs['sub']
                                if kwargs['sub'] != unicode(None) else None,
                                sector_id=kwargs['sec'])
            context['system'] = Configuracion.objects.get(
                                    periodo=globalVariable.get_year)
            context['currency'] = Moneda.objects.filter(
                                    flag=True).order_by('moneda')
            context['exchange'] = TipoCambio.objects.filter(
                                    fecha=globalVariable.date_now())
            context['alerts'] = Alertasproyecto.objects.filter(
                                Q(proyecto_id=kwargs['pro']),
                                Q(subproyecto_id=kwargs['sub']
                                    if kwargs['sub'] != unicode(None)
                                    else None), Q(sector_id=kwargs['sec']),
                                Q(flag=True)).order_by('-registrado')
            # end block global
            # Deductive
            context['dsectors'] = Sectore.objects.filter(
                                    proyecto_id=kwargs['pro'],
                                    subproyecto_id=None)
            materials = Metradoventa.objects.filter(
                            proyecto_id=kwargs['pro'],
                            subproyecto_id=kwargs['sub']
                            if kwargs['sub'] != unicode(None) else None,
                            sector_id=kwargs['sec']
                            ).order_by('materiales__matnom')
            met = MetProject.objects.filter(
                    proyecto_id=kwargs['pro'],
                    subproyecto_id=kwargs['sub']
                    if kwargs['sub'] != unicode(None) else None,
                    sector_id=kwargs['sec']).order_by('materiales__matnom')
            if materials:
                    context['materials'] = materials
            if context['project'].status == 'AC' and met:
                del context['materials']
                data = list()
                for x in met:
                    stock = None
                    stock = Inventario.objects.filter(
                            materiales_id=x.materiales_id,
                            periodo=globalVariable.get_year)
                    if not stock:
                        stock = '-'
                    else:
                        stock = stock[0].stock
                    data.append(
                        {
                            'materiales_id': x.materiales_id,
                            'name': x.materiales.matnom,
                            'measure': x.materiales.matmed,
                            'unit': x.materiales.unidad.uninom,
                            'brand': x.brand.brand,
                            'brand_id': x.brand_id,
                            'model': x.model.model,
                            'model_id': x.model_id,
                            'quantity': x.quantityorder,
                            'cantidad': x.cantidad,
                            'price': x.precio,
                            'sales': x.sales,
                            'stock': stock,
                            'comment': x.comment,
                            'tag': x.tag
                        }
                    )
                context['meter'] = data
                context['niple'] = globalVariable.tipo_nipples
                context['store'] = Almacene.objects.filter(
                                    flag=True).order_by('nombre')
            return render_to_response(self.template_name,
                                      context,
                                      context_instance=RequestContext(request))
        except TemplateDoesNotExist, e:
            messages.error(request, 'Template not Exist %s' % e)
            raise Http404('Page Not Found')

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            context = dict()
            try:
                if 'type' in request.POST:
                    if request.POST.get('type') == 'plane':
                        form = SectorFilesForm(request.POST, request.FILES)
                        if form.is_valid():
                            form.save()
                        context['status'] = True
                    if request.POST.get('type') == 'delplane':
                        obj = SectorFiles.objects.get(
                                proyecto_id=request.POST.get('pro'),
                                subproyecto_id=request.POST.get('sub')
                                if request.POST.get('sub') else None,
                                sector_id=request.POST.get('sec'),
                                files=request.POST.get('files'))
                        obj.delete()
                        uploadFiles.removeTmp('%s/%s' % (
                            globalVariable.relative_path,
                            request.POST.get('files')))
                        context['status'] = True
                    if request.POST.get('type') == 'add':
                        if 'edit' in request.POST:
                            obj = Metradoventa.objects.get(
                                proyecto_id=request.POST.get('proyecto'),
                                subproyecto_id=request.POST.get('subproyecto')
                                if request.POST.get('subproyecto') else None,
                                sector_id=request.POST.get('sector'),
                                materiales_id=request.POST.get('materiales'))
                            form = MetradoventaForm(request.POST, instance=obj)
                        else:
                            form = MetradoventaForm(request.POST)
                        if form.is_valid():
                            if 'edit' in request.POST:
                                edit = form.save(commit=False)
                                edit.sales = float(request.POST.get('sales'))
                                edit.save()
                            else:
                                obj = Metradoventa.objects.filter(
                                    proyecto_id=kwargs['pro'],
                                    subproyecto_id=request.POST.get(
                                                                'subproyecto')
                                    if request.POST.get('subproyecto')
                                    else None, sector_id=kwargs['sec'],
                                    materiales_id=request.POST.get(
                                                                'materiales'))
                                if obj:
                                    obj[0].cantidad = (
                                        obj[0].cantidad + float(
                                            request.POST.get('cantidad')))
                                    obj[0].precio = request.POST.get('precio')
                                    obj[0].sales = float(
                                                    request.POST.get('sales'))
                                    obj[0].save()
                                else:
                                    add = form.save(commit=False)
                                    add.sales = float(
                                                    request.POST.get('sales'))
                                    # print request.POST.get('sales')
                                    add.save()
                                if not 'edit' in request.POST and 'details' in request.POST:
                                    # save Details Material Group
                                    for x in json.loads(
                                            request.POST.get('details')):
                                        try:
                                            obj = Metradoventa.objects.get(
                                                proyecto_id=kwargs['pro'],
                                                subproyecto_id=request.POST.get('subproyecto') if request.POST.get('subproyecto') else None,
                                                sector_id=kwargs['sec'],
                                                materiales_id=x['materials']
                                            )
                                            obj.cantidad = obj.cantidad + (float(x['quantity']) * float(request.POST.get('cantidad')))
                                            obj.save()
                                        except ObjectDoesNotExist:
                                            obj = Metradoventa()
                                            obj.proyecto_id = kwargs['pro']
                                            obj.subproyecto_id = (
                                                request.POST.get('subproyecto')
                                                if request.POST.get(
                                                    'subproyecto') else '')
                                            obj.sector_id = kwargs['sec']
                                            obj.materiales_id = x['materials']
                                            obj.cantidad = (
                                                float(
                                                    x['quantity']) * float(
                                                    request.POST.get(
                                                                'cantidad')))
                                            obj.precio = 0
                                            obj.sales = 0
                                            obj.brand_id = 'BR000'
                                            obj.model_id = 'MO000'
                                            obj.flag = True
                                            obj.save()
                            context['status'] = True
                        else:
                            context['status'] = False
                    if request.POST.get('type') == 'del':
                        obj = Metradoventa.objects.filter(
                                proyecto_id=request.POST.get('pro'),
                                subproyecto_id=request.POST.get('sub')
                                if request.POST.get('sub') else None,
                                sector_id=request.POST.get('sec'),
                                materiales_id=request.POST.get('materials'))
                        obj.delete()
                        context['status'] = True
                    if request.POST.get('type') == 'killdata':
                        obj = Metradoventa.objects.filter(
                                proyecto_id=request.POST.get('pro'),
                                subproyecto_id=request.POST.get('sub')
                                if request.POST.get('sub') else None,
                                sector_id=request.POST.get('sec'))
                        obj.delete()
                        context['status'] = True
                # else:
                #     context['status'] = False
                if 'addnip' in request.POST:
                    if 'id' in request.POST:
                        obj = Nipple.objects.get(pk=request.POST.get('id'))
                        form = NippleForm(request.POST, instance=obj)
                    else:
                        form = NippleForm(request.POST)
                    if form.is_valid():
                        form.save()
                        context['status'] = True
                if 'delnip' in request.POST:
                    obj = Nipple.objects.get(pk=request.POST.get('pk'))
                    obj.delete()
                    context['status'] = True
                if 'delnipall' in request.POST:
                    Nipple.objects.filter(
                        proyecto_id=request.POST.get('proyecto'),
                        subproyecto_id=request.POST.get('subproyecto')
                        if request.POST.get('subproyecto') != '' else None,
                        sector_id=request.POST.get('sector'),
                        materiales_id=request.POST.get('materiales')).delete()
                    context['status'] = True
                if 'upcomment' in request.POST:
                    obj = MetProject.objects.get(
                        proyecto_id=request.POST.get('pro'),
                        subproyecto_id=request.POST.get('sub')
                        if request.POST.get('sub') != '' else None,
                        sector_id=request.POST.get('sec'),
                        materiales_id=request.POST.get('mat'))
                    if obj:
                        obj.comment = request.POST.get('comment')
                        obj.save()
                        context['status'] = True
                if 'saveorders' in request.POST:
                    form = addOrdersForm(request.POST, request.FILES)
                    if form.is_valid():
                        # save bedside Orders
                        add = form.save(commit=False)
                        id = genkeys.GenerateIdOrders()
                        add.pedido_id = id
                        add.status = 'PE'
                        add.save()
                        # save to detail
                        details = json.loads(request.POST.get('details'))
                        # print details
                        for x in details:
                            obj = Detpedido()
                            obj.pedido_id = id
                            obj.materiales_id = x['idmat']
                            obj.cantidad = x['quantity']
                            obj.cantshop = x['quantity']
                            br = search.searchBrands()
                            br.brand = x['brand']
                            obj.brand_id = br.autoDetected()['pk'].strip()
                            mo = search.searchModels()
                            mo.model = x['model']
                            obj.model_id = mo.autoDetected()['pk'].strip()
                            obj.comment = x['comment']
                            obj.save()
                            # update quantity in Metproject
                            pro = MetProject.objects.get(
                                    proyecto_id=kwargs['pro'],
                                    subproyecto_id=kwargs['sub']
                                    if kwargs['sub'] != unicode(kwargs['sub'])
                                    else None, sector_id=kwargs['sec'],
                                    materiales_id=x['idmat'])
                            if pro.quantityorder == pro.cantidad:
                                pro.quantityorder = (
                                                pro.cantidad - x['quantity'])
                            else:
                                pro.quantityorder = (
                                    pro.quantityorder - x['quantity'])
                            if pro.quantityorder > 0 and (
                                    pro.quantityorder < pro.cantidad):
                                pro.tag = '1'
                            else:
                                pro.tag = '2'
                            pro.save()
                        # save to nipples
                        nipples = json.loads(request.POST.get('nipples'))
                        for x in nipples:
                            obj = Niple()
                            obj.pedido_id = id
                            obj.proyecto_id = request.POST.get('proyecto')
                            obj.subproyecto_id = request.POST.get(
                                                                'subproyecto')
                            obj.sector_id = request.POST.get('sector')
                            obj.empdni = request.user.get_profile().empdni_id
                            obj.materiales_id = x['idmat']
                            obj.cantidad = x['quantity']
                            obj.cantshop = x['quantity']
                            obj.metrado = x['meter']
                            obj.tipo = x['type'].strip()
                            obj.comment = x['comment'].strip()
                            obj.save()
                            # update table od nipples - operations
                            nip = Nipple.objects.get(
                                    proyecto_id=request.POST.get('proyecto'),
                                    subproyecto_id=request.POST.get(
                                                                'subproyecto')
                                    if request.POST.get('subproyecto') != ''
                                    else None, sector_id=request.POST.get(
                                                                    'sector'),
                                    materiales_id=x['idmat'], id=x['idnip'])
                            if nip.cantshop == nip.cantidad:
                                nip.cantshop = (nip.cantidad - x['quantity'])
                            else:
                                nip.cantshop = (nip.cantshop - x['quantity'])
                            if nip.cantshop > 0 and(
                                    nip.cantshop < nip.cantidad):
                                nip.tag = '1'
                            else:
                                nip.tag = '2'
                            nip.save()
                        context['nro'] = id
                        context['status'] = True
                if 'approvedadditional' in request.POST:
                    # details = json.loads(request.POST.get('details'))
                    details = Metradoventa.objects.filter(
                                proyecto_id=kwargs['pro'],
                                subproyecto_id=kwargs['sub'] if kwargs[
                                    'sub'] != unicode(None) else None,
                                sector_id=kwargs['sec'])
                    if details:
                        for x in details:
                            # save Metprojet addtional
                            obj = MetProject()
                            obj.proyecto_id = kwargs['pro']
                            obj.subproyecto_id = kwargs['sub'] if kwargs[
                                'sub'] != unicode(None) else ''
                            obj.sector_id = kwargs['sec']
                            obj.materiales_id = x.materiales_id
                            obj.cantidad = x.cantidad
                            obj.quantityorder = x.cantidad
                            obj.precio = x.precio
                            obj.brand_id = x.brand_id
                            obj.model_id = x.model_id
                            obj.sales = x.sales
                            obj.flag = True
                            obj.save()
                    context['status'] = True
                if 'modifystart' in request.POST:
                    update = UpdateMetProject.objects.filter(
                            proyecto_id=kwargs['pro'],
                            subproyecto_id=kwargs['sub'] if kwargs[
                                'sub'] != unicode(None) else None,
                            sector_id=kwargs['sec'],
                            flag=True
                        ).order_by('materials__matnom')
                    amountp = 0
                    list_ = list()
                    if not update:
                        for x in MetProject.objects.filter(
                            proyecto_id=kwargs['pro'],
                            subproyecto_id=kwargs['sub']
                            if kwargs['sub'] != unicode(None)
                            else None, sector_id=kwargs['sec']).order_by(
                                'materiales__matnom'):
                            obj = UpdateMetProject()
                            obj.proyecto_id = kwargs['pro']
                            obj.subproyecto_id = kwargs['sub'] if kwargs[
                                'sub'] != unicode(None) else ''
                            obj.sector_id = kwargs['sec']
                            obj.materials_id = x.materiales_id
                            obj.brand_id = x.brand_id
                            obj.model_id = x.model_id
                            obj.quantity = x.cantidad
                            obj.price = x.precio
                            obj.sales = x.sales
                            obj.comment = x.comment
                            obj.quantityorders = x.quantityorder
                            obj.tag = x.tag
                            obj.save()
                            list_.append(
                                {
                                    'materials': x.materiales_id,
                                    'name': x.materiales.matnom,
                                    'measure': x.materiales.matmed,
                                    'unit': x.materiales.unidad.uninom,
                                    'brand_id': x.brand_id,
                                    'model_id': x.model_id,
                                    'brand': x.brand.brand,
                                    'model': x.model.model,
                                    'quantity': x.cantidad,
                                    'orders': x.quantityorder,
                                    'price': x.precio,
                                    'sales': float(x.sales),
                                    'amount': x.amountPurchase,
                                    'tag': x.tag
                                }
                            )
                            amountp += x.amountPurchase
                    else:
                        for x in update:
                            list_.append(
                                {
                                    'materials': x.materials_id,
                                    'name': unicode(x.materials.matnom),
                                    'measure': unicode(x.materials.matmed),
                                    'unit': x.materials.unidad.uninom,
                                    'brand_id': x.brand_id,
                                    'model_id': x.model_id,
                                    'brand': x.brand.brand,
                                    'model': x.model.model,
                                    'quantity': x.quantity,
                                    'orders': x.quantityorders,
                                    'price': x.price,
                                    'sales': float(x.sales),
                                    'amount': '{0:.3f}'.format(x.amount),
                                    'tag': x.tag
                                }
                            )
                            amountp += x.amount
                    context['details'] = list_
                    context['apurchase'] = amountp
                    context['listBrand'] = [
                        {
                            'brand_id': x.brand_id,
                            'brand': x.brand
                        }
                        for x in Brand.objects.filter(
                                            flag=True).order_by('brand')
                    ]
                    context['listModel'] = [
                        {
                            'model_id': x.model_id,
                            'model': x.model
                        }
                        for x in Model.objects.filter(
                                                flag=True).order_by('model')
                    ]
                    context['status'] = True
                if 'updatematerialMeter' in request.POST:
                    obj = UpdateMetProject.objects.get(
                            proyecto_id=kwargs['pro'],
                            subproyecto_id=kwargs['sub']
                            if kwargs['sub'] != unicode(None) else None,
                            sector_id=kwargs['sec'],
                            materials_id=request.POST.get('materials'),
                            brand_id=request.POST.get('brands'),
                            model_id=request.POST.get('models'),
                            flag=True)
                    obj.brand_id = request.POST.get('brand')
                    obj.model_id = request.POST.get('model')
                    pending = 0
                    metrado = 0
                    try:
                        meter = MetProject.objects.get(
                            proyecto_id=kwargs['pro'],
                            subproyecto_id=kwargs['sub']
                            if kwargs['sub'] != unicode(None) else None,
                            sector_id=kwargs['sec'],
                            materiales_id=request.POST.get('materials'),
                            brand_id=request.POST.get('brands'),
                            model_id=request.POST.get('models'))
                        pending = float(meter.quantityorder)
                        metrado = meter.cantidad
                    except ObjectDoesNotExist:
                        pending = 0
                    quantity = float(request.POST.get('quantity'))
                    if quantity > metrado:
                        obj.quantityorders += (quantity - metrado)
                    if metrado > quantity:
                        obj.quantityorders -= (metrado - quantity)
                        if obj.quantityorders <= 0:
                            obj.quantityorders = 0
                            obj.tag = '2'
                    if obj.quantityorders > 0:
                        obj.tag = '1'

                    orders = obj.quantityorders

                    # print 'pending proccess', pending
                    # if quantity > obj.quantity:
                    #     if pending > 0 or pending == 0:
                    #         if orders == obj.quantity:
                    #             obj.tag = '0'
                    #         else:
                    #             obj.tag = '1'
                    #         obj.quantityorders = (
                    #             pending + (quantity - metrado))
                    #     if pending == quantity:
                    #         obj.quantityorders = quantity
                    #         obj.tag = '0'
                    # elif quantity < obj.quantity:
                    #     if pending == 0:
                    #         obj.tag = '2'
                    #         obj.quantityorders = pending
                    #     if pending >= quantity:
                    #         obj.quantityorders = quantity
                    #         obj.tag = '0'
                    #     if pending > 0 and pending <= quantity:
                    #         obj.quantityorders = (
                    #                 pending - (obj.quantity - quantity))
                    #         if (pending - (obj.quantity - quantity)) == 0:
                    #             obj.tag = '2'
                    #         else:
                    #             obj.tag = '1'
                    # else:
                    #     if pending == 0 and quantity == metrado:
                    #         obj.tag = '2'
                    #     else:
                    #         obj.tag = '0'
                    obj.quantity = quantity
                    if request.user.get_profile().empdni.charge.area.lower() != 'operaciones':
                        obj.price = float(request.POST.get('price'))
                        obj.sales = float(request.POST.get('sales'))
                    obj.save()
                    context['status'] = True
                if 'deletematerialMeter' in request.POST:
                    obj = UpdateMetProject.objects.get(
                        proyecto_id=kwargs['pro'],
                        subproyecto_id=kwargs['sub'] if kwargs['sub'] != unicode(None) else None,
                        sector_id=kwargs['sec'],
                        materials_id=request.POST.get('materials'),
                        brand_id=request.POST.get('brand'),
                        model_id=request.POST.get('model'),
                        flag=True
                    )
                    obj.delete()
                    context['status'] = True
                if 'deleteallupdatemeter' in request.POST:
                    obj = UpdateMetProject.objects.filter(
                        proyecto_id=kwargs['pro'],
                        subproyecto_id=kwargs['sub'] if kwargs['sub'] != unicode(None) else None,
                        sector_id=kwargs['sec'],
                        flag=True
                    )
                    obj.delete()
                    context['status'] = True
                if 'addupdatemeter' in request.POST:
                    try:
                        obj = UpdateMetProject.objects.get(
                                proyecto_id=kwargs['pro'],
                                subproyecto_id=kwargs['sub'] if kwargs['sub'] != unicode(None) else None,
                                sector_id=kwargs['sec'],
                                materials_id=request.POST.get('materials'),
                                brand_id=request.POST.get('brand'),
                                model=request.POST.get('model'),
                                flag=True
                            )
                        context['status'] = False
                        # quantity = (obj.quantity + float(request.POST.get('quantity')))
                        context['raise'] = 'El Material ya se encuentra ingresado editelo.'
                    except ObjectDoesNotExist, e:
                        obj = UpdateMetProject()
                        obj.proyecto_id = kwargs['pro']
                        obj.subproyecto_id = kwargs['sub'] if kwargs['sub'] != unicode(None) else ''
                        obj.sector_id = kwargs['sec']
                        obj.materials_id = request.POST.get('materials')
                        obj.brand_id = request.POST.get('brand')
                        obj.model_id = request.POST.get('model')
                        obj.quantity = float(request.POST.get('quantity'))
                        obj.price = float(request.POST.get('price'))
                        obj.sales = float(request.POST.get('sales'))
                        obj.comment = ''
                        obj.quantityorders = request.POST.get('quantity')
                        obj.tag = '0'
                        obj.flag = True
                        obj.save()
                        context['status'] = True
                    # save details materail group
                    if 'details' in request.POST:
                        # print request.POST.get('details')
                        for x in json.loads(request.POST.get('details')):
                            try:
                                obj = UpdateMetProject.objects.get(
                                    proyecto_id=kwargs['pro'],
                                    subproyecto_id=kwargs['sub']
                                    if kwargs['sub'] != unicode(None)
                                    else None, sector_id=kwargs['sec'],
                                    materials_id=x['materials'], flag=True)
                                quantity = (
                                        obj.quantity + (
                                            float(x['quantity']) * float(
                                                request.POST.get('quantity'))))
                                if obj.tag != '0':
                                    if obj.tag == '1':
                                        if obj.quantity > quantity:
                                            obj.tag = '2'
                                    elif obj.tag == '2':
                                        if obj.quantity < quantity:
                                            obj.tag = '1'
                                    else:
                                        obj.tag = '0'
                                obj.quantity = quantity
                                obj.save()
                            except ObjectDoesNotExist:
                                obj = UpdateMetProject()
                                obj.proyecto_id = kwargs['pro']
                                obj.subproyecto_id = kwargs['sub'] if kwargs['sub'] != unicode(None) else ''
                                obj.sector_id = kwargs['sec']
                                obj.materials_id = x['materials']
                                obj.brand_id = 'BR000'
                                obj.model_id = 'MO000'
                                obj.quantity = (float(x['quantity']) * float(request.POST.get('quantity')))
                                if request.user.get_profile().empdni.charge.area.lower() != 'operaciones':
                                    obj.price = float(
                                                    request.POST.get('price'))
                                    obj.sales = float(
                                                    request.POST.get('sales'))
                                else:
                                    obj.price = 0
                                    obj.sales = 0
                                obj.comment = ''
                                obj.quantityorders = x['quantity']
                                obj.tag = '0'
                                obj.flag = True
                                obj.save()
                        context['status'] = True
                if 'generateDeductiveOne' in request.POST:
                    key = genkeys.GenerateIdDeductive()
                    # Generate bedside deductive
                    bed = Deductive()
                    bed.deductive_id = key
                    bed.proyecto_id = kwargs['pro']
                    bed.subproyecto_id = kwargs['sub'] if kwargs['sub'] != unicode(None) else ''
                    bed.sector_id = kwargs['sec']
                    bed.rtype = 'LO'
                    bed.relations = ''
                    bed.save()
                    # deductive inputs details
                    details = json.loads(request.POST.get('details'))
                    for x in details:
                        detin = DeductiveInputs()
                        detin.deductive_id = key
                        detin.materials_id = x['materials']
                        sr = search.searchBrands()
                        sr.brand = x['brand']
                        sr = sr.autoDetected()
                        detin.brand_id = sr['pk']
                        sr = search.searchModels()
                        sr.model = x['model']
                        sr = sr.autoDetected()
                        detin.model_id = sr['pk']
                        detin.quantity = x['quantity']
                        detin.price = x['price']
                        detin.related = x['output']
                        if x['output'] != '':
                            det = x['output'].split(',')
                            for o in det:
                                row = o.split('|')
                                deto = DeductiveOutputs()
                                deto.deductive_id = key
                                deto.materials_id = row[0]
                                deto.quantity = row[1]
                                deto.save()
                        detin.save()
                    context['deductive'] = key
                    context['status'] = True
                if 'approvedModifyFinal' in request.POST:
                    # request parameters #  first save history sector
                    # meter = json.loads(request.POST.get('meter'))
                    # history = json.loads(request.POST.get('history'))
                    # delete details sector of the meter

                    sec = MetProject.objects.filter(
                            proyecto_id=kwargs['pro'],
                            subproyecto_id=kwargs['sub']
                            if kwargs['sub'] != unicode(None)
                            else None, sector_id=kwargs['sec'])
                    token = globalVariable.get_Token()
                    meter = sec
                    for x in sec:
                        h = HistoryMetProject()
                        h.token = token
                        h.proyecto_id = kwargs['pro']
                        h.subproyecto_id = kwargs['sub'] if kwargs['sub'] != unicode(None) else ''
                        h.sector_id = kwargs['sec']
                        h.materials_id = x.materiales_id
                        h.quantity = x.cantidad
                        h.brand_id = x.brand_id
                        h.model_id = x.model_id
                        h.price = x.precio
                        h.sales = x.sales
                        h.comment = x.comment
                        h.quantityorders = x.quantityorder if x.tag != '0' else x.cantidad
                        h.tag = x.tag
                        h.flag = x.flag
                        h.save()
                        # if x.materials_id[:3] == '115':
                        #     for n in Nipple.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id= kwargs['sub'] if kwargs['sub'] != unicode(None) else '', sector_id=kwargs['sec'], materiales_id=x.materials_id):
                        #         n.flag = False
                        #         n.save()
                        x.delete()
                    # save new details sector of the meter
                    update = UpdateMetProject.objects.filter(
                                proyecto_id=kwargs['pro'],
                                subproyecto_id=kwargs['sub']
                                if kwargs['sub'] != unicode(None)
                                else None, sector_id=kwargs['sec'])
                    for x in update:
                        s = MetProject()
                        s.proyecto_id = kwargs['pro']
                        s.subproyecto_id = kwargs['sub'] if kwargs['sub'] != unicode(None) else ''
                        s.sector_id = kwargs['sec']
                        s.materiales_id = x.materials_id
                        s.brand_id = x.brand_id
                        s.model_id = x.model_id
                        s.cantidad = x.quantity
                        s.precio = x.price
                        s.sales = x.sales
                        s.flag = True
                        s.comment = x.comment
                        s.quantityorder = x.quantityorders
                        if x.quantityorders  == 0:
                            s.tag = '2'
                        if x.quantityorders > 0 and x.quantityorders < x.quantity:
                            s.tag = '1'
                        if x.quantityorders == x.quantity:
                            s.tag = '0'
                        dev = meter.filter(materiales_id=x.materials_id)
                        if dev.count():
                            if x.quantity > dev[0].cantidad:
                                d = RestoreStorage()
                                d.token = token
                                d.proyecto_id = kwargs['pro']
                                d.subproyecto_id = kwargs['sub'] if kwargs['sub'] != unicode(None) else ''
                                d.sector_id = kwargs['sec']
                                d.materials_id = x.materials_id
                                d.brand_id = x.brand_id
                                d.model_id = x.model_id
                                d.quantity = float(x.quantity - dev[0].cantidad)
                                d.price = x.price
                                d.save()
                        s.save()
                        x.delete()
                    # up = UpdateMetProject.objects.filter(
                    #         proyecto_id=kwargs['pro'],
                    #         subproyecto_id=kwargs['sub']
                    #         if kwargs['sub'] != unicode(None)
                    #         else None, sector_id=kwargs['sec'])
                    # for x in up:
                    #     x.delete()
                    context['status'] = True
                if 'searchdescdeductive' in request.POST:
                    list_ = list()
                    obj = None
                    if request.POST.get('typedeductive') == 'ONE':
                        obj = MetProject.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id=None,sector_id=request.POST.get('sector') ,materiales__matnom__icontains=request.POST.get('text'))
                    elif request.POST.get('typedeductive') == 'CUS':
                        sec = request.POST.get('cus').split(',')
                        obj = MetProject.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id=None,sector_id__in=sec,materiales__matnom__icontains=request.POST.get('text'))
                    elif request.POST.get('typedeductive') == 'ALL':
                        obj = MetProject.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id=None,materiales__matnom__icontains=request.POST.get('text'))
                    o = obj
                    obj = obj.order_by('materiales__materiales_id').distinct('materiales__materiales_id')
                    for x in obj:
                        quan = o.filter(materiales_id=x.materiales_id).aggregate(quantity=Sum('cantidad'))
                        o = o.filter(materiales_id=x.materiales_id)
                        for c in o:
                            print c.subproyecto_id
                        list_.append({'materials': x.materiales_id, 'name': x.materiales.matnom, 'measure': x.materiales.matmed, 'unit': x.materiales.unidad.uninom, 'quantity': quan['quantity'], 'price': x.precio})
                    context['list'] = list_
                    context['status'] = True
                if 'registerdeductivegl' in request.POST:
                    key = genkeys.GenerateIdDeductive()
                    bed = Deductive()
                    bed.deductive_id = key
                    bed.proyecto_id = kwargs['pro']
                    bed.subproyecto_id = kwargs['sub'] if kwargs['sub'] != unicode(None) else ''
                    bed.sector_id = kwargs['sec']
                    bed.rtype = request.POST.get('rtype')
                    bed.relations = request.POST.get('relations') if 'relations' in request.POST else ''
                    bed.save()
                    # save deductive inputs details
                    inputs = json.loads(request.POST.get('inputs'))
                    for x in inputs:
                        detin = DeductiveInputs()
                        detin.deductive_id = key
                        detin.materials_id = x['materials']
                        detin.brand_id = 'BR000'
                        detin.model_id = 'MO000'
                        detin.quantity = x['quantity']
                        detin.price = x['price']
                        detin.related = x['output']
                        detin.save()
                    # save deductive outputs details
                    outputs = json.loads(request.POST.get('outputs'))
                    if outputs:
                        for x in outputs:
                            dout = DeductiveOutputs()
                            dout.deductive_id = key
                            dout.materials_id = x['materials']
                            dout.quantity = x['quantity']
                            dout.save()
                    # register thoe history of sectors if relations
                    relations = None
                    token = globalVariable.get_Token()
                    if request.POST.get('rtype') == 'ONE':
                        relations = MetProject.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id=kwargs['sub'] if kwargs['sub'] != unicode(kwargs['sub']) else '',sector_id=request.POST.get('relations'))
                        for x in relations:
                            h = HistoryMetProject()
                            h.token = token
                            h.proyecto_id = kwargs['pro']
                            h.subproyecto_id = ''
                            h.sector_id = request.POST.get('relations')
                            h.materials_id = x.materiales_id
                            h.quantity = x.cantidad
                            h.brand_id = x.brand_id
                            h.model_id = x.model_id
                            h.price = x.precio
                            h.comment = x.comment
                            h.quantityorders = x.quantityorder
                            h.tag = x.tag
                            h.flag = x.flag
                            h.save()
                    elif request.POST.get('rtype') == 'CUS':
                        for s in json.loads(request.POST.get('relations')):
                            #print s
                            relations = MetProject.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id=None,sector_id=s)
                            for x in relations:
                                h = HistoryMetProject()
                                h.token = token
                                h.proyecto_id = kwargs['pro']
                                h.subproyecto_id = kwargs['sub'] if kwargs['sub'] != unicode(None) else ''
                                h.sector_id = s
                                h.materials_id = x.materiales_id
                                h.quantity = x.cantidad
                                h.brand_id = x.brand_id
                                h.model_id = x.model_id
                                h.price = x.precio
                                h.comment = x.comment
                                h.quantityorders = x.quantityorder
                                h.tag = x.tag
                                h.flag = x.flag
                                h.save()
                    elif request.POST.get('rtype') == 'ALL':
                        sectors = Sectore.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id='')
                        for s in sectors:
                            relations = MetProject.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id=kwargs['sub'] if kwargs['sub'] != unicode(kwargs['sub']) else '',sector_id=s.sector_id)
                            for x in relations:
                                h = HistoryMetProject()
                                h.token = token
                                h.proyecto_id = kwargs['pro']
                                h.subproyecto_id = kwargs['sub'] if kwargs['sub'] != unicode(None) else ''
                                h.sector_id = s
                                h.materials_id = x.materiales_id
                                h.quantity = x.cantidad
                                h.brand_id = x.brand_id
                                h.model_id = x.model_id
                                h.price = x.precio
                                h.comment = x.comment
                                h.quantityorders = x.quantityorder
                                h.tag = x.tag
                                h.flag = x.flag
                                h.save()
                    # for the registers
                    # delete materials in list outputs
                    if outputs:
                        if request.POST.get('rtype') == 'ONE':
                            for x in outputs:
                                outs = MetProject.objects.get(proyecto_id=kwargs['pro'],sector_id=request.POST.get('relations'), materiales_id=x['materials'])
                                if (outs.cantidad - x['quantity']) > 0:
                                    d = RestoreStorage()
                                    d.token = token
                                    d.proyecto_id = kwargs['pro']
                                    d.subproyecto_id = ''
                                    d.sector_id = request.POST.get('relations')
                                    d.materials_id = x['materials']
                                    d.brand_id = 'BR000'
                                    d.model_id = 'MO000'
                                    d.quantity = float((outs.cantidad - x['quantity']))
                                    d.price = x['price']
                                    d.save()
                                    # update qunaity
                                    outs.cantidad = (outs.cantidad - x['quantity'])
                                    outs.save()
                                else:
                                    outs.delete()
                        elif request.POST.get('rtype') == 'CUS':
                            #print outputs
                            for x in outputs:
                                for o in json.loads(request.POST.get('relations')):
                                    #print o, x['materials']
                                    outs = MetProject.objects.get(proyecto_id=kwargs['pro'],sector_id=o, materiales_id=x['materials'])
                                    if (outs.cantidad - x['quantity']) > 0:
                                        d = RestoreStorage()
                                        d.token = token
                                        d.proyecto_id = kwargs['pro']
                                        d.subproyecto_id = ''
                                        d.sector_id = o
                                        d.materials_id = x['materials']
                                        d.brand_id = 'BR000'
                                        d.model_id = 'MO000'
                                        d.quantity = float((outs.cantidad - x['quantity']))
                                        d.price = x['price']
                                        d.save()
                                        outs.cantidad = (outs.cantidad - x['quantity'])
                                        outs.save()
                                    else:
                                        outs.delete()
                        elif request.POST.get('rtype') == 'ALL':
                            for x in outputs:
                                for o in Sectore.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id=''):
                                    outs = MetProject.objects.get(proyecto_id=kwargs['pro'],sector_id=o.sector_id, materiales_id=x['materials'])
                                    if (outs.cantidad - x['quantity']) > 0:
                                        d = RestoreStorage()
                                        d.token = token
                                        d.proyecto_id = kwargs['pro']
                                        d.subproyecto_id = ''
                                        d.sector_id = request.POST.get('relations')
                                        d.materials_id = x['materials']
                                        d.brand_id = 'BR000'
                                        d.model_id = 'MO000'
                                        d.quantity = float((outs.cantidad - x['quantity']))
                                        d.price = x['price']
                                        d.save()
                                        outs.cantidad = (outs.cantidad - x['quantity'])
                                        outs.save()
                                    else:
                                        outs.delete()

                    if request.POST.get('rtype') == 'ONE':
                        for x in inputs:
                            s = MetProject()
                            s.proyecto_id = kwargs['pro']
                            s.subproyecto_id = ''
                            s.sector_id = request.POST.get('relations')
                            s.materiales_id = x['materials']
                            s.brand_id = 'BR000'
                            s.model_id = 'MO000'
                            s.cantidad = x['quantity']
                            s.precio = x['price']
                            s.flag = True
                            s.comment = ''
                            s.quantityorder = x['quantity']
                            # s.tag = x['tag']
                            s.save()
                    if request.POST.get('rtype') == 'CUS':
                        # print json.loads(request.POST.get('relations'))
                        for sn in json.loads(request.POST.get('relations')):
                            # print s, 'CUS '
                            for x in inputs:
                                s = MetProject()
                                s.proyecto_id = kwargs['pro']
                                s.subproyecto_id = ''
                                s.sector_id = sn
                                s.materiales_id = x['materials']
                                s.brand_id = 'BR000'
                                s.model_id = 'MO000'
                                s.cantidad = float(x[
                                    'quantity']) / len(
                                    json.loads(request.POST.get('relations')))
                                s.precio = x['price']
                                s.flag = True
                                s.comment = ''
                                s.quantityorder = float(x[
                                    'quantity']) / len(
                                    json.loads(request.POST.get('relations')))
                                # s.tag = x['tag']
                                s.save()
                    if request.POST.get('rtype') == 'ALL':
                        ln = Sectore.objects.filter(
                            proyecto_id=kwargs['pro'],
                            subproyecto_id='').count()
                        for s in Sectore.objects.filter(
                                proyecto_id=kwargs['pro'], subproyecto_id=''):
                            for x in inputs:
                                s = MetProject()
                                s.proyecto_id = kwargs['pro']
                                s.subproyecto_id = ''
                                s.sector_id = s.sector_id
                                s.materiales_id = x['materials']
                                s.brand_id = 'BR000'
                                s.model_id = 'MO000'
                                s.cantidad = float(x['quantity']) / ln
                                s.precio = x['price']
                                s.flag = True
                                s.comment = ''
                                s.quantityorder = float(x['quantity']) / ln
                                # s.tag = x['tag']
                                s.save()

                    for x in inputs:
                        s = MetProject()
                        s.proyecto_id = kwargs['pro']
                        s.subproyecto_id = kwargs[
                            'sub'] if kwargs['sub'] != unicode(None) else ''
                        s.sector_id = kwargs['sec']
                        s.materiales_id = x['materials']
                        s.brand_id = 'BR000'
                        s.model_id = 'MO000'
                        s.cantidad = x['quantity']
                        s.precio = x['price']
                        s.flag = True
                        s.comment = ''
                        s.quantityorder = x['quantity']
                        # s.tag = x['tag']
                        s.save()
                    context['status'] = True
                if 'alertmsg' in request.POST:
                    if 'edit' in request.POST:
                        obj = Alertasproyecto.objects.get(
                            pk=request.POST.get('edit'))
                        form = AlertasproyectoForm(request.POST, instance=obj)
                    else:
                        form = AlertasproyectoForm(request.POST)
                    if form.is_valid():
                        add = form.save(commit=False)
                        add.empdni_id = request.user.get_profile().empdni_id
                        add.charge_id = request.user.get_profile(
                                            ).empdni.charge_id
                        add.sector_id = kwargs['sec']
                        add.save()
                        context['status'] = True
                    else:
                        context['status'] = False
                if 'readerPrices' in request.POST:
                    period = search.searchPeriodProject(code=kwargs['pro'])
                    filename = '%s/storage/projects/%s/%s/prices.xlsx' % (
                        globalVariable.relative_path, period, kwargs['pro'])
                    # print filename
                    if os.path.exists(filename):
                        sess = 'PRICES%s' % (kwargs['pro'])
                        if sess in request.session:
                            del request.session[sess]
                        if not sess in request.session:
                            request.session[sess] = list()
                            request.session[sess] = uploadFiles.readQuotation(
                                filename)
                        context['status'] = True
                    else:
                        context['status'] = False
                if 'diffmodifysec' in request.POST:
                    lp = [x.materiales_id for x in MetProject.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id=kwargs['sub'] if kwargs['sub'] != unicode(None) else None, sector_id=kwargs['sec'])]
                    lded = UpdateMetProject.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id=kwargs['sub'] if kwargs['sub'] != unicode(None) else None, sector_id=kwargs['sec']).exclude(materials_id__in=lp)
                    if lded.count():
                        context['deductive'] = [
                            {
                                'materials': x.materials_id,
                                'name': x.materials.matnom,
                                'meter': x.materials.matmed,
                                'unit': x.materials.unidad.uninom,
                                'brand': x.brand.brand,
                                'model': x.model.model,
                                'quantity': x.quantity,
                                'purchase': x.price,
                                'sales': float(x.sales),
                                'amount': (x.quantity * x.price)
                            }
                            for x in lded
                        ]
                        mm = [x.materials_id for x in UpdateMetProject.objects.filter(proyecto_id=kwargs['pro'], subproyecto_id=kwargs['sub'] if kwargs['sub'] != unicode(None) else None, sector_id=kwargs['sec'])]
                        ldel = MetProject.objects.filter(
                            proyecto_id=kwargs['pro']).exclude(
                            materiales_id__in=mm)
                        if ldel.count():
                            context['mdelete'] = [
                                {
                                    'materials': x.materiales_id,
                                    'name': x.materials.matnom,
                                    'meter': x.materials.matmed,
                                    'unit': x.materials.unidad.uninom,
                                    'brand': x.brand.brand,
                                    'model': x.model.model,
                                    'quantity': x.cantidad,
                                    'purchase': x.precio,
                                    'sales': float(x.sales),
                                    'amount': (x.cantidad * x.precio)
                                }
                                for x in ldel
                            ]
                        context['status'] = True
                    else:
                        context['status'] = False
                if 'savepreorders' in request.POST:
                    form = PreOrdersForm(request.POST)
                    if form.is_valid():
                        add = form.save(commit=False)
                        key = genkeys.generatePreOrdersId()
                        add.preorder_id = key
                        add.project_id = kwargs['pro']
                        add.subproject_id = kwargs[
                            'sub'] if unicode(kwargs['sub']) != 'None' else ''
                        add.sector_id = kwargs['sec']
                        add.performed_id = request.user.get_profile().empdni_id
                        add.save()
                        # details
                        details = json.loads(request.POST.get('order'))
                        for x in details:
                            det = DetailsPreOrders()
                            det.preorder_id = key
                            det.materials_id = x['materials']
                            det.brand_id = x['brand']
                            det.model_id = x['model']
                            det.quantity = x['quantity']
                            det.orders = x['quantity']
                            det.save()
                        context['code'] = key
                        context['status'] = True
                    else:
                        context['status'] = False
                        context['raise'] = 'Forms invalid.'
                if 'savewithoutprice' in request.POST:
                    lst = json.loads(request.POST['list'])
                    for x in lst:
                        prices = UpdateMetProject.objects.get(
                            proyecto_id=kwargs['pro'],
                            subproyecto_id=kwargs['sub']
                            if kwargs['sub'] != unicode(None)
                            else None, sector_id=kwargs['sec'],
                            materials_id=x['materials'],
                            brand_id=x['brand'],
                            model_id=x['model'])
                        prices.price = x['purchase']
                        prices.sales = x['sales']
                        prices.save()
                    context['status'] = True
                if 'readNOrders' in request.POST:
                    path = '/storage/Temp/'
                    name = uploadFiles.upload(path, request.FILES['fOrders'])
                    wb = load_workbook(filename=name, read_only=True)
                    ws = wb['ORDEN']
                    nrow = ws.max_row
                    # ncol = ws.max_column
                    accessory = list()
                    for y in range(7, nrow):
                        cell = ws.cell(row=y, column=1)
                        # if cell.value is not None:
                        a = str(cell.value)
                        if len(a) > 13 and len(a) <= 15:
                            # print type(cell.value)
                            # print cell.value
                            # Save List Accesories
                            accessory.append({
                                'materials': a,
                                'quantity': float(ws.cell(row=y, column=4).value),
                                'observation': ws.cell(row=y, column=5).value if not ws.cell(row=y, column=5).value is None else ''
                            })
                            print accessory
                        if a == 'NIPLES':
                            # Save list nipples
                            for yn in range(y+1, nrow+1):
                                try:
                                    c = str(ws.cell(row=yn, column=1).value)
                                except Exception, e:
                                    break
                                # print accessory
                                if len(c) > 13 and len(c) <= 15:
                                    mc = c
                                    accessory.append({
                                        'materials': mc,
                                        'quantity': 0,
                                        'nipples': []})
                                    continue
                                # tmp = accessory
                                # print len(accessory)
                                np = filter(lambda n: n['materials'] == mc, accessory)
                                # np = next(n for n in accessory if n['materials'] == mc)
                                print np
                                if np:
                                    for d in accessory:
                                        if d['materials'] == mc:
                                            if ws.cell(row=yn, column=2).value == None:
                                                continue
                                            print ws.cell(row=yn, column=2).value
                                            print ws.cell(row=yn, column=3).value
                                            print ws.cell(row=yn, column=5).value
                                            typen = str(ws.cell(row=yn, column=2).value).split('=')[-1]
                                            me = float(str(ws.cell(row=yn, column=3).value))
                                            qu = float(str(ws.cell(row=yn, column=5).value))
                                            d['nipples'].append({
                                                'materials': mc,
                                                'type': typen.strip(),
                                                'measure': me,
                                                'quantity': qu,
                                                'observation': ws.cell(row=yn, column=6).value if not ws.cell(row=yn, column=6).value is None else '',
                                                })
                                            d['quantity']+=((me*qu)/100)
                            break
                        # else:
                        #     continue
                    # Compare list with order
                    ntexists = list()
                    orders = list()
                    for ac in accessory:
                        try:
                            lo = MetProject.objects.get(
                                    proyecto_id=kwargs['pro'],
                                    subproyecto_id=kwargs['sub'] if kwargs[
                                        'sub'] != unicode(None) else None,
                                    sector_id=kwargs['sec'],
                                    materiales_id=ac['materials'])
                            ac['name'] = lo.materiales.matnom
                            ac['measure'] = lo.materiales.matmed
                            ac['unit'] = lo.materiales.unidad.uninom
                            if ac['quantity'] <= lo.quantityorder:
                                orders.append({
                                    'materials': lo.materiales_id,
                                    'quantity': ac['quantity'],
                                    'observation': ac['observation'] if 'observation' in ac else '',
                                    'brand': lo.brand_id,
                                    'model': lo.model_id})
                            else:
                                ntexists.append({
                                    'materials': lo.materiales_id,
                                    'quantity': (ac['quantity']-lo.quantityorder),
                                    'brand': lo.brand_id,
                                    'model': lo.model_id,
                                    'observation': ac['observation'] if 'observation' in ac else ''
                                    })
                        except MetProject.DoesNotExist:
                            ex = list(filter(lambda x: x['materials'] == ac['materials'], ntexists))
                            if ex:
                                ex = ex[0]
                                for i in ntexists:
                                    if i['materials'] == ex['materials']:
                                        i['quantity'] += ac['quantity']
                            else:
                                ntexists.append({
                                    'materials': ac['materials'],
                                    'quantity': ac['quantity'],
                                    'brand': 'BR000',
                                    'model': 'MO000',
                                    'observation': ac['observation'] if 'observation' in ac else ''})
                            try:
                                ma = Materiale.objects.get(materiales_id=ac['materials'])
                                ac['name'] = ma.matnom
                                ac['measure'] = ma.matmed
                                ac['unit'] = ma.unidad.uninom
                            except Materiale.DoesNotExist:
                                pass
                    if len(ntexists):
                        # clear update meter project
                        UpdateMetProject.objects.filter(
                            proyecto_id=kwargs['pro'],
                            subproyecto_id=kwargs['sub'] if kwargs[
                                'sub'] != unicode(None) else None,
                            sector_id=kwargs['sec']).delete()
                        # add list to update list
                        for x in MetProject.objects.filter(
                            proyecto_id=kwargs['pro'],
                            subproyecto_id=kwargs['sub']
                            if kwargs['sub'] != unicode(None)
                            else None, sector_id=kwargs['sec']).order_by(
                                'materiales__matnom'):
                            obj = UpdateMetProject()
                            obj.proyecto_id = kwargs['pro']
                            obj.subproyecto_id = kwargs['sub'] if kwargs[
                                'sub'] != unicode(None) else ''
                            obj.sector_id = kwargs['sec']
                            obj.materials_id = x.materiales_id
                            obj.brand_id = x.brand_id
                            obj.model_id = x.model_id
                            obj.quantity = x.cantidad
                            obj.price = x.precio
                            obj.sales = x.sales
                            obj.comment = x.comment
                            obj.quantityorders = x.quantityorder
                            obj.tag = x.tag
                            obj.save()
                        # add update list materials not exists
                        for x in ntexists:
                            ppurchase = 0
                            psales = 0
                            m = MetProject.objects.filter(
                                materiales_id=x['materials']
                                ).order_by('-proyecto__registrado')
                            if m:
                                m = m[0]
                                ppurchase = m.precio
                                psales = m.sales
                            try:
                                u = UpdateMetProject.objects.get(
                                        proyecto_id=kwargs['pro'],
                                        subproyecto_id=kwargs['sub'] if kwargs[
                                            'sub'] != unicode(None) else None,
                                        sector_id=kwargs['sec'],
                                        materials_id=x['materials'])
                                u.quantity += x['quantity']
                                u.quantityorders += x['quantity']
                                u.save()
                            except UpdateMetProject.DoesNotExist:
                                obj = UpdateMetProject()
                                obj.proyecto_id = kwargs['pro']
                                obj.subproyecto_id = kwargs['sub'] if kwargs[
                                    'sub'] != unicode(None) else ''
                                obj.sector_id = kwargs['sec']
                                obj.materials_id = x['materials']
                                obj.brand_id = 'BR000'
                                obj.model_id = 'MO000'
                                obj.quantity = x['quantity']
                                obj.price = ppurchase
                                obj.sales = psales
                                obj.comment = x['observation']
                                obj.quantityorders = x['quantity']
                                obj.tag = '0'
                                obj.save()
                        context['result'] = 'modify'
                    else:
                        context['requeriment'] = accessory
                        context['issue'] = str(ws.cell(row=5, column=2).value)
                        context['traslate'] = ws.cell(row=4, column=4).value
                        context['filename'] = name
                        # uploadFiles.deleteFile(name)
                        context['result'] = 'showTable'
                    context['status'] = True
                if 'processRequeriment' in request.POST:
                    # create order to storage
                    # save Nipples
                    data = json.loads(request.POST['data'])
                    print 'requeriment'
                    accessory = data['requeriment']
                    # print accessory
                    for ac in accessory:
                        if 'nipples' in ac:
                            for x in ac['nipples']:
                                np = Nipple.objects.filter(
                                        Q(proyecto_id=kwargs['pro']),
                                        Q(subproyecto_id=kwargs['sub'] if kwargs['sub'] != unicode(None) else None),
                                        Q(sector_id=kwargs['sec']),
                                        Q(materiales_id=ac['materials']),
                                        Q(metrado=x['measure']),
                                        Q(tipo=x['type']),
                                        Q(flag=True),)
                                if np:
                                    n = 0
                                    for i in np:
                                        if i.id > n:
                                            n = i
                                    np = n
                                    np.cantidad += x['quantity']
                                    if np.cantshop == 0:
                                        np.tag = '2'
                                    else:
                                        np.tag = '1'
                                    np.save()
                                else:
                                    np = Nipple()
                                    np.proyecto_id = kwargs['pro']
                                    np.subproyecto_id = kwargs['sub'] if kwargs['sub'] != unicode(None) else None
                                    np.sector_id = kwargs['sec']
                                    np.materiales_id = ac['materials']
                                    np.metrado = x['measure']
                                    np.tipo = str(x['type']).strip()
                                    np.cantidad = x['quantity']
                                    np.cantshop = 0
                                    np.comment = x['observation']
                                    np.tag = '2'
                                    np.save()
                    # generate orders Storage
                    gkey = genkeys.GenerateIdOrders()
                    pe = Pedido()
                    pe.pedido_id = gkey
                    pe.proyecto_id = kwargs['pro']
                    pe.subproyecto_id = kwargs['sub'] if kwargs[
                        'sub'] != unicode(None) else None
                    pe.sector_id = kwargs['sec']
                    pe.almacen_id = 'AL01'
                    pe.asunto = data['issue']
                    pe.empdni_id = request.user.get_profile().empdni_id
                    pe.traslado = data['traslate']
                    pe.obser = kwargs['pro']
                    # pe.orderfile = request.FILES['fOrders']
                    pe.status = 'PE'
                    pe.save()
                    ext = data['filename'].split('.')[-1]
                    p = ''
                    try:
                        p = Proyecto.objects.get(proyecto_id=kwargs['pro'])
                        p = p.registrado.strftime('%Y')
                    except Proyecto.DoesNotExist, e:
                        p = globalVariable.get_year
                    path = '%s/storage/projects/%s/%s/' % (
                        globalVariable.relative_path,
                        p,
                        kwargs['pro'])
                    print 'origin', data['filename']
                    print 'move', path
                    if not os.path.exists(path):
                        os.makedirs(path, 0777)
                        os.chmod(path, 0777)
                    path = '%s%s%s' % (path, gkey, ext)
                    shutil.move(data['filename'], path)
                    # details orders
                    for d in accessory:
                        dlm = MetProject.objects.get(
                                proyecto_id=kwargs['pro'],
                                subproyecto_id=kwargs['sub']
                                if kwargs['sub'] != unicode(None)
                                else None, sector_id=kwargs['sec'],
                                materiales_id=d['materials'])
                        dlm.quantityorder -= d['quantity']
                        if dlm.quantityorder <= 0:
                            dlm.quantityorder = 0
                            dlm.tag = '2'
                        dlm.save()
                        de = Detpedido()
                        de.pedido_id = gkey
                        de.materiales_id = d['materials']
                        de.brand_id = 'BR000'
                        de.model_id = 'MO000'
                        de.cantidad = d['quantity']
                        de.cantshop = d['quantity']
                        de.cantguide = 0
                        de.tag = '0'
                        de.comment = d['observation'] if 'observation' in d else ''
                        de.flag = True
                        de.save()
                        if 'nipples' in d:
                            for dn in d['nipples']:
                                n = Niple()
                                n.pedido_id = gkey
                                n.proyecto_id = kwargs['pro']
                                n.subproyecto_id = kwargs['sub'] if kwargs[
                                    'sub'] != unicode(None) else None
                                n.sector_id = kwargs['sec']
                                n.empdni_id = request.user.get_profile().empdni_id
                                n.materiales_id = d['materials']
                                n.cantidad = dn['quantity']
                                n.metrado = dn['measure']
                                n.cantshop = dn['quantity']
                                n.cantguide = 0
                                n.tipo = str(dn['type']).strip()
                                n.flag = True
                                n.tag = '0'
                                n.comment = dn['observation'] if 'observation' in dn else ''
                                n.save()
                    context['result'] = 'orders'
                    context['orders'] = gkey
                    context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = e.__str__()
                context['status'] = False
            return self.render_to_json_response(context)


class ListGuideByProject(JSONResponseMixin, TemplateView):

    template_name = 'sales/listguide.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        try:
            if kwargs['sec'] == unicode(None):
                guide = GuiaRemision.objects.filter(
                    pedido__proyecto_id=kwargs['pro']).order_by('registrado')
            else:
                guide = GuiaRemision.objects.filter(
                    pedido__proyecto_id=kwargs['pro'],
                    pedido__sector_id=kwargs['sec']).order_by('-registrado')
                context['sec'] = Sectore.objects.get(sector_id=kwargs['sec'])
            if guide:
                context['guides'] = [
                    {
                      'guide': c.guia_id,
                      'register': c.registrado,
                      'traslate': c.traslado,
                      'customer': c.ruccliente_id,
                      'companyname': c.ruccliente.razonsocial,
                      'currency': c.pedido.proyecto.currency.moneda,
                      'status': c.status,
                      'orders': c.pedido_id,
                      'details': [
                        {
                            'materials': x.materiales_id,
                            'name': x.materiales.matnom,
                            'meter': x.materiales.matmed,
                            'unit': x.materiales.unidad.uninom,
                            'quantity': x.cantguide,
                            'nipple': [
                                {
                                    'type': [ globalVariable.tipo_nipples[k] for k in globalVariable.tipo_nipples if n.tipo == k],
                                    'meter': n.metrado,
                                    'quantity': n.cantguide
                                }
                                for n in NipleGuiaRemision.objects.filter(guia_id=c.guia_id, materiales_id=x.materiales_id)
                            ] if x.materiales_id[0:3] == '115' else []
                        }
                        for x in DetGuiaRemision.objects.filter(guia_id=c.guia_id)
                      ]
                    }
                    for c in guide
                ]
                context['pro'] = Proyecto.objects.get(pk=kwargs['pro'])
                context['dates'] = [guide[0].registrado, guide.latest('registrado').registrado]
            return render_to_response(self.template_name, context, context_instance=RequestContext(request))
        except TemplateDoesNotExist, e:
            raise Http404(e)


class ListOrdersByProject(JSONResponseMixin, TemplateView):

    template_name = 'sales/listorders.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        try:
            if kwargs['sec'] == unicode(None):
                orders = Pedido.objects.filter(
                    proyecto_id=kwargs['pro']).order_by('registrado')
            else:
                orders = Pedido.objects.filter(
                    proyecto_id=kwargs['pro'],
                    sector_id=kwargs['sec']).order_by('-registrado')
                context['sec'] = Sectore.objects.get(sector_id=kwargs['sec'])
            if orders:
                context['orders'] = [{
                    'orders': o.pedido_id,
                    'register': o.registrado,
                    'transfer': o.traslado,
                    'perform': o.empdni.name_complete,
                    'status': o.status,
                    'details': [{
                        'materials': x.materiales_id,
                        'name': x.materiales.matnom,
                        'meter': x.materiales.matmed,
                        'unit': x.materiales.unidad.uninom,
                        'brand': x.brand.brand,
                        'model': x.model.model,
                        'quantity': x.cantidad,
                        'nipple': [{
                            'type': [globalVariable.tipo_nipples[k] for k in globalVariable.tipo_nipples if n.tipo == k],
                            'meter': n.metrado,
                            'quantity': n.cantidad}
                            for n in Niple.objects.filter(
                                pedido_id=o.pedido_id,
                                materiales_id=x.materiales_id)
                            ] if x.materiales_id[0:3] == '115' else []}
                        for x in Detpedido.objects.filter(
                            pedido_id=o.pedido_id)]}
                    for o in orders]
                context['pro'] = Proyecto.objects.get(pk=kwargs['pro'])
                context['dates'] = [
                    orders[0].registrado,
                    orders.latest('registrado').registrado]
            return render_to_response(
                self.template_name,
                context,
                context_instance=RequestContext(request))
        except TemplateDoesNotExist, e:
            raise Http404(e)


class ServicesProjectView(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        try:
            context['pro'] = Proyecto.objects.get(
                            proyecto_id=kwargs['pro'])
            svc = ServiceOrder.objects.filter(project_id=kwargs['pro'])
            dsvc = DetailsServiceOrder.objects.filter(
                serviceorder_id__in=[x.serviceorder_id for x in svc])
            lst = list()
            for x in svc.order_by('-register'):
                am = dsvc.filter(
                        serviceorder_id=x.serviceorder_id
                        ).aggregate(amount=Sum(
                            'quantity', field='quantity*price'))['amount']
                cs = ''
                if x.currency_id == 'CU02':
                    am = (am/x.project.exchange)
                    cs = ' - USD'
                conf = Configuracion.objects.get(
                        periodo=x.register.strftime('%Y'))
                dst = (((x.dsct*am)/100)+am)
                if x.sigv:
                    am = (((conf.igv*dst)/100)+dst)
                else:
                    am = dst
                lst.append({
                    'id': x.serviceorder_id,
                    'supplier': x.supplier.razonsocial,
                    'register': x.register.strftime('%d-%m-%Y'),
                    'symbol': '%s%s' % (x.currency.simbolo, cs),
                    'amount': am})
                am = None
            context['services'] = lst
            context['total'] = sum([x['amount'] for x in lst])
            if context['total'] is None:
                context['total'] = 0
            context['diff'] = (
                float(context['pro'].aservices) - context['total'])
            return render(request, 'sales/servicesproject.html', context)
        except TemplateDoesNotExist, e:
            raise Http404(e)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        # context = dict()
        if 'saservices' in request.POST:
            p = Proyecto.objects.get(proyecto_id=kwargs['pro'])
            p.aservices = request.POST['aservices']
            p.save()
            return redirect('servicesp_view', pro=kwargs['pro'])
