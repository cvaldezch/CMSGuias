#-*- Encoding: utf-8 -*-
from django.conf import settings
import datetime
import os
import urllib
from openpyxl import load_workbook


def upload(absolutePath,archive,options={}):
    #defaults= {"date": False,"time": False}
    try:
        # path absolute
        path = '%s%s'%(settings.MEDIA_ROOT, absolutePath)
        # verify path exists if not exists path makers the folders
        if not os.path.exists(path):
            os.makedirs(path, 0774)
            os.chmod(path, 0774)
        # verify aggregate options to name of file
        name = None
        if len(options) > 0:
            ext = archive.name.split('.')
            ext = ext[ext.__len__() - 1]
            ext = ext.lower()
            name = '%s.%s'%(options['name'], ext)
        else:
          name = archive.name
        filename = '%s%s'%(path, name)
        # recover full address of filename
        dirfilename = open(filename, 'wb+')
        # walk all file and save im new address
        for bit in archive.chunks():
            dirfilename.write(bit)
        # close file
        dirfilename.close()
        # changes file permissions
        os.chmod(filename, 0777)
        return filename
    except Exception, e:
        return False

def removeTmp(absolutePath):
    try:
        os.remove(absolutePath)
    except Exception, e:
        raise e

def deleteFile(uriPath, partial=False):
    try:
        path = None
        # path absolute
        if partial:
            path = '%s%s'%(settings.MEDIA_ROOT, uriPath)
        else:
            path = uriPath
        if path is not None:
            print os.remove(path)
        else:
            return False
    except Exception, e:
        print e
        return False

def fileExists(paths, partial=False):
    try:
        if partial:
            # path absolute
            path = '%s%s'%(settings.MEDIA_ROOT, paths)
        else:
            path = paths
        return os.path.lexists(path)
    except Exception, e:
        print e
        return False

def descompressRAR(filename, path_to_extract):
    try:
        if filename == '':
            return 'File name is nothing.'
        if path_to_extract == '':
            return 'Path to extract nothing.'
        if filename != '' and path_to_extract != '':
            path_to_extract = '%s%s'%(settings.MEDIA_ROOT, path_to_extract)
            cmd = 'unrar x -y %s %s'%(filename, path_to_extract)
            os.system(cmd)
            return 'success'
    except Exception, e:
        return e.__str__()

def listDir(path):
    r = ['<ul class="jqueryFileTree" style="display: none;">']
    try:
        r = ['<ul class="jqueryFileTree" style="display: none;">']
        #d = urllib.unquote(path) #request.POST.get('dir','c:\\temp'
        dirstart = str(settings.MEDIA_ROOT).split('/')[1]
        if not path.startswith('/%s/'%dirstart):
            path = '%s%s'%(settings.MEDIA_ROOT, path)
        path = urllib.unquote(path)
        for f in os.listdir(path):
            ff = os.path.join(path,f)
            if os.path.isdir(ff):
                r.append('<li class="directory collapsed"><a href="#" rel="%s/">%s</a></li>' % (ff,f))
            else:
                e = os.path.splitext(f)[1][1:] # get .ext and remove dot
                e = e.lower()
                media = ff.split('/media/')
                print media
                r.append('<li class="file ext_%s"><a href="#" rel="/media/%s">%s</a></li>' % (e,media[1],f))
        r.append('</ul>')
    except Exception, e:
        r.append('Could not load directory.')

    r.append('</ul>')
    return r

def readQuotation(filename):
    workbook = load_workbook(filename=filename, read_only=False, use_iterators=True)
    # first recover the prices for each item
    head = list()
    counter = 0
    prices = dict()
    # get sheet price
    sheet = workbook.get_sheet_by_name('PRECIOS')
    for row in sheet.iter_rows():
        if counter == 1620:
            break
        if len(unicode(row[3].value)) == 15:
            price = 0
            if str(row[7].value).replace('.','',1).isdigit():
                price = float(row[7].value)
            # print 'purchase ', price
            sale = 0
            # print 'sale ', sale
            if str(row[8].value).replace('.','',1).isdigit():
                sale = float(row[8].value)
            prices[str(row[3].value).strip()] = {
                                                    'purchase': price,
                                                    'sale': sale
                                                }
        counter += 1

    # print prices
    # quantity for sector or sheet
    sectors = dict()
    book = workbook.get_sheet_names()
    for n in book:
        if n == 'PRECIOS':
            continue
        page = workbook.get_sheet_by_name(n)
        counter = 0
        for cell in page.iter_rows():
            if str(cell[0].value).upper().strip() != 'ACTIVE' and counter == 0:
                break
            else:
                if counter == 1:
                    name_sector = str(cell[0].value).split(',')
                    for s in name_sector:
                        sectors[s] = list()
                # print cell[6].value
                if len(str(cell[0].value).strip()) == 15:
                    if not str(cell[6].value).isdigit():
                        continue
                    quantity = 0
                    if cell[6].value:
                        quantity = float(cell[6].value)
                    for s in name_sector:
                        purchase = 0
                        sale = 0
                        #print prices['220018030014001']
                        if str(cell[0].value) in prices:
                            purchase = prices[str(cell[0].value).strip()]['purchase']
                            sale = prices[str(cell[0].value).strip()]['sale']
                        sectors[s].append(
                            {
                                'materials': str(cell[0].value),
                                'quantity': quantity,
                                'purchase': purchase,
                                'sale': sale
                            }
                        )
                counter += 1
        if sectors:
            head.append(sectors)
            sectors = dict()
    return head